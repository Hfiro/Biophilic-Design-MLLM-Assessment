"""
Author: Hoda Alem
Date: November 2025
Context: Part of PhD research for a PhD in Environmental Design and Planning
         at Myers-Lawson School of Construction, Virginia Tech, Blacksburg, VA.
Purpose: Plot aggregate performance over prompt iterations: categorical accuracy
         (0â€“1â€“2 exact matches) and scalar error (1â€“10 MSE) against Expert.
Features:
- Reads Expert truths and per-space result workbooks.
- Computes per-iteration accuracy (0â€“1â€“2 exact match) and MSE (1â€“10).
- Draws two stacked subplots with per-attribute grouped bars and an average
  line across attributes; includes expert baseline lines.
- Saves a single figure for quick, high-level trend comparison.
"""

import argparse
import math
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple


def _normalize_tokens(s: str) -> List[str]:
    import re
    s = (s or "").lower()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return [t for t in s.split() if t]


def _norm_key(s: str) -> str:
    return " ".join(_normalize_tokens(s))


def _match_space_dict(mapping: Dict[str, Dict[str, float]], space: str) -> Dict[str, float]:
    if space in mapping:
        return mapping[space]
    norm = _norm_key(space)
    for name, values in mapping.items():
        if _norm_key(name) == norm:
            return values
    return {}


def _match_attr_value(mapping: Dict[str, float], attribute: str) -> Optional[float]:
    if attribute in mapping:
        return mapping[attribute]
    norm = _norm_key(attribute)
    for name, value in mapping.items():
        if _norm_key(name) == norm:
            return value
    return None


def _resolve_sheet_name(wb, desired: str) -> Optional[str]:
    # Exact match
    if desired in wb.sheetnames:
        return desired
    tokens = set(_normalize_tokens(desired))
    cands = [s for s in wb.sheetnames if s.lower() != 'index']
    best_subset = [s for s in cands if tokens and tokens.issubset(set(_normalize_tokens(s)))]
    if best_subset:
        return sorted(best_subset, key=lambda x: (len(_normalize_tokens(x)), x))[0]
    scored = sorted(((len(tokens & set(_normalize_tokens(s))), s) for s in cands), reverse=True)
    if scored and scored[0][0] > 0:
        return scored[0][1]
    return None


def _load_wb(path: Path):
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:
        raise SystemExit(
            "Missing dependency: openpyxl. Install with `pip install openpyxl` or `pip install -r requirements.txt`."
        ) from e
    return load_workbook(filename=str(path), read_only=True, data_only=True)


def _collect_data_rows(ws) -> List[Dict[str, Any]]:
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers: Dict[str, int] = {}
    for i, v in enumerate(header):
        if isinstance(v, str):
            headers[v.strip()] = i
    data: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        rec: Dict[str, Any] = {}
        for k, idx in headers.items():
            rec[k] = row[idx] if idx < len(row) else None
        if isinstance(rec.get("run"), (int, float)):
            data.append(rec)
    return data


def _read_expert_truths(path: Path, sheet_name: str = "Expert") -> Dict[str, Dict[str, Dict[str, float]]]:
    """Return mapping: space -> attribute -> {'0-2': v, '1-10': v}.

    Layout:
    - Row 1: space names repeated for two adjacent columns per space
    - Row 2: rating tag; we ignore and take fixed positions (col c = 0-2, col c+1 = 1-10)
    - Col 1: index (ignored)
    - Col 2: attribute names
    - Data starts row 3
    """
    wb = _load_wb(path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {path}")
    ws = wb[sheet_name]

    header1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header2 = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    attr_col = 2

    def _label_is_02(label: Any) -> bool:
        s = str(label or '').lower()
        return ('0' in s and '2' in s) or ('0, 1, 2' in s)

    # Primary detection: use row 2 labels ("0, 1, 2" and "1 - 10") paired with space names in row 1.
    space_cols: Dict[str, int] = {}
    for c in range(3, ws.max_column + 1):  # 1-based inclusive
        name = str(header1[c - 1] or '').strip()
        lab = str(header2[c - 1] or '').strip()
        if name and _label_is_02(lab):
            space_cols.setdefault(name, c)

    # Fallback: consecutive duplicate names in row 1
    if not space_cols:
        for c in range(3, ws.max_column):
            name1 = str(header1[c - 1] or '').strip()
            name2 = str(header1[c] or '').strip()
            if name1 and name2 and name1 == name2:
                space_cols.setdefault(name1, c)

    # Second fallback: assume pairs start at column 3 and repeat every two columns
    if not space_cols:
        c = 3
        while c <= ws.max_column:
            space = str(header1[c - 1] or '').strip()
            if space:
                space_cols.setdefault(space, c)
            c += 2

    out: Dict[str, Dict[str, Dict[str, float]]] = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row is None:
            continue
        attr_raw = row[attr_col - 1] if attr_col - 1 < len(row) else None
        if not attr_raw:
            continue
        attribute = str(attr_raw).strip()
        for space, c02 in space_cols.items():
            # c02 points to the 0-2 column in the pair; next column is 1-10
            v02 = row[c02 - 1] if c02 - 1 < len(row) else None
            c10 = c02 + 1
            v10 = row[c10 - 1] if c10 - 1 < len(row) else None
            try:
                if v02 is None or (isinstance(v02, float) and math.isnan(v02)):
                    val = None
                elif isinstance(v02, (int, float)):
                    val = float(v02)
                else:
                    s = str(v02).strip()
                    val = float(s) if s else None
            except Exception:
                val = None
            try:
                if v10 is None or (isinstance(v10, float) and math.isnan(v10)):
                    val10 = None
                elif isinstance(v10, (int, float)):
                    val10 = float(v10)
                else:
                    s10 = str(v10).strip()
                    val10 = float(s10) if s10 else None
            except Exception:
                val10 = None

            if val is not None or val10 is not None:
                out.setdefault(space, {}).setdefault(attribute, {})
                if val is not None:
                    out[space][attribute]['0-2'] = val
                if val10 is not None:
                    out[space][attribute]['1-10'] = val10
    return out


def _get_attr_scale_value(attr_map: Dict[str, Dict[str, float]], attribute: str, scale_tag: str) -> Optional[float]:
    """Fetch scale value for an attribute using normalized name matching."""
    # Exact
    entry = attr_map.get(attribute)
    if isinstance(entry, dict) and scale_tag in entry:
        return entry.get(scale_tag)
    # Normalized match
    target = _norm_key(attribute)
    for name, vals in attr_map.items():
        if _norm_key(name) == target and scale_tag in vals:
            return vals.get(scale_tag)
    return None

def _attributes_from_expert(path: Path, sheet_name: str = "Expert") -> List[str]:
    wb = _load_wb(path)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    attrs: List[str] = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        v = row[1] if len(row) > 1 else None
        if v:
            attrs.append(str(v).strip())
    return [a for a in attrs if a]

def _available_iterations(result_paths: List[Path]) -> List[int]:
    iters: set[int] = set()
    for p in result_paths:
        wb = _load_wb(p)
        for sheet in wb.sheetnames:
            if sheet in {"meta", "prompts"}:
                continue
            ws = wb[sheet]
            rows = _collect_data_rows(ws)
            for rec in rows:
                try:
                    it = int(rec.get('iteration'))
                    iters.add(it)
                except Exception:
                    continue
    return sorted(iters)


def _space_from_results(path: Path) -> Optional[str]:
    wb = _load_wb(path)
    if 'meta' not in wb.sheetnames:
        return None
    ws = wb['meta']
    space = None
    for r in ws.iter_rows(min_row=1, values_only=True):
        if r and str(r[0]).strip().lower() == 'space':
            space = r[1]
            break
    return str(space).strip() if space else None


def compute_match_percentages(
    result_paths: List[Path],
    expert02: Dict[str, Dict[str, float]],
    attributes: List[str],
    iterations: List[int] = [1, 2, 3, 4],
) -> Dict[int, Dict[str, float]]:
    """Return mapping: iteration -> attribute -> percent matches (0..100).

    Aggregates across all spaces and runs for each attribute and iteration.
    """
    # Build map space -> path
    per_space_paths: Dict[str, Path] = {}
    for p in result_paths:
        s = _space_from_results(p)
        if s:
            per_space_paths[s] = p

    # For each iteration and attribute, count matches across spaces and runs
    out: Dict[int, Dict[str, float]] = {it: {} for it in iterations}
    for attr in attributes:
        # For each space available in both expert and results
        for it in iterations:
            matched = 0
            total = 0
            for space, path in per_space_paths.items():
                exp_attr_map = expert02.get(space)
                if exp_attr_map is None:
                    exp_attr_map = _match_space_dict(expert02, space)
                if not exp_attr_map:
                    continue
                exp_val = _match_attr_value(exp_attr_map, attr)
                if exp_val is None:
                    continue
                wb = _load_wb(path)
                sheet_name = _resolve_sheet_name(wb, attr)
                if not sheet_name:
                    continue
                ws = wb[sheet_name]
                rows = _collect_data_rows(ws)
                for rec in rows:
                    try:
                        rit = int(rec.get('iteration'))
                    except Exception:
                        continue
                    if rit != it:
                        continue
                    rv = rec.get('rating_0_2')
                    if isinstance(rv, (int, float)):
                        total += 1
                        if float(rv) == float(exp_val):
                            matched += 1
            out[it][attr] = (100.0 * matched / total) if total > 0 else 0.0
    return out


def compute_mse(
    result_paths: List[Path],
    expert10: Dict[str, Dict[str, float]],
    attributes: List[str],
    iterations: List[int] = [1, 2, 3, 4],
) -> Dict[int, Dict[str, float]]:
    """Return mapping iteration -> attribute -> MSE of 1â€“10 ratings vs expert across spaces and runs."""
    # Map space->path
    per_space_paths: Dict[str, Path] = {}
    for p in result_paths:
        s = _space_from_results(p)
        if s:
            per_space_paths[s] = p

    out: Dict[int, Dict[str, float]] = {it: {} for it in iterations}
    for attr in attributes:
        for it in iterations:
            sse = 0.0
            n = 0
            for space, path in per_space_paths.items():
                exp_attr_map = expert10.get(space)
                if exp_attr_map is None:
                    exp_attr_map = _match_space_dict(expert10, space)
                if not exp_attr_map:
                    continue
                exp_val = _match_attr_value(exp_attr_map, attr)
                if exp_val is None:
                    continue
                wb = _load_wb(path)
                sheet_name = _resolve_sheet_name(wb, attr)
                if not sheet_name:
                    continue
                ws = wb[sheet_name]
                rows = _collect_data_rows(ws)
                for rec in rows:
                    try:
                        rit = int(rec.get('iteration'))
                    except Exception:
                        continue
                    if rit != it:
                        continue
                    sv = rec.get('score_1_10')
                    if isinstance(sv, (int, float)):
                        diff = float(sv) - float(exp_val)
                        sse += diff * diff
                        n += 1
            out[it][attr] = (sse / n) if n > 0 else 0.0
    return out


def plot_dual_progression(percentages: Dict[int, Dict[str, float]], mses: Dict[int, Dict[str, float]], attributes: List[str], out: Path) -> None:
    """Plot grouped bars per attribute over iterations and overlay average lines."""
    try:
        import matplotlib.pyplot as plt  # type: ignore
        import numpy as np  # type: ignore
    except Exception as e:
        raise SystemExit(
            "Missing dependency: matplotlib (and numpy). Install with `pip install matplotlib numpy`."
        ) from e

    iters = sorted(percentages.keys())
    n_attr = len(attributes)
    x = np.arange(len(iters))
    width = min(0.16, 0.8 / max(1, n_attr))

    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(11, 10), sharex=True)

    # Top: Accuracy
    for i, attr in enumerate(attributes):
        vals = [percentages[it].get(attr, 0.0) for it in iters]
        ax1.bar(x + (i - (n_attr-1)/2)*width, vals, width=width, label=attr)
    # Piecewise linear average across attributes (centered at iteration positions)
    avg_acc = [float(np.mean([percentages[it].get(attr, 0.0) for attr in attributes])) for it in iters]
    ax1.plot(
        x,
        avg_acc,
        '-o',
        color='tab:cyan',
        linewidth=3.0,
        markersize=8,
        markerfacecolor='tab:cyan',
        markeredgecolor='tab:cyan',
        label='Avg across attributes',
    )
    # Optional expert baseline for 0-2 accuracy will be added later if provided
    ax1.set_ylabel('Accuracy (%, 0,1,2 match)')
    ax1.set_ylim(0, 106)
    ax1.grid(axis='y', alpha=0.3)
    ax1.text(0.01, 0.02, '(a)', transform=ax1.transAxes, ha='left', va='bottom', fontsize=11)
    # Legend moved to bottom subplot to avoid overlap

    # Bottom: MSE
    for i, attr in enumerate(attributes):
        vals = [mses[it].get(attr, 0.0) for it in iters]
        ax2.bar(x + (i - (n_attr-1)/2)*width, vals, width=width, label=attr)
    avg_mse = [float(np.mean([mses[it].get(attr, 0.0) for attr in attributes])) for it in iters]
    ax2.plot(
        x,
        avg_mse,
        '-o',
        color='tab:cyan',
        linewidth=3.0,
        markersize=8,
        markerfacecolor='tab:cyan',
        markeredgecolor='tab:cyan',
        label='Avg across attributes',
    )
    # Optional expert baseline for 1-10 MSE will be added later if provided
    ax2.set_ylabel('Mean Squared Error (1â€“10)')
    ax2.set_ylim(0, 6)
    ax2.grid(axis='y', alpha=0.3)
    ax2.set_xticks(x)
    ax2.set_xticklabels([str(it) for it in iters])
    ax2.set_xlabel('Prompt Iteration')
    ax2.legend(ncol=min(3, n_attr), fontsize=10)
    ax2.text(0.01, 0.02, '(b)', transform=ax2.transAxes, ha='left', va='bottom', fontsize=11)

    fig.tight_layout()
    fig.savefig(out, dpi=200)
    plt.close(fig)


def main(argv: Optional[List[str]] = None) -> int:
    import glob
    import sys
    ap = argparse.ArgumentParser(description='Plot accuracy progression of 0,1,2 matches across prompt iterations')
    ap.add_argument('--glob', default='results_*.xlsx', help='Glob for per-space result workbooks')
    ap.add_argument('--xlsx', default='BD GPT Prompts.xlsx', help='Workbook containing Expert sheet')
    ap.add_argument('--out', default='accuracy_progression.png', help='Output image path')
    # Expert baseline options
    ap.add_argument('--baseline-acc', type=float, help='Override baseline for 0-2 accuracy (percentage)')
    ap.add_argument('--baseline-mse', type=float, help='Override baseline for 1-10 MSE')
    ap.add_argument('--expert-sheets', nargs='*', help='Optional list of sheet names with individual expert ratings (default: all non-Prompts/Spaces/Expert)')
    args = ap.parse_args(argv or sys.argv[1:])

    result_paths = [Path(p) for p in sorted(glob.glob(args.glob)) if p.lower().endswith('.xlsx')]
    if not result_paths:
        raise SystemExit('No result files matched.')

    expert_both = _read_expert_truths(Path(args.xlsx))
    # Derive attributes from Expert sheet order, fallback to alphabetical from results
    attributes = _attributes_from_expert(Path(args.xlsx))
    if not attributes:
        # Fallback: union of sheet names across results
        names: set[str] = set()
        for p in result_paths:
            wb = _load_wb(p)
            names.update(s for s in wb.sheetnames if s.lower() not in {"meta", "prompts"})
        attributes = sorted(names)

    expert02 = {s: {a: v.get('0-2') for a, v in amap.items() if '0-2' in v} for s, amap in expert_both.items()}
    expert10 = {s: {a: v.get('1-10') for a, v in amap.items() if '1-10' in v} for s, amap in expert_both.items()}

    iterations = _available_iterations(result_paths)
    percentages = compute_match_percentages(result_paths, expert02, attributes, iterations)
    mses = compute_mse(result_paths, expert10, attributes, iterations)

    # Compute or accept expert baselines
    baseline_acc = args.baseline_acc
    baseline_mse = args.baseline_mse
    # For debug/inspection
    dbg_matches = dbg_total = 0
    dbg_sse = 0.0
    dbg_n = 0
    if baseline_acc is None or baseline_mse is None:
        # Try computing from expert sheets inside the same workbook
        try:
            wb = _load_wb(Path(args.xlsx))
            ignore = {'prompts', 'spaces', 'expert', 'index'}
            candidate_sheets = args.expert_sheets if args.expert_sheets else [s for s in wb.sheetnames if s.lower() not in ignore]
            # Build expert maps per sheet
            matches = 0
            total = 0
            sse = 0.0
            n = 0
            for sheet in candidate_sheets:
                try:
                    indiv = _read_expert_truths(Path(args.xlsx), sheet_name=sheet)
                except Exception:
                    continue
                # Compare against consolidated expert02/expert10
                for space, attrs_map in expert02.items():
                    # 0-2
                    indiv_space02 = indiv.get(space) or _match_space_dict(indiv, space)
                    if not indiv_space02:
                        continue
                    for attr, gt02 in attrs_map.items():
                        # use normalized attribute matching
                        v02 = _get_attr_scale_value(indiv_space02, attr, '0-2')
                        if v02 is not None:
                            total += 1
                            if float(v02) == float(gt02):
                                matches += 1
                # 1-10
                for space, attrs_map10 in expert10.items():
                    indiv_space10 = indiv.get(space) or _match_space_dict(indiv, space)
                    if not indiv_space10:
                        continue
                    for attr, gt10 in attrs_map10.items():
                        v10 = _get_attr_scale_value(indiv_space10, attr, '1-10')
                        if v10 is not None:
                            diff = float(v10) - float(gt10)
                            sse += diff * diff
                            n += 1
            if baseline_acc is None and total > 0:
                baseline_acc = 100.0 * matches / total
            if baseline_mse is None and n > 0:
                baseline_mse = sse / n
            # capture for debug prints
            dbg_matches, dbg_total, dbg_sse, dbg_n = matches, total, sse, n
        except Exception:
            pass

    # Plot
    try:
        import matplotlib.pyplot as plt  # type: ignore
    except Exception:
        plt = None
    # Render with optional baselines
    try:
        import matplotlib.pyplot as _plt
        import numpy as _np
        iters = sorted(percentages.keys())
        n_attr = len(attributes)
        x = _np.arange(len(iters))
        width = min(0.16, 0.8 / max(1, n_attr))
        fig, (ax1, ax2) = _plt.subplots(2, 1, figsize=(11, 10), sharex=True)
        # Top bars
        for i, attr in enumerate(attributes):
            vals = [percentages[it].get(attr, 0.0) for it in iters]
            ax1.bar(x + (i - (n_attr-1)/2)*width, vals, width=width, label=attr)
        avg_acc = [float(_np.mean([percentages[it].get(attr, 0.0) for attr in attributes])) for it in iters]
        ax1.plot(x, avg_acc, '-o', color='tab:cyan', linewidth=3.0, markersize=8, markerfacecolor='tab:cyan', markeredgecolor='tab:cyan', label='Avg across attributes')
        if baseline_acc is not None:
            ax1.axhline(float(baseline_acc), color='black', linestyle='--', linewidth=2.0, dashes=(10, 6), label='Average Expert Performance')
        ax1.set_ylabel('Accuracy (%, 0,1,2 match)')
        ax1.set_ylim(0, 106)
        ax1.grid(axis='y', alpha=0.3)
        ax1.text(0.01, 0.02, '(a)', transform=ax1.transAxes, ha='left', va='bottom', fontsize=11)
        # Bottom bars
        for i, attr in enumerate(attributes):
            vals = [mses[it].get(attr, 0.0) for it in iters]
            ax2.bar(x + (i - (n_attr-1)/2)*width, vals, width=width, label=attr)
        avg_mse = [float(_np.mean([mses[it].get(attr, 0.0) for attr in attributes])) for it in iters]
        ax2.plot(x, avg_mse, '-o', color='tab:cyan', linewidth=3.0, markersize=8, markerfacecolor='tab:cyan', markeredgecolor='tab:cyan', label='Avg across attributes')
        if baseline_mse is not None:
            ax2.axhline(float(baseline_mse), color='black', linestyle='--', linewidth=2.0, dashes=(10, 6), label='Average Expert Performance')
        ax2.set_ylabel('Mean Squared Error (1â€“10)')
        ax2.set_ylim(0, 6)
        ax2.grid(axis='y', alpha=0.3)
        ax2.set_xticks(x)
        ax2.set_xticklabels([str(it) for it in iters])
        ax2.set_xlabel('Prompt Iteration')
        ax2.legend(ncol=min(3, n_attr), fontsize=10)
        ax2.text(0.01, 0.02, '(b)', transform=ax2.transAxes, ha='left', va='bottom', fontsize=11)
        # Print computed baseline values for verification
        if baseline_acc is not None:
            print(f"Baseline (0â€“2) accuracy: {float(baseline_acc):.2f}%" + (f"  (matches={dbg_matches} / total={dbg_total})" if dbg_total else ""))
        else:
            print("Baseline (0â€“2) accuracy: not set")
        if baseline_mse is not None:
            print(f"Baseline (1â€“10) MSE: {float(baseline_mse):.4f}" + (f"  (n={dbg_n})" if dbg_n else ""))
        else:
            print("Baseline (1â€“10) MSE: not set")

        fig.tight_layout()
        fig.savefig(Path(args.out), dpi=200)
        _plt.close(fig)
        print(f'Saved {args.out}')
        return 0
    except Exception:
        # Fallback to original plot function if something goes sideways
        plot_dual_progression(percentages, mses, attributes, Path(args.out))
        print(f'Saved {args.out}')
        return 0
    print(f'Saved {args.out}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
