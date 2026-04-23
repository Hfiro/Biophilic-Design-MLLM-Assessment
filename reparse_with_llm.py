"""
Author: Hoda Alem
Date: November 2025
Context: Part of PhD research for a PhD in Environmental Design and Planning
         at Myers-Lawson School of Construction, Virginia Tech, Blacksburg, VA.

Stand-alone LLM-based reparser for a single results workbook.

Features (as requested):
- Target ONE Excel results file (not a glob).
- Allow filtering by sheet (attribute), iteration, and run (instance);
  each can be a single value or "all" (default: all).
- For the selected rows, read the raw response and call the API twice:
  1) Extract the 0â€“2 rating (return just 0,1,2 or nothing)
  2) Extract the 1â€“10 score (return just a number 1..10 or nothing)
- Write exactly the two numbers to rating_0_2 and score_1_10 columns.
- No changes to justification and no extra validation.
"""

import argparse
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple


def _headers(ws) -> Dict[str, int]:
    row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    out: Dict[str, int] = {}
    for i, v in enumerate(row, start=1):
        if isinstance(v, str):
            out[v.strip()] = i
    return out


def _cell_val(row, idx: Dict[str, int], key: str) -> Any:
    col = idx.get(key)
    if not col:
        return None
    return row[col - 1] if col - 1 < len(row) else None


def _is_number(x: Any) -> bool:
    return isinstance(x, (int, float))


def _load_env():
    try:
        from biophilic_eval import _load_env_from_dotenv  # type: ignore
        _load_env_from_dotenv()
    except Exception:
        pass


PROMPT_RATING = (
    "From the text below, extract the categorical rating on a 0â€“2 scale. "
    "Choose the number explicitly tied to the 0â€“2 scale (e.g., '/2', 'out of 2', '0â€“2 rating', 'presence rating'). "
    "Output exactly one of: 0, 1, or 2. No other text. If no explicit 0â€“2 rating is present, output nothing. "
    "Text: \"\"\"{text}\"\"\""
)

PROMPT_SCORE = (
    "From the text below, extract the score on the 1â€“10 scale. "
    "Choose the number explicitly tied to the 1â€“10 scale (e.g., '/10', 'out of 10', '1â€“10 score', 'overall score'). "
    "Output exactly one number between 1 and 10 (integer or decimal). No other text. If no explicit 1â€“10 score is present, output nothing. "
    "Text: \"\"\"{text}\"\"\""
)


def _ask_once(prompt_text: str, model: str, timeout: float) -> Optional[str]:
    from openai import OpenAI
    client = OpenAI()
    resp = client.chat.completions.create(
        model=model,
        messages=[{"role": "user", "content": prompt_text}],
        temperature=0.0,
        timeout=timeout,
    )
    return (resp.choices[0].message.content or "").strip()


def _extract_rating_and_score_via_llm(raw_text: str, model: str, timeout: float) -> Tuple[Optional[int], Optional[float]]:
    # First, rating 0â€“2
    try:
        rtext = _ask_once(PROMPT_RATING.format(text=raw_text), model=model, timeout=timeout)
    except Exception:
        rtext = None
    rating: Optional[int] = None
    if rtext:
        try:
            rating = int(str(rtext).strip())
        except Exception:
            rating = None

    # Then, score 1â€“10
    try:
        stext = _ask_once(PROMPT_SCORE.format(text=raw_text), model=model, timeout=timeout)
    except Exception:
        stext = None
    score: Optional[float] = None
    if stext:
        try:
            score = float(str(stext).strip())
        except Exception:
            score = None

    return rating, score


def reparse_file(
    path: Path,
    model: str,
    timeout: float,
    dry_run: bool,
    copy_mode: bool,
    only_missing: bool,
    limit: Optional[int],
    sheets_filter: Optional[List[str]] = None,
    iterations_filter: Optional[List[int]] = None,
    runs_filter: Optional[List[int]] = None,
    print_only: bool = False,
) -> Tuple[int, int]:
    from openpyxl import load_workbook
    try:
        from biophilic_eval import _first_sentences  # justification helper
    except Exception:
        _first_sentences = None  # type: ignore

    target_path = path
    if copy_mode:
        target_path = path.with_name(f"{path.stem}_reparsed{path.suffix}")

    wb = load_workbook(filename=str(path), read_only=False, data_only=True)
    total = 0
    updated = 0
    remaining = limit if limit is not None else None

    for sheet in wb.sheetnames:
        if sheet in {"meta", "prompts"}:
            continue
        if sheets_filter and sheet not in sheets_filter:
            continue
        ws = wb[sheet]
        idx = _headers(ws)
        if not idx:
            continue
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row is None:
                continue
            run_val = _cell_val(row, idx, "run")
            if not _is_number(run_val):
                continue
            try:
                it_val = _cell_val(row, idx, "iteration")
                it_num = int(it_val) if it_val is not None else None
            except Exception:
                it_num = None
            if iterations_filter and it_num not in iterations_filter:
                continue
            if runs_filter and int(run_val) not in runs_filter:
                continue

            raw = _cell_val(row, idx, "raw")
            if not isinstance(raw, str) or not raw.strip():
                continue  # no text to parse

            r02 = _cell_val(row, idx, "rating_0_2")
            s10 = _cell_val(row, idx, "score_1_10")
            if only_missing and (r02 not in (None, "") and s10 not in (None, "")):
                continue

            total += 1
            if dry_run and not print_only:
                print(f"{path.name} | sheet={sheet} | space={space} | attribute={attr} | run={run_val}")
                continue

            rating, score = _extract_rating_and_score_via_llm(raw, model=model, timeout=timeout)
            if rating is None and score is None:
                continue

            def _set(name: str, value: Any):
                col = idx.get(name)
                if col:
                    ws.cell(row=row_idx, column=col, value=value)

            if print_only:
                print(f"sheet={sheet} | iteration={it_num} | run={int(run_val)} | rating_0_2={'' if rating is None else rating} | score_1_10={'' if score is None else score}")
            else:
                _set("rating_0_2", rating if rating is not None else "")
                _set("score_1_10", score if score is not None else "")
                updated += 1

            if remaining is not None:
                remaining -= 1
                if remaining <= 0:
                    break
        if remaining is not None and remaining <= 0:
            break

    if not dry_run and not print_only and updated > 0:
        wb.save(str(target_path))
    wb.close()
    return total, updated


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Reparse raw text in a results workbook using the API to extract rating_0_2 and score_1_10.")
    parser.add_argument("--file", required=True, help="Path to one results_*.xlsx file")
    parser.add_argument("--model", default="gpt-4o-mini", help="LLM model for extraction (text-only)")
    parser.add_argument("--timeout", type=float, default=60.0, help="Per-request timeout")
    parser.add_argument("--dry-run", action="store_true", help="List rows that would be reparsed (no API calls)")
    parser.add_argument("--print-only", action="store_true", help="Do not modify files; call API and print parsed numbers for matching rows")
    parser.add_argument("--copy", action="store_true", help="Write changes to <file>_reparsed.xlsx (default unless --in-place)")
    parser.add_argument("--in-place", action="store_true", help="Modify the input workbook directly")
    parser.add_argument("--only-missing", action="store_true", help="Only fill rows missing rating and/or score")
    parser.add_argument("--limit", type=int, help="Maximum rows to reparse")
    parser.add_argument("--sheets", nargs="*", help="Restrict to specific sheets (attributes)")
    parser.add_argument("--iterations", nargs="*", type=int, help="Restrict to specific iteration numbers")
    parser.add_argument("--runs", nargs="*", type=int, help="Restrict to specific run numbers")
    args = parser.parse_args(argv or sys.argv[1:])

    _load_env()

    p = Path(args.file)
    if not (p.exists() and p.suffix.lower() == ".xlsx"):
        raise SystemExit(f"File not found or not an .xlsx: {p}")

    copy_mode = True
    if args.in_place:
        copy_mode = False
    if args.copy:
        copy_mode = True

    t, u = reparse_file(
        p,
        model=args.model,
        timeout=float(args.timeout),
        dry_run=bool(args.dry_run),
        copy_mode=copy_mode,
        only_missing=bool(args.only_missing),
        limit=args.limit,
        sheets_filter=args.sheets,
        iterations_filter=args.iterations,
        runs_filter=args.runs,
        print_only=bool(args.print_only),
    )
    print(f"{p.name}: candidates={t}, updated={u}")
    if args.dry_run and not args.print_only:
        print("No changes written (dry run)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
