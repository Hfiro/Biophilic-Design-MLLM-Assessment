"""
Author: Hoda Alem
Date: November 2025
Context: Part of PhD research for a PhD in Environmental Design and Planning
         at Myers-Lawson School of Construction, Virginia Tech, Blacksburg, VA.
Purpose: Automate evaluation of biophilic design attributes from five images per
         built space using OpenAI vision models, driven by prompts from an Excel
         workbook.
Features:
- Reads attributes, prompts, and space lists from a Prompts workbook (XLSX).
- Encodes five JPG images as data URLs and calls OpenAI chat/vision API.
- Extracts 0â€“1â€“2 categorical rating and 1â€“10 scalar score from model text.
- Supports fallback model, timeouts, and basic error handling.
- Writes per-space results to XLSX/CSV with meta, prompts, raw responses, and
  per-iteration summaries (median/MAD for 0â€“2; mean/variance for 1â€“10).
- Utility modes to list attributes, spaces, and prompts without running the API.
"""

import os
import sys
import base64
import mimetypes
import argparse
import re
import csv
import time
from pathlib import Path
import statistics as stats
import unicodedata
from datetime import datetime

from typing import Optional, List, Dict, Any, Pattern, Tuple

try:
    # Optional: load .env if python-dotenv is installed
    from dotenv import load_dotenv  # type: ignore
except Exception:
    load_dotenv = None  # graceful fallback


DEFAULT_PROMPT_TMPL = (
    "Please review these five pictures of this built space. Then, using Stephen Kellertâ€™s "
    "Biophilic Design framework for \"{attribute}\", as defined below:\n\n"
    "\"Color has long been instrumental in human evolution and survival, enhancing the "
    "ability to locate food, resources, and water; identify danger; facilitate visual access; "
    "foster mobility; and more. People for, good and obvious reasons, are attracted to bright "
    "flowering colors, rainbows, beautiful sunsets, glistening water, blue skies, and other colorful "
    "features of the natural world. Natural colors, such as earth tones, are thus often used to good "
    "effect by designers.\"\n\n"
    "Please rate the incorporation of \"{attribute}\" in the built space on a scale of 0, 1, or 2:\n"
    "  0 = Zero\n  1 = Some\n  2 = A Large Amount\n\n"
    "Next, on a scale from 1 to 10, score the incorporation of \"{attribute}\" in this space. "
    "In maximum 3 sentences, provide a concise justification for both your 0,1,2 rating and your 1-10 score."
)


def _encode_image_as_data_url(path: str) -> str:
    mime, _ = mimetypes.guess_type(path)
    if not mime:
        mime = "image/jpeg"
    with open(path, "rb") as f:
        b64 = base64.b64encode(f.read()).decode("utf-8")
    return f"data:{mime};base64,{b64}"


def _load_env_from_dotenv() -> None:
    env_path = os.path.join(os.getcwd(), ".env")
    if not os.path.exists(env_path):
        return
    if load_dotenv is not None:
        try:
            load_dotenv(dotenv_path=env_path, override=False)
            return
        except Exception:
            pass
    try:
        with open(env_path, "r", encoding="utf-8") as f:
            for raw in f:
                line = raw.strip()
                if not line or line.startswith("#"):
                    continue
                if "=" not in line:
                    continue
                key, value = line.split("=", 1)
                key = key.strip()
                value = value.strip().strip('"').strip("'")
                os.environ.setdefault(key, value)
    except Exception:
        pass


def analyze_color(
    image_data_urls: List[str],
    prompt_text: str,
    model: str = "gpt-4o",
    detail: str = "high",
    allow_fallback: bool = True,
    timeout: float = 90,
) -> tuple[str, str]:
    if not image_data_urls or len(image_data_urls) != 5:
        raise ValueError("Expected exactly 5 image data URLs.")

    # Lazy import to avoid requiring openai when using utility modes
    try:
        from openai import OpenAI
        from openai import (
            APIError,
            RateLimitError,
            AuthenticationError,
            BadRequestError,
        )
    except Exception as e:
        raise SystemExit(
            "Missing dependency: openai. Install with `pip install --upgrade openai` or `pip install -r requirements.txt`."
        ) from e

    client = OpenAI()

    def _call(the_model: str, the_detail: str):
        return client.chat.completions.create(
            model=the_model,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt_text},
                        *[
                            {
                                "type": "image_url",
                                "image_url": {"url": url, "detail": the_detail},
                            }
                            for url in image_data_urls
                        ],
                    ],
                }
            ],
            temperature=0.2,
            timeout=timeout,
        )

    try:
        resp = _call(model, detail)
        return resp.choices[0].message.content.strip(), model
    except (RateLimitError, APIError) as e:
        err_text = str(e)
        quota_like = any(
            s in err_text.lower() for s in ["quota", "insufficient", "rate limit", "429"]
        )
        if allow_fallback and quota_like and model != "gpt-4o-mini":
            try:
                resp = _call("gpt-4o-mini", "low")
                return resp.choices[0].message.content.strip(), "gpt-4o-mini"
            except Exception as e2:
                raise RuntimeError(
                    "Quota or rate limit encountered; fallback to gpt-4o-mini also failed. "
                    "Consider adding billing, lowering usage limits, or trying later.\n"
                    f"Original error: {e}\nFallback error: {e2}"
                )
        raise
    except (AuthenticationError, BadRequestError) as e:
        raise RuntimeError(
            "Authentication or request error. Verify OPENAI_API_KEY, model access, and image paths.\n"
            f"Details: {e}"
        )


def main(argv: Optional[list] = None) -> int:
    argv = sys.argv[1:] if argv is None else argv

    parser = argparse.ArgumentParser(description="Evaluate biophilic design attributes from five images using GPT-Vision.")
    parser.add_argument("--model", default=os.environ.get("OPENAI_MODEL", "gpt-4o"), help="OpenAI model to use")
    parser.add_argument("--detail", default=os.environ.get("OPENAI_IMAGE_DETAIL", "high"), choices=["low", "high"], help="Image detail level")
    parser.add_argument("--no-fallback", action="store_true", help="Disable fallback to gpt-4o-mini on quota errors")
    parser.add_argument("--out", default=None, help="Path to output CSV file (default: auto based on space & iteration)")
    parser.add_argument("--sleep", type=float, default=0.0, help="Seconds to sleep between runs")
    parser.add_argument("--timeout", type=float, default=float(os.environ.get("OPENAI_TIMEOUT", 90)), help="Per-request timeout in seconds")
    parser.add_argument("--xlsx", default="BD GPT Prompts.xlsx", help="Path to the prompts workbook (.xlsx)")
    parser.add_argument("--list-attributes", metavar="XLSX", help="Print attributes from the Prompts sheet (col 2, from row 3) and exit")
    parser.add_argument("--list-spaces", metavar="XLSX", help="Print built-space names from the Spaces sheet (col 1, from row 2) and exit")
    parser.add_argument("--list-prompts", metavar="XLSX", help="Print 5Ã—4 prompts from Prompts sheet (col 2 attr, cols 3-6 prompts, from row 3) and exit")
    args = parser.parse_args(argv)

    # Utility mode: list attributes from the given xlsx and exit
    if args.list_attributes:
        try:
            attrs = _read_attributes_from_prompts_sheet(Path(args.list_attributes))
        except Exception as e:
            print(f"Failed to read attributes: {e}", file=sys.stderr)
            return 2
        if not attrs:
            print("No attributes found.")
        else:
            print("Attributes:")
            for i, a in enumerate(attrs, start=1):
                print(f"  {i}. {a}")
        return 0

    # Utility mode: list spaces from the given xlsx and exit
    if args.list_spaces:
        try:
            spaces = _read_spaces_from_sheet(Path(args.list_spaces))
        except Exception as e:
            print(f"Failed to read spaces: {e}", file=sys.stderr)
            return 2
        if not spaces:
            print("No spaces found.")
        else:
            print("Spaces:")
            for i, s in enumerate(spaces, start=1):
                print(f"  {i}. {s}")
        return 0

    # Utility mode: list prompts from the Prompts sheet and exit
    if args.list_prompts:
        try:
            prompts = _read_attribute_prompts_from_sheet(Path(args.list_prompts))
        except Exception as e:
            print(f"Failed to read prompts: {e}", file=sys.stderr)
            return 2
        if not prompts:
            print("No prompts found.")
        else:
            print("Prompts (attribute â†’ iteration â†’ prompt):")
            # Group by attribute for nicer display
            by_attr: Dict[str, List[Dict[str, Any]]] = {}
            for rec in prompts:
                by_attr.setdefault(rec["attribute"], []).append(rec)
            for attr, rows in by_attr.items():
                rows = sorted(rows, key=lambda r: r["iteration"])  # type: ignore
                print(f"- {attr}:")
                for r in rows:
                    it = r["iteration"]
                    text = (r["prompt"] or "").replace("\n", " ")
                    print(f"  [{it}] {text}")
            print(f"Total prompts: {len(prompts)}")
        return 0

    _load_env_from_dotenv()

    # Gather the three required inputs
    iterations = _prompt_iterations()
    runs = _prompt_runs()
    space_name = _prompt_space_name()

    # Resolve workbook path (default to --xlsx if exists; otherwise prompt)
    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"Workbook not found at default path: {xlsx_path}")
        xlsx_path = _prompt_xlsx_path()

    # Read prompts for selected iteration(s)
    all_prompts = _read_attribute_prompts_from_sheet(xlsx_path)
    # Resolve 'All' sentinel (-1) to all iterations present
    if -1 in iterations:
        iter_set = {int(p.get("iteration")) for p in all_prompts if p.get("iteration") is not None}
    else:
        iter_set = set(iterations)
    prompts_for_iter = [p for p in all_prompts if p.get("iteration") in iter_set]
    if not prompts_for_iter:
        label = ",".join(map(str, iterations))
        print(f"No prompts found for iteration(s) {label} in {xlsx_path}", file=sys.stderr)
        return 2

    # Allow user to pick a specific attribute subset before running
    attribute_options = sorted(
        {str(p.get("attribute", "")).strip() for p in prompts_for_iter if p.get("attribute")}
    )
    if attribute_options:
        selected_attributes = _prompt_attribute_selection(attribute_options)
        prompts_for_iter = [
            p for p in prompts_for_iter if str(p.get("attribute", "")).strip() in selected_attributes
        ]
        if not prompts_for_iter:
            print("No prompts to run after applying the selected attribute filter.", file=sys.stderr)
            return 2

    # Determine which spaces to run
    spaces: List[str]
    if space_name.strip().lower() in {"all", "a"}:
        spaces = _read_spaces_from_sheet(xlsx_path)
        if not spaces:
            print("No spaces found in workbook's Spaces sheet.", file=sys.stderr)
            return 2
    else:
        spaces = [space_name.strip()]

    # Ensure API key is set (only now needed since we're calling the API)
    if not os.environ.get("OPENAI_API_KEY"):
        print(
            "Environment variable OPENAI_API_KEY is not set.\n"
            "Put it in .env as OPENAI_API_KEY=sk-... or export in your shell.",
            file=sys.stderr,
        )
        return 2

    # If not provided, default to an .xlsx so we can include multiple sheets
    iteration_label = ("All" if len(iterations) > 1 else str(iterations[0]))

    for space in spaces:
        space_folder = Path(space)
        if not space_folder.is_absolute():
            space_folder = Path(os.getcwd()) / space_folder
        if not space_folder.exists() or not space_folder.is_dir():
            print(f"Space folder not found: {space_folder}", file=sys.stderr)
            continue
        try:
            image_paths = _collect_five_jpgs(space_folder)
        except Exception as e:
            print(f"{space_folder.name}: {e}", file=sys.stderr)
            continue
        image_data_urls = [_encode_image_as_data_url(str(p)) for p in image_paths]

        out_path = args.out or f"results_{space_folder.name}_iter{iteration_label}.xlsx"

        rows: List[Dict[str, Any]] = []
        for p_index, rec in enumerate(prompts_for_iter, start=1):
            attribute = rec.get("attribute", "").strip()
            prompt_text = rec.get("prompt", "").strip()
            if not attribute or not prompt_text:
                continue

            for i in range(1, runs + 1):
                print(
                    f"[{space_folder.name}] Attr {p_index}/{len(prompts_for_iter)} {attribute} | Run {i}/{runs}: calling {args.model} (detail={args.detail})...",
                    flush=True,
                )
                try:
                    text, used_model = analyze_color(
                        image_data_urls=image_data_urls,
                        prompt_text=prompt_text,
                        model=args.model,
                        detail=args.detail,
                        allow_fallback=not args.no_fallback,
                        timeout=args.timeout,
                    )
                except Exception as e:
                    print(f"[{space_folder.name}] {attribute} Run {i} error: {e}", file=sys.stderr)
                    rows.append(
                        {
                            "task": p_index,
                            "space": space_folder.name,
                            "attribute": attribute,
                            "iteration": rec.get("iteration"),
                            "run": i,
                            "model": "",
                            "rating_0_2": "",
                            "score_1_10": "",
                            "justification": f"ERROR: {e}",
                            "raw": "",
                        }
                    )
                    if args.sleep:
                        time.sleep(args.sleep)
                    continue

                rating, score = _extract_rating_and_score(text)
                rows.append(
                    {
                        "task": p_index,
                        "space": space_folder.name,
                        "attribute": attribute,
                        "iteration": rec.get("iteration"),
                        "run": i,
                        "model": used_model,
                        "rating_0_2": rating if rating is not None else "",
                        "score_1_10": score if score is not None else "",
                        "justification": _first_sentences(text, max_sentences=3),
                        "raw": text,
                    }
                )
                print(
                    f"[{space_folder.name}] Attr {p_index}/{len(prompts_for_iter)} {attribute} | Run {i}/{runs}: done.",
                    flush=True,
                )
                if args.sleep:
                    time.sleep(args.sleep)

        # Output per space
        try:
            if str(out_path).lower().endswith(".xlsx"):
                _write_xlsx(
                    out_path,
                    rows,
                    meta={
                        "space": space_folder.name,
                        "iterations": ",".join(map(str, iterations)),
                        "model": args.model,
                        "detail": args.detail,
                        "runs": runs,
                        "timestamp": datetime.now().isoformat(timespec="seconds"),
                    },
                    prompts=prompts_for_iter,
                )
            else:
                _write_csv(out_path, rows)
            print(f"Saved results to {out_path}")
        except Exception as e:
            print(f"Failed to write output for {space_folder.name}: {e}", file=sys.stderr)
            continue

        # Summary across all rows for this space
        ratings = [r["rating_0_2"] for r in rows if isinstance(r.get("rating_0_2"), int)]
        scores = [r["score_1_10"] for r in rows if isinstance(r.get("score_1_10"), (int, float))]
        if ratings:
            print(f"[{space_folder.name}] Median rating (0-2): {stats.median(ratings)}")
        else:
            print(f"[{space_folder.name}] Median rating (0-2): n/a")
        if scores:
            print(f"[{space_folder.name}] Mean score (1-10): {round(stats.fmean(scores), 3)}")
        else:
            print(f"[{space_folder.name}] Mean score (1-10): n/a")

        print(
            f"Completed {len(rows)} rows across {len(prompts_for_iter)} attribute(s) for space '{space_folder.name}'."
        )

    return 0


def _collect_five_jpgs(folder: Path) -> List[Path]:
    candidates = sorted(
        [p for p in folder.iterdir() if p.is_file() and p.suffix.lower() in {".jpg", ".jpeg"}]
    )
    if len(candidates) != 5:
        raise FileNotFoundError(
            f"Expected exactly 5 JPG files in {folder}, found {len(candidates)}"
        )
    return candidates


def _prompt_folder() -> Path:
    while True:
        raw = input("Enter the subfolder containing exactly five JPG images: ").strip()
        if not raw:
            print("Please enter a folder path.")
            continue
        folder = Path(os.path.expanduser(raw))
        if not folder.exists() or not folder.is_dir():
            print(f"Not a valid folder: {folder}")
            continue
        try:
            _ = _collect_five_jpgs(folder)
            return folder
        except Exception as e:
            print(e)
            continue


def _prompt_runs() -> int:
    while True:
        raw = input("Enter number of runs (positive integer): ").strip()
        try:
            n = int(raw)
            if n > 0:
                return n
        except Exception:
            pass
        print("Please enter a valid positive integer.")


def _prompt_iterations() -> List[int]:
    while True:
        raw = input("Select prompt iteration (>=1) or 'All': ").strip()
        if not raw:
            print("Please enter a positive integer or 'All'.")
            continue
        if raw.lower() in {"all", "a"}:
            return [-1]
        try:
            n = int(raw)
            if 1 <= n:
                return [n]
        except Exception:
            pass
        print("Please enter a positive integer (e.g., 1) or 'All'.")


def _prompt_space_name() -> str:
    while True:
        raw = input("Enter the built space name (matches folder name) or 'All': ").strip()
        if raw:
            return raw
        print("Please enter a non-empty name.")


def _prompt_attribute_selection(options: List[str]) -> List[str]:
    """Prompt the user to select one attribute or all available."""
    if not options:
        return []
    lookup = {opt.lower(): opt for opt in options}
    display = ", ".join(options)
    while True:
        raw = input(f"Enter the attribute name ({display}) or 'All': ").strip()
        if not raw:
            print("Please enter an attribute name or 'All'.")
            continue
        lowered = raw.lower()
        if lowered in {"all", "a"}:
            return options
        match = lookup.get(lowered)
        if match:
            return [match]
        print("Attribute not recognized. Please enter one of:", display)


def _prompt_yes_no(question: str) -> bool:
    raw = input(question).strip().lower()
    return raw in {"y", "yes"}


def _prompt_xlsx_path() -> Path:
    while True:
        raw = input("Enter path to the Excel workbook (.xlsx): ").strip()
        if not raw:
            print("Please enter a file path.")
            continue
        p = Path(os.path.expanduser(raw))
        if p.exists() and p.is_file() and p.suffix.lower() == ".xlsx":
            return p
        print("Not a valid .xlsx file. Try again.")


def _read_tasks_from_xlsx(path: Path) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:
        raise SystemExit(
            "Missing dependency: openpyxl. Install with `pip install openpyxl` or `pip install -r requirements.txt`."
        ) from e

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb.active

    # Expect header row with at least these columns:
    # space_folder | attribute | prompt
    headers = {}
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    for idx, val in enumerate(header_row):
        if isinstance(val, str):
            headers[val.strip().lower()] = idx

    required = ["space_folder", "attribute", "prompt"]
    for req in required:
        if req not in headers:
            raise ValueError(
                f"Workbook missing required column '{req}'. Found columns: {list(headers.keys())}"
            )

    tasks: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        space_folder = str(row[headers["space_folder"]]).strip() if row[headers["space_folder"]] else ""
        if not space_folder:
            continue
        attribute = str(row[headers["attribute"]]).strip() if row[headers["attribute"]] else "Color"
        prompt_val = row[headers["prompt"]]
        if isinstance(prompt_val, str) and prompt_val.strip():
            prompt_text = prompt_val.strip()
        else:
            prompt_text = DEFAULT_PROMPT_TMPL.format(attribute=attribute)

        tasks.append({
            "space_folder": space_folder,
            "attribute": attribute,
            "prompt": prompt_text,
        })

    if not tasks:
        raise ValueError("No tasks found in workbook (no data rows or missing values).")
    return tasks


def _read_attributes_from_prompts_sheet(path: Path, sheet_name: str = "Prompts", column_index: int = 2, start_row: int = 3) -> List[str]:
    """Read all attributes from the given workbook.

    - Uses worksheet named `sheet_name` (default 'Prompts')
    - Reads from 1-based `column_index` (default 2 => column B)
    - Starts at `start_row` (default 3)
    - Returns a list with up to `max_items` non-empty strings
    """
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:
        raise SystemExit(
            "Missing dependency: openpyxl. Install with `pip install openpyxl` or `pip install -r requirements.txt`."
        ) from e

    if not path.exists() or not path.is_file():
        raise FileNotFoundError(f"Workbook not found: {path}")

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]

    attrs: List[str] = []
    r = start_row
    while r <= ws.max_row:
        val = ws.cell(row=r, column=column_index).value
        if isinstance(val, str):
            v = val.strip()
            if v:
                attrs.append(v)
        elif val is not None:
            # Coerce non-strings to strings
            v = str(val).strip()
            if v:
                attrs.append(v)
        r += 1

    return attrs


def _read_attribute_prompts_from_sheet(
    path: Path,
    sheet_name: str = "Prompts",
    attr_col: int = 2,
    first_prompt_col: int = 3,
    start_row: int = 3,
) -> List[Dict[str, Any]]:
    """Read attribute prompts grid dynamically (all attributes Ã— all iterations).

    - `sheet_name`: worksheet name ('Prompts')
    - `attr_col`: 1-based column index with attribute names (default B=2)
    - `first_prompt_col`: starting 1-based column for prompts (default C=3)
    - `iterations`: number of prompt iterations per attribute (default 4 => C..F)
    - `start_row`: 1-based first row with data (default 3)
    - `max_attrs`: maximum attributes to read (default 5 rows)
    Returns a list of dicts: {attribute, iteration, prompt}
    """
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:
        raise SystemExit(
            "Missing dependency: openpyxl. Install with `pip install openpyxl` or `pip install -r requirements.txt`."
        ) from e

    if not path.exists() or not path.is_file():
        raise FileNotFoundError(f"Workbook not found: {path}")

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]

    out: List[Dict[str, Any]] = []
    r = start_row
    while r <= ws.max_row:
        attr_val = ws.cell(row=r, column=attr_col).value
        if isinstance(attr_val, str):
            attribute = attr_val.strip()
        elif attr_val is not None:
            attribute = str(attr_val).strip()
        else:
            attribute = ""

        if not attribute:
            # Stop if no attribute; or skip blank rows and continue
            r += 1
            continue

        i = 0
        while True:
            col = first_prompt_col + i
            if col > ws.max_column:
                break
            cell_val = ws.cell(row=r, column=col).value
            if isinstance(cell_val, str):
                prompt = cell_val.strip()
            elif cell_val is not None:
                prompt = str(cell_val).strip()
            else:
                prompt = ""

            if prompt:
                out.append({
                    "attribute": attribute,
                    "iteration": i + 1,
                    "prompt": prompt,
                })
            else:
                empty_tail = True
                for look_ahead in range(1, 3):
                    c2 = col + look_ahead
                    if c2 <= ws.max_column and ws.cell(row=r, column=c2).value not in (None, ""):
                        empty_tail = False
                        break
                if empty_tail:
                    break
            i += 1

        r += 1

    return out


def _read_spaces_from_sheet(path: Path, sheet_name: str = "Spaces", column_index: int = 1, start_row: int = 2) -> List[str]:
    """Read built-space names from the given workbook.

    - Uses worksheet named `sheet_name` (default 'Spaces')
    - Reads from 1-based `column_index` (default 1 => column A)
    - Starts at `start_row` (default 2)
    - Returns all non-empty strings in that column
    """
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:
        raise SystemExit(
            "Missing dependency: openpyxl. Install with `pip install openpyxl` or `pip install -r requirements.txt`."
        ) from e

    if not path.exists() or not path.is_file():
        raise FileNotFoundError(f"Workbook not found: {path}")

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]

    spaces: List[str] = []
    r = start_row
    while r <= ws.max_row:
        val = ws.cell(row=r, column=column_index).value
        if isinstance(val, str):
            v = val.strip()
            if v:
                spaces.append(v)
        elif val is not None:
            v = str(val).strip()
            if v:
                spaces.append(v)
        r += 1

    return spaces


_WORD_NUMBER_MAP = {
    "zero": 0,
    "one": 1,
    "two": 2,
    "three": 3,
    "four": 4,
    "five": 5,
    "six": 6,
    "seven": 7,
    "eight": 8,
    "nine": 9,
    "ten": 10,
}

_NUMBER_TOKEN_PATTERN = re.compile(r"\b(zero|one|two|three|four|five|six|seven|eight|nine|ten|\d+(?:\.\d+)?)\b", re.IGNORECASE)

_RATING_PATTERNS = [
    re.compile(r"(?:0\s*(?:[-â€“â€”]|to)\s*2\s*(?:rating|score|scale)?\s*(?:is|was|=|:)?\s*[\*\s]{0,5})(?P<value>[0-2])\b", re.IGNORECASE),
    re.compile(r"(?:rating|rated|rate)\s*(?:it\s*)?(?:at|as|is|=|:)?\s*[\*\s]{0,5}(?P<value>[0-2])(?:\s*(?:out\s*of\s*(?:2|two)|/\s*(?:2|two)))?\b", re.IGNORECASE),
    re.compile(r"\b(?P<value>[0-2])\s*/\s*(?:2|two)\b", re.IGNORECASE),
    re.compile(r"\b(?P<value>[0-2])\s*(?:out\s*of\s*(?:2|two))\b", re.IGNORECASE),
    re.compile(r"(?:presence\s*(?:rating|score|assessment)\s*(?:is|was|=|:)?\s*[\*\s]{0,5})(?P<value>[0-2])\b", re.IGNORECASE),
]

_SCORE_PATTERNS = [
    re.compile(r"(?:1\s*(?:[-â€“â€”]|to)\s*10\s*(?:rating|score|scale)?\s*(?:is|was|=|:)?\s*[\*\s]{0,5})(?P<value>(?:10|[0-9](?:\.\d+)?))\b", re.IGNORECASE),
    re.compile(r"(?:score|scored|rating|rated)\s*(?:it\s*)?(?:at|as|is|=|:)?\s*[\*\s]{0,5}(?P<value>(?:10|[0-9](?:\.\d+)?))\s*(?:out\s*of\s*(?:10|ten)|/\s*(?:10|ten))\b", re.IGNORECASE),
    re.compile(r"\b(?P<value>(?:10|[0-9](?:\.\d+)?))\s*/\s*(?:10|ten)\b", re.IGNORECASE),
    re.compile(r"\b(?P<value>(?:10|[0-9](?:\.\d+)?))\s*(?:out\s*of\s*(?:10|ten))\b", re.IGNORECASE),
    re.compile(r"(?:integration\s*(?:score|rating|assessment)\s*(?:is|was|=|:)?\s*[\*\s]{0,5})(?P<value>(?:10|[0-9](?:\.\d+)?))\b", re.IGNORECASE),
    re.compile(r"(?:overall\s*(?:score|rating)\s*(?:is|was|=|:)?\s*[\*\s]{0,5})(?P<value>(?:10|[0-9](?:\.\d+)?))\b", re.IGNORECASE),
]

_RATING_LINE_INDICATORS = [
    "0-2",
    "0 - 2",
    "0 to 2",
    "(0-2",
    "(0 to 2",
    "0/2",
    "out of 2",
    "two-point",
    "two point",
    "presence rating",
    "presence score",
    "presence assessment",
]

_SCORE_LINE_INDICATORS = [
    "1-10",
    "1 - 10",
    "1 to 10",
    "(1-10",
    "(1 to 10",
    "/10",
    "out of 10",
    "ten-point",
    "ten point",
    "integration score",
    "integration rating",
    "integration assessment",
    "overall score",
]

_RATING_KEYWORDS = [
    "0-2",
    "0 - 2",
    "0 to 2",
    "rating",
    "rate",
    "categorical",
    "scale",
    "score",
    "assessment",
    "presence rating",
    "presence score",
    "presence assessment",
    "presence evaluation",
]

_RATING_STRONG_KEYWORDS = [
    "0-2",
    "0 - 2",
    "0 to 2",
    "0/2",
    "presence rating",
    "presence score",
]

_SCORE_KEYWORDS = [
    "1-10",
    "1 - 10",
    "1 to 10",
    "score",
    "rating",
    "scale",
    "/10",
    "out of 10",
    "ten-point",
    "integration score",
    "integration rating",
    "integration assessment",
    "overall score",
    "integration evaluation",
]

_SCORE_STRONG_KEYWORDS = [
    "1-10",
    "1 - 10",
    "1 to 10",
    "/10",
    "out of 10",
    "integration score",
    "integration rating",
]


def _normalize_response_text(text: str) -> str:
    normalized = unicodedata.normalize("NFKC", text or "")
    normalized = normalized.replace("â€“", "-").replace("â€”", "-").replace("âˆ’", "-")
    return normalized


def _coerce_number(raw: str) -> Optional[float]:
    token = raw.strip()
    if not token:
        return None
    lower = token.lower()
    if lower in _WORD_NUMBER_MAP:
        return float(_WORD_NUMBER_MAP[lower])
    try:
        return float(token)
    except Exception:
        return None


def _validate_value(value: float, allowed_set: Optional[set[int]], min_value: Optional[float], max_value: Optional[float], allow_float: bool) -> bool:
    candidate = value
    if not allow_float:
        if abs(candidate - round(candidate)) > 1e-6:
            return False
        candidate = round(candidate)
    if min_value is not None and candidate < min_value:
        return False
    if max_value is not None and candidate > max_value:
        return False
    if allowed_set is not None:
        check = int(round(candidate))
        if check not in allowed_set:
            return False
    return True


def _token_is_scale_endpoint(source: str, start: int, end: int) -> bool:
    lower = source.lower()
    prev_char = lower[start - 1] if start > 0 else ""
    next_char = lower[end] if end < len(lower) else ""
    if prev_char in "-/" or next_char in "-/":
        return True
    after = lower[end:end + 4]
    if after.startswith(" to "):
        return True
    before = lower[max(0, start - 4):start]
    if before.endswith(" to "):
        return True
    return False


def _extract_value_from_line(
    line: str,
    indicator_ranges: List[Tuple[int, int]],
    allowed_set: Optional[set[int]],
    min_value: Optional[float],
    max_value: Optional[float],
    allow_float: bool,
) -> Optional[float]:
    candidates: List[Tuple[int, float]] = []
    for match in _NUMBER_TOKEN_PATTERN.finditer(line):
        raw = match.group(1)
        value = _coerce_number(raw)
        if value is None:
            continue
        if not _validate_value(value, allowed_set, min_value, max_value, allow_float):
            continue
        start, end = match.span(1)
        if _token_is_scale_endpoint(line, start, end):
            continue
        preceding = line[max(0, start - 6):start].lower()
        if max_value is not None and abs(value - max_value) < 1e-6 and ("out of" in preceding or "/" in preceding):
            continue
        candidates.append((start, value))

    if not candidates:
        return None

    after_candidates = [
        (start, value)
        for start, value in candidates
        if any(start >= end for (begin, end) in indicator_ranges)
    ]

    def distance_to_indicator(pos: int) -> int:
        distances = []
        for begin, end in indicator_ranges:
            center = (begin + end) // 2
            distances.append(abs(pos - center))
        return min(distances) if distances else 0

    if after_candidates:
        chosen_start, chosen_value = min(after_candidates, key=lambda item: item[0])
    else:
        chosen_start, chosen_value = min(
            candidates, key=lambda item: (distance_to_indicator(item[0]), item[0])
        )

    if not allow_float:
        return float(int(round(chosen_value)))
    return chosen_value


def _search_lines_for_value(normalized_text: str, indicator_terms: List[str], allowed_set: Optional[set[int]], min_value: Optional[float], max_value: Optional[float], allow_float: bool) -> Optional[float]:
    lowered_terms = [term.lower() for term in indicator_terms]
    for raw_line in normalized_text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        lower_line = line.lower()
        indicator_ranges: List[Tuple[int, int]] = []
        for term in lowered_terms:
            start = lower_line.find(term)
            while start != -1:
                indicator_ranges.append((start, start + len(term)))
                start = lower_line.find(term, start + 1)
        if not indicator_ranges:
            continue
        value = _extract_value_from_line(line, indicator_ranges, allowed_set, min_value, max_value, allow_float)
        if value is not None:
            return value
    return None


def _extract_value_with_patterns(normalized_text: str, patterns: List[Pattern[str]], allowed_set: Optional[set[int]], min_value: Optional[float], max_value: Optional[float], allow_float: bool) -> Optional[float]:
    for pattern in patterns:
        for match in pattern.finditer(normalized_text):
            raw = match.group("value")
            value = _coerce_number(raw)
            if value is None:
                continue
            if not _validate_value(value, allowed_set, min_value, max_value, allow_float):
                continue
            if not allow_float:
                return float(int(round(value)))
            return value
    return None


def _find_value_with_context(normalized_text: str, keywords: List[str], strong_keywords: List[str], allowed_set: Optional[set[int]], min_value: Optional[float], max_value: Optional[float], allow_float: bool) -> Optional[float]:
    matches: List[tuple[int, int, float]] = []
    lowered_keywords = [kw.lower() for kw in keywords]
    lowered_strong = [kw.lower() for kw in strong_keywords]
    lower_text = normalized_text.lower()
    for idx, match in enumerate(_NUMBER_TOKEN_PATTERN.finditer(normalized_text)):
        raw = match.group(1)
        value = _coerce_number(raw)
        if value is None:
            continue
        if not _validate_value(value, allowed_set, min_value, max_value, allow_float):
            continue
        start, end = match.span(1)
        if _token_is_scale_endpoint(normalized_text, start, end):
            continue
        window_start = max(0, start - 40)
        window_end = min(len(normalized_text), end + 40)
        window = lower_text[window_start:window_end]
        if not any(keyword in window for keyword in lowered_keywords):
            continue
        weight = 1
        if any(strong in window for strong in lowered_strong):
            weight += 1
        matches.append((weight, idx, value))
    if not matches:
        return None
    weight, order, chosen = sorted(matches, key=lambda item: (-item[0], item[1]))[0]
    if not allow_float:
        return float(int(round(chosen)))
    return chosen


def _extract_scale_value(
    normalized_text: str,
    patterns: List[Pattern[str]],
    indicator_terms: List[str],
    keywords: List[str],
    strong_keywords: List[str],
    allowed_set: Optional[set[int]] = None,
    min_value: Optional[float] = None,
    max_value: Optional[float] = None,
    allow_float: bool = True,
) -> Optional[float]:
    value = _search_lines_for_value(normalized_text, indicator_terms, allowed_set, min_value, max_value, allow_float)
    if value is not None:
        return value
    value = _extract_value_with_patterns(normalized_text, patterns, allowed_set, min_value, max_value, allow_float)
    if value is not None:
        return value
    return _find_value_with_context(normalized_text, keywords, strong_keywords, allowed_set, min_value, max_value, allow_float)


def _extract_rating_and_score(text: str) -> tuple[Optional[int], Optional[float]]:
    normalized = _normalize_response_text(text)
    rating_value = _extract_scale_value(
        normalized_text=normalized,
        patterns=_RATING_PATTERNS,
        indicator_terms=_RATING_LINE_INDICATORS,
        keywords=_RATING_KEYWORDS,
        strong_keywords=_RATING_STRONG_KEYWORDS,
        allowed_set={0, 1, 2},
        allow_float=False,
    )
    score_value = _extract_scale_value(
        normalized_text=normalized,
        patterns=_SCORE_PATTERNS,
        indicator_terms=_SCORE_LINE_INDICATORS,
        keywords=_SCORE_KEYWORDS,
        strong_keywords=_SCORE_STRONG_KEYWORDS,
        min_value=1.0,
        max_value=10.0,
        allow_float=True,
    )

    rating = int(rating_value) if rating_value is not None else None
    score: Optional[float]
    if score_value is None:
        score = None
    else:
        if abs(score_value - round(score_value)) < 1e-6:
            score = float(int(round(score_value)))
        else:
            score = round(score_value, 3)
    return rating, score


def _first_sentences(text: str, max_sentences: int = 3) -> str:
    parts = re.split(r"(?<=[.!?])\s+", text.strip())
    return " ".join(parts[:max_sentences]).strip()


def _write_csv(path: str, rows: List[Dict[str, Any]]) -> None:
    if not rows:
        return
    fieldnames = [
        "task",
        "space",
        "attribute",
        "iteration",
        "run",
        "model",
        "rating_0_2",
        "score_1_10",
        "justification",
        "raw",
    ]
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for r in rows:
            writer.writerow(r)


def _write_xlsx(path: str, rows: List[Dict[str, Any]], meta: Dict[str, Any], prompts: List[Dict[str, Any]]) -> None:
    try:
        from openpyxl import Workbook
    except Exception as e:
        raise SystemExit(
            "Missing dependency: openpyxl. Install with `pip install openpyxl` or `pip install -r requirements.txt`."
        ) from e

    wb = Workbook()

    # Meta sheet
    ws_meta = wb.active
    ws_meta.title = "meta"
    for i, (k, v) in enumerate(meta.items(), start=1):
        ws_meta.cell(row=i, column=1, value=str(k))
        ws_meta.cell(row=i, column=2, value=v)
    _autowidth(ws_meta)

    # Prompts sheet
    ws_prompts = wb.create_sheet("prompts")
    ws_prompts.append(["attribute", "iteration", "prompt"])
    for rec in prompts:
        ws_prompts.append([rec.get("attribute"), rec.get("iteration"), rec.get("prompt")])
    _autowidth(ws_prompts)

    # One sheet per attribute with per-iteration blocks and summaries
    headers = [
        "task",
        "space",
        "attribute",
        "iteration",
        "run",
        "model",
        "rating_0_2",
        "score_1_10",
        "justification",
        "raw",
    ]
    by_attr: Dict[str, List[Dict[str, Any]]] = {}
    for r in rows:
        by_attr.setdefault(str(r.get("attribute", "")).strip() or "(unknown)", []).append(r)

    for attr, items in by_attr.items():
        title = _sanitize_sheet_title(attr or "attribute")
        ws = wb.create_sheet(title)
        ws.append(headers)
        # Group rows by iteration and write each block followed by its summaries
        by_iter: Dict[int, List[Dict[str, Any]]] = {}
        for r in items:
            try:
                it = int(r.get("iteration")) if r.get("iteration") is not None else None
            except Exception:
                it = None
            if it is None:
                # If iteration is missing, place into a pseudo-iteration 0 at the top
                by_iter.setdefault(0, []).append(r)
            else:
                by_iter.setdefault(it, []).append(r)

        for it in sorted(by_iter.keys()):
            block = by_iter[it]
            # Rows for this iteration
            for r in block:
                ws.append([r.get(h, "") for h in headers])

            # Compute stats for this iteration
            ratings = [v for v in (rr.get("rating_0_2") for rr in block) if isinstance(v, (int, float))]
            scores = [v for v in (rr.get("score_1_10") for rr in block) if isinstance(v, (int, float))]

            mean_score = round(stats.fmean(scores), 6) if scores else None
            var_score = None
            if len(scores) >= 2:
                try:
                    var_score = round(stats.variance(scores), 6)
                except Exception:
                    var_score = None

            med_rating = stats.median(ratings) if ratings else None
            mad_rating = None
            if ratings:
                try:
                    med = stats.median(ratings)
                    mad_rating = stats.median([abs(x - med) for x in ratings])
                except Exception:
                    mad_rating = None

            # Append a blank row then two summary rows for this iteration
            ws.append([""] * len(headers))
            row_mean_median = [""] * len(headers)
            row_mean_median[headers.index("rating_0_2")] = med_rating
            row_mean_median[headers.index("score_1_10")] = mean_score
            ws.append(row_mean_median)

            row_var_disp = [""] * len(headers)
            row_var_disp[headers.index("rating_0_2")] = mad_rating
            row_var_disp[headers.index("score_1_10")] = var_score
            ws.append(row_var_disp)

            # Spacer row between iterations
            ws.append([""] * len(headers))

        _autowidth(ws)

    # Save workbook
    wb.save(path)


def _sanitize_sheet_title(name: str) -> str:
    # Excel sheet title constraints
    invalid = set('[]:*?/\\')
    clean = ''.join(ch for ch in name if ch not in invalid)
    if not clean:
        clean = "sheet"
    return clean[:31]


def _autowidth(ws) -> None:
    try:
        for col in ws.columns:
            max_len = 0
            col_letter = getattr(col[0], "column_letter", None)
            for cell in col:
                try:
                    val = cell.value
                    if val is None:
                        continue
                    max_len = max(max_len, len(str(val)))
                except Exception:
                    pass
            if col_letter:
                ws.column_dimensions[col_letter].width = min(max_len + 2, 80)
    except Exception:
        # Best-effort; ignore any autowidth issues
        pass


if __name__ == "__main__":
    raise SystemExit(main())
