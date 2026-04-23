"""
Author: Hoda Alem
Date: November 2025
Context: Part of PhD research for a PhD in Environmental Design and Planning
         at Myers-Lawson School of Construction, Virginia Tech, Blacksburg, VA.
Purpose: Aggregate per-space result workbooks into CSV and a consolidated
         workbook for plotting and analysis; join with Expert ground truths.
Features:
- Parses per-run rows from results_*.xlsx and recomputes per-iteration stats.
- Computes 0â€“2 median, Q1/Q3, IQR and 1â€“10 mean, standard deviation.
- Reads Expert sheet (0â€“2 and 1â€“10) from the prompts workbook.
- Exports combined_summary.xlsx with a sheet per attribute and rows per space
  (including Expert rows), used by plotting utilities.
- Optional CSV export of per-iteration summaries and a listing mode.
"""

import os
import sys
import glob
import math
import argparse
import statistics as stats
import re
from pathlib import Path
from typing import Dict, List, Any, Tuple, DefaultDict, Optional


def _norm_key(value: Any) -> str:
    """Normalize labels for case-insensitive, whitespace-insensitive matching."""
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _match_space_dict(mapping: Dict[str, Dict[str, Any]], space: str) -> Optional[Dict[str, Any]]:
    """Return the attribute dict for the given space (exact or normalized match)."""
    if space in mapping:
        return mapping[space]
    norm = _norm_key(space)
    for name, attrs in mapping.items():
        if _norm_key(name) == norm:
            return attrs
    return None


def _match_attr_entry(mapping: Dict[str, Any], attribute: str) -> Optional[Dict[str, Any]]:
    """Return the expert entry for the given attribute (exact or normalized match)."""
    if attribute in mapping:
        return mapping[attribute]
    norm = _norm_key(attribute)
    for name, entry in mapping.items():
        if _norm_key(name) == norm:
            return entry
    return None


def _find_expert_entry(
    expert_map: Dict[str, Dict[str, Dict[str, Any]]],
    space: str,
    attribute: str,
) -> Optional[Dict[str, Any]]:
    """Lookup expert rating/score by matching on normalized labels."""
    space_dict = _match_space_dict(expert_map, space)
    if not space_dict:
        return None
    return _match_attr_entry(space_dict, attribute)


def _load_wb(path: Path):
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:
        raise SystemExit(
            "Missing dependency: openpyxl. Install with `pip install openpyxl` or `pip install -r requirements.txt`."
        ) from e
    return load_workbook(filename=str(path), read_only=True, data_only=True)


def _headers_from_sheet(ws) -> Dict[str, int]:
    row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers: Dict[str, int] = {}
    for idx, val in enumerate(row):
        if isinstance(val, str):
            headers[val.strip()] = idx
    return headers


def _is_number(x: Any) -> bool:
    return isinstance(x, (int, float)) and not (isinstance(x, float) and math.isnan(x))


def _collect_data_rows(ws) -> List[Dict[str, Any]]:
    headers = _headers_from_sheet(ws)
    data: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        rec: Dict[str, Any] = {}
        for k, i in headers.items():
            rec[k] = row[i] if i < len(row) else None
        # Keep only true per-run rows (have a run index)
        run_val = rec.get("run")
        if _is_number(run_val):
            data.append(rec)
    return data


def _summarize_block(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    ratings = [r.get("rating_0_2") for r in rows if _is_number(r.get("rating_0_2"))]
    scores = [r.get("score_1_10") for r in rows if _is_number(r.get("score_1_10"))]

    # Convert to float for consistency
    ratings = [float(x) for x in ratings]
    scores = [float(x) for x in scores]

    out: Dict[str, Any] = {
        "n_rows": len(rows),
        "median_rating_0_2": (stats.median(ratings) if ratings else None),
        "mad_rating_0_2": None,
        "mean_score_1_10": (round(stats.fmean(scores), 6) if scores else None),
        "std_score_1_10": None,
    }
    if ratings:
        try:
            med = stats.median(ratings)
            out["mad_rating_0_2"] = stats.median([abs(x - med) for x in ratings])
        except Exception:
            out["mad_rating_0_2"] = None
    if len(scores) >= 2:
        try:
            out["std_score_1_10"] = round(stats.stdev(scores), 6)
        except Exception:
            out["std_score_1_10"] = None
    return out


def list_files(pattern: str) -> List[Path]:
    return [Path(p) for p in sorted(glob.glob(pattern)) if p.lower().endswith(".xlsx")]


def analyze_file(path: Path) -> Dict[str, Any]:
    """Parse a single per-space results workbook.

    Returns a dict with 'meta' and 'per_attribute' summaries computed from per-run rows.
    """
    wb = _load_wb(path)
    meta = {}
    if "meta" in wb.sheetnames:
        ws_meta = wb["meta"]
        for r in ws_meta.iter_rows(min_row=1, values_only=True):
            if r and r[0]:
                meta[str(r[0])] = r[1] if len(r) > 1 else None

    per_attr: List[Dict[str, Any]] = []
    for sheet in wb.sheetnames:
        if sheet in {"meta", "prompts"}:
            continue
        ws = wb[sheet]
        data = _collect_data_rows(ws)
        # Group by iteration
        by_iter: Dict[Any, List[Dict[str, Any]]] = {}
        for rec in data:
            by_iter.setdefault(rec.get("iteration"), []).append(rec)
        for it, rows in sorted(by_iter.items(), key=lambda kv: (kv[0] is None, kv[0])):
            s = _summarize_block(rows)
            per_attr.append({
                "space": meta.get("space") or Path(path).stem,
                "attribute": sheet,
                "iteration": it,
                **s,
                "file": str(path),
            })

    return {"meta": meta, "per_attribute": per_attr}


def export_summary_csv(paths: List[Path], out_csv: Path) -> None:
    import csv
    fieldnames = [
        "file",
        "space",
        "attribute",
        "iteration",
        "n_rows",
        "median_rating_0_2",
        "mad_rating_0_2",
        "mean_score_1_10",
        "std_score_1_10",
    ]
    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for p in paths:
            result = analyze_file(p)
            for row in result["per_attribute"]:
                w.writerow({k: row.get(k) for k in fieldnames})


def main(argv: Optional[List[str]] = None) -> int:
    argv = sys.argv[1:] if argv is None else argv

    ap = argparse.ArgumentParser(description="Aggregate results_*.xlsx files produced by biophilic_eval.py")
    ap.add_argument("--glob", default="results_*.xlsx", help="Glob to match result workbooks")
    ap.add_argument("--list", action="store_true", help="List matched files and exit")
    ap.add_argument("--summary-csv", metavar="OUT", help="Export per-attribute, per-iteration summary to CSV")
    ap.add_argument("--consolidated-xlsx", metavar="OUT", help="Write one workbook: one sheet per space with 16 iteration columns per attribute")
    ap.add_argument("--xlsx", default="BD GPT Prompts.xlsx", help="Path to workbook containing Expert sheet")
    args = ap.parse_args(argv)

    files = list_files(args.glob)
    if args.list:
        if not files:
            print("No files matched.")
            return 0
        for p in files:
            info = analyze_file(p)
            space = info["meta"].get("space") or p.stem
            attrs = sorted({row["attribute"] for row in info["per_attribute"]})
            iters = sorted({row["iteration"] for row in info["per_attribute"]})
            print(f"- {p}: space={space}, attributes={len(attrs)} ({', '.join(attrs)}), iterations={iters}")
        return 0

    if args.summary_csv:
        if not files:
            print("No files matched.", file=sys.stderr)
            return 2
        export_summary_csv(files, Path(args.summary_csv))
        print(f"Wrote {args.summary_csv}")
        return 0

    if args.consolidated_xlsx:
        if not files:
            print("No files matched.", file=sys.stderr)
            return 2
        export_consolidated_xlsx(files, Path(args.consolidated_xlsx), expert_xlsx=Path(args.xlsx))
        print(f"Wrote {args.consolidated_xlsx}")
        return 0

    # Default behavior: just list
    return main(["--list"])  # type: ignore


def _iqr(values: List[float]) -> Optional[float]:
    try:
        if len(values) < 2:
            return None
        qs = stats.quantiles(values, n=4, method="inclusive")
        # qs -> [Q1, Q2 (median), Q3]
        return float(qs[2] - qs[0])
    except Exception:
        return None

def _quartiles(values: List[float]) -> Tuple[Optional[float], Optional[float]]:
    try:
        if len(values) < 2:
            return None, None
        qs = stats.quantiles(values, n=4, method="inclusive")
        # qs -> [Q1, Q2 (median), Q3]
        return float(qs[0]), float(qs[2])
    except Exception:
        return None, None


def _gather_by_space(paths: List[Path]) -> Dict[str, Dict[str, Dict[int, Dict[str, Any]]]]:
    """Return nested dict: space -> attribute -> iteration -> summary dict.

    Summary dict contains keys: median_rating_0_2, iqr_rating_0_2, mean_score_1_10, var_score_1_10.
    """
    out: Dict[str, Dict[str, Dict[int, Dict[str, Any]]]] = {}
    for p in paths:
        info = analyze_file(p)
        space = str(info["meta"].get("space") or p.stem)
        space_map = out.setdefault(space, {})
        # Recompute stats from per-run data (analyze_file already summarized per-iteration).
        # We recompute here to add IQR; we need raw rows, but analyze_file didn't return them.
        # So reload sheets and recompute directly.
        wb = _load_wb(p)
        for sheet in wb.sheetnames:
            if sheet in {"meta", "prompts"}:
                continue
            ws = wb[sheet]
            data_rows = _collect_data_rows(ws)
            attr_label: Optional[str] = None
            for rec in data_rows:
                raw_attr = rec.get("attribute")
                if isinstance(raw_attr, str):
                    stripped = raw_attr.strip()
                    if stripped:
                        attr_label = stripped
                        break
            display_name = attr_label or sheet
            by_iter: Dict[int, List[Dict[str, Any]]] = {}
            for rec in data_rows:
                it = rec.get("iteration")
                try:
                    it = int(it) if it is not None else None
                except Exception:
                    it = None
                if it is None:
                    continue
                by_iter.setdefault(it, []).append(rec)

            attr_map = space_map.setdefault(display_name, {})
            for it, rows in by_iter.items():
                ratings = [float(r.get("rating_0_2")) for r in rows if _is_number(r.get("rating_0_2"))]
                scores = [float(r.get("score_1_10")) for r in rows if _is_number(r.get("score_1_10"))]
                med_rating = stats.median(ratings) if ratings else None
                q1_rating, q3_rating = _quartiles(ratings) if ratings else (None, None)
                iqr_rating = (q3_rating - q1_rating) if (q1_rating is not None and q3_rating is not None) else _iqr(ratings) if ratings else None
                mean_score = round(stats.fmean(scores), 6) if scores else None
                std_score = None
                if len(scores) >= 2:
                    try:
                        std_score = round(stats.stdev(scores), 6)
                    except Exception:
                        std_score = None
                attr_map[it] = {
                    "median_rating_0_2": med_rating,
                    "q1_rating_0_2": q1_rating,
                    "q3_rating_0_2": q3_rating,
                    "iqr_rating_0_2": iqr_rating,
                    "mean_score_1_10": mean_score,
                    "std_score_1_10": std_score,
                }
    return out


def export_consolidated_xlsx(paths: List[Path], out_xlsx: Path, expert_xlsx: Path = Path("BD GPT Prompts.xlsx")) -> None:
    """Create one workbook with a sheet per attribute and rows per built space.

    Columns per iteration: median_rating_0_2, q1_rating_0_2, q3_rating_0_2, iqr_rating_0_2, mean_score_1_10, std_score_1_10
    Across iterations 1..4 => 24 columns. Appends expert ground truth values under median/mean columns; dispersion columns left blank for experts.
    """
    try:
        from openpyxl import Workbook
    except Exception as e:
        raise SystemExit(
            "Missing dependency: openpyxl. Install with `pip install openpyxl` or `pip install -r requirements.txt`."
        ) from e

    by_space = _gather_by_space(paths)
    # Load expert truths if available
    expert_map: Dict[str, Dict[str, Dict[str, Any]]] = {}
    try:
        if expert_xlsx.exists():
            expert_map = _read_expert_truths(expert_xlsx)
    except Exception as e:
        print(f"Warning: failed to read experts from {expert_xlsx}: {e}")

    all_attrs: List[str] = sorted({attr for space in by_space.values() for attr in space.keys()})
    all_spaces: List[str] = sorted(by_space.keys())
    all_iters: List[int] = sorted({it for space in by_space.values() for attr_map in space.values() for it in attr_map.keys()})

    wb = Workbook()
    ws_idx = wb.active
    ws_idx.title = "index"
    ws_idx.append(["attributes", len(all_attrs)])
    ws_idx.append(["spaces", len(all_spaces)])

    def sheet_title(name: str) -> str:
        invalid = set('[]:*?/\\')
        clean = ''.join(ch for ch in name if ch not in invalid)
        return (clean or "sheet")[:31]

    for attribute in all_attrs:
        ws = wb.create_sheet(sheet_title(attribute))
        headers: List[str] = ["space"]
        for it in all_iters:
            headers.extend([
                f"it{it}_median_rating_0_2",
                f"it{it}_q1_rating_0_2",
                f"it{it}_q3_rating_0_2",
                f"it{it}_iqr_rating_0_2",
                f"it{it}_mean_score_1_10",
                f"it{it}_std_score_1_10",
            ])
        ws.append(headers)

        for space in all_spaces:
            iter_map = by_space.get(space, {}).get(attribute, {})
            # ChatGPT row
            row: List[Any] = [space]
            for it in all_iters:
                s = iter_map.get(it, {})
                row.extend([
                    s.get("median_rating_0_2"),
                    s.get("q1_rating_0_2"),
                    s.get("q3_rating_0_2"),
                    s.get("iqr_rating_0_2"),
                    s.get("mean_score_1_10"),
                    s.get("std_score_1_10"),
                ])
            ws.append(row)

            # Expert row (second row for this space): repeat expert values under value columns; dispersion blank
            exp_entry = _find_expert_entry(expert_map, space, attribute)
            exp_row: List[Any] = [f"{space} (Expert)"]
            rating02 = None if not exp_entry else exp_entry.get("rating_0_2")
            score10 = None if not exp_entry else exp_entry.get("score_1_10")
            for it in all_iters:
                exp_row.extend([
                    rating02,
                    None,
                    None,
                    None,
                    score10,
                    None,
                ])
            ws.append(exp_row)

        try:
            for col in ws.columns:
                max_len = 0
                col_letter = getattr(col[0], "column_letter", None)
                for cell in col:
                    val = cell.value
                    if val is None:
                        continue
                    max_len = max(max_len, len(str(val)))
                if col_letter:
                    ws.column_dimensions[col_letter].width = min(max_len + 2, 80)
        except Exception:
            pass

    wb.save(out_xlsx)


def _read_expert_truths(path: Path, sheet_name: str = "Expert") -> Dict[str, Dict[str, Dict[str, Any]]]:
    """Read expert ground-truth ratings.

    Expected layout (as specified):
    - Row 1: built space names (repeated for two adjacent columns per space)
    - Row 2: rating tag per column (e.g., '0-2' or '1-10')
    - Column 1: index (ignored)
    - Column 2: attribute names
    - Data starts at row 3, column 3

    Returns nested mapping: space -> attribute -> {rating_0_2, score_1_10}
    """
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:
        raise SystemExit(
            "Missing dependency: openpyxl. Install with `pip install openpyxl` or `pip install -r requirements.txt`."
        ) from e

    if not path.exists() or not path.is_file():
        raise FileNotFoundError(f"Workbook not found: {path}")

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]

    # Read header rows
    header1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header2 = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))

    def norm_space(h: Any) -> str:
        s = str(h or "").strip()
        return s

    def norm_tag(h: Any) -> str:
        s = str(h or "").lower()
        # Normalize dashes and the word 'to'
        s = s.replace("â€“", "-")  # en dash
        s = s.replace("â€”", "-")  # em dash
        s = s.replace("to", "-")
        # Remove all whitespace including non-breaking spaces
        s = re.sub(r"\s+", "", s, flags=re.UNICODE)
        # Map alternate label for categorical rating
        if s in {"0,1,2", "012"}:
            return "0-2"
        return s
        s = s.strip().rstrip("-:")
        return s

    # Determine attribute column and space columns by fixed positions
    # attr names in column 2; starting at column 3, for each space: col c = 0-2 rating, col c+1 = 1-10 rating
    attr_col = 2
    start_col = 3
    pairs: Dict[str, Dict[str, int]] = {}
    c = start_col
    while c <= ws.max_column:
        space = norm_space(header1[c - 1] if c - 1 < len(header1) else None)
        if space:
            pairs.setdefault(space, {})['0-2'] = c
            if c + 1 <= ws.max_column:
                pairs[space]['1-10'] = c + 1
        c += 2

    result: Dict[str, Dict[str, Dict[str, Any]]] = {}
    # Iterate data rows starting from row 2
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row is None:
            continue
        attr_raw = row[attr_col - 1] if attr_col - 1 < len(row) else None
        if not attr_raw:
            continue
        attribute = str(attr_raw).strip()
        for space, cols in pairs.items():
            c02 = cols.get('0-2')
            c10 = cols.get('1-10')
            v02 = row[c02 - 1] if c02 and c02 - 1 < len(row) else None
            v10 = row[c10 - 1] if c10 and c10 - 1 < len(row) else None
            # Coerce to numbers when possible
            def to_num(x):
                try:
                    if x is None or (isinstance(x, float) and math.isnan(x)):
                        return None
                    if isinstance(x, (int, float)):
                        return float(x)
                    s = str(x).strip()
                    if not s:
                        return None
                    return float(s)
                except Exception:
                    return None

            rating_0_2 = to_num(v02)
            score_1_10 = to_num(v10)

            result.setdefault(space, {}).setdefault(attribute, {})
            result[space][attribute] = {
                "rating_0_2": rating_0_2,
                "score_1_10": score_1_10,
            }

    return result


if __name__ == "__main__":
    raise SystemExit(main())
