"""
Author: Hoda Alem
Date: November 2025
Context: Part of PhD research for a PhD in Environmental Design and Planning
         at Myers-Lawson School of Construction, Virginia Tech, Blacksburg, VA.

Faceted Expert vs Model scatter by attribute, one panel per attribute.

Outputs two figures:
- out02: 0â€“2 scale, per-attribute panels (x=Expert, y=Model), normal jitter.
- out10: 1â€“10 scale, per-attribute panels (x=Expert, y=Model), normal jitter.

Color encoding selectable: iteration (default) or space.
"""

import argparse
import math
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple


def _normalize_tokens(s: str) -> List[str]:
    import re
    s = (s or "").lower()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return [t for t in s.split() if t]


def _norm_key(s: str) -> str:
    return " ".join(_normalize_tokens(s or ""))


def _load_wb(path: Path):
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:
        raise SystemExit("Missing dependency: openpyxl. Install with `pip install openpyxl`.") from e
    return load_workbook(filename=str(path), read_only=True, data_only=True)


def _read_expert_truths(path: Path, sheet_name: str = "Expert") -> Dict[str, Dict[str, Dict[str, float]]]:
    wb = _load_wb(path)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"Sheet '{sheet_name}' not found in {path}")
    ws = wb[sheet_name]
    header1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    start_col = 3
    space_cols: List[Tuple[str, int]] = []
    c = start_col
    while c <= ws.max_column:
        space = str(header1[c - 1] or '').strip()
        if space:
            space_cols.append((space, c))
        c += 2
    out: Dict[str, Dict[str, Dict[str, float]]] = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        attr_raw = row[1] if len(row) > 1 else None
        if not attr_raw:
            continue
        attribute = str(attr_raw).strip()
        for space, c02 in space_cols:
            v02 = row[c02 - 1] if c02 - 1 < len(row) else None
            v10 = row[c02] if c02 < len(row) else None  # c02+1
            def to_f(x):
                try:
                    return float(x) if isinstance(x, (int, float)) else float(str(x).strip()) if x else None
                except Exception:
                    return None
            val02 = to_f(v02)
            val10 = to_f(v10)
            if val02 is None and val10 is None:
                continue
            out.setdefault(space, {}).setdefault(attribute, {})
            if val02 is not None:
                out[space][attribute]['0-2'] = val02
            if val10 is not None:
                out[space][attribute]['1-10'] = val10
    return out


def _headers(ws) -> Dict[str, int]:
    row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    idx: Dict[str, int] = {}
    for i, v in enumerate(row):
        if isinstance(v, str):
            idx[v.strip()] = i
    return idx


def _is_number(x: Any) -> bool:
    import math as _m
    return isinstance(x, (int, float)) and not (isinstance(x, float) and _m.isnan(x))


def _collect_points(paths: List[Path]) -> List[Dict[str, Any]]:
    pts: List[Dict[str, Any]] = []
    for p in paths:
        wb = _load_wb(p)
        space_name: Optional[str] = None
        if 'meta' in wb.sheetnames:
            try:
                ws_meta = wb['meta']
                for r in ws_meta.iter_rows(min_row=1, values_only=True):
                    if r and str(r[0]).strip().lower() == 'space':
                        space_name = str(r[1]).strip() if len(r) > 1 and r[1] is not None else None
                        break
            except Exception:
                pass
        for sheet in wb.sheetnames:
            if sheet in {"meta", "prompts"}:
                continue
            ws = wb[sheet]
            idx = _headers(ws)
            if not idx:
                continue
            def val(row, k):
                i = idx.get(k)
                return row[i] if i is not None and i < len(row) else None
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row is None:
                    continue
                rv = val(row, 'run'); it = val(row, 'iteration')
                if not _is_number(rv) or not _is_number(it):
                    continue
                r02 = val(row, 'rating_0_2'); s10 = val(row, 'score_1_10')
                if not _is_number(r02) or not _is_number(s10):
                    continue
                attr = val(row, 'attribute') or sheet
                sp = val(row, 'space') or space_name or p.stem
                pts.append({
                    'space': str(sp),
                    'attribute': str(attr),
                    'iteration': int(it),
                    'rating_0_2': float(r02),
                    'score_1_10': float(s10),
                })
    return pts


def _match_expert(expert: Dict[str, Dict[str, Dict[str, float]]], space: str, attribute: str) -> Optional[Dict[str, float]]:
    if space in expert and attribute in expert[space]:
        return expert[space][attribute]
    s_norm = _norm_key(space); a_norm = _norm_key(attribute)
    for s, amap in expert.items():
        if _norm_key(s) != s_norm:
            continue
        for a, vals in amap.items():
            if _norm_key(a) == a_norm:
                return vals
    return None


def _build_pairs(points: List[Dict[str, Any]], expert: Dict[str, Dict[str, Dict[str, float]]]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    p02: List[Dict[str, Any]] = []
    p10: List[Dict[str, Any]] = []
    for pt in points:
        ex = _match_expert(expert, pt['space'], pt['attribute'])
        if not ex:
            continue
        if '0-2' in ex:
            p02.append({'attribute': pt['attribute'], 'space': pt['space'], 'iteration': pt['iteration'], 'x': ex['0-2'], 'y': pt['rating_0_2']})
        if '1-10' in ex:
            p10.append({'attribute': pt['attribute'], 'space': pt['space'], 'iteration': pt['iteration'], 'x': ex['1-10'], 'y': pt['score_1_10']})
    return p02, p10


def _colors_for(keys: List[str]):
    import matplotlib.pyplot as plt
    cmap = plt.get_cmap('tab20')
    return {k: cmap((i % 20) / 20.0) for i, k in enumerate(sorted(set(keys)))}


def _pearson_r(xs: List[float], ys: List[float]) -> Optional[float]:
    if len(xs) < 2:
        return None
    mean_x = sum(xs) / len(xs)
    mean_y = sum(ys) / len(ys)
    num = sum((x - mean_x) * (y - mean_y) for x, y in zip(xs, ys))
    den_x = sum((x - mean_x) ** 2 for x in xs)
    den_y = sum((y - mean_y) ** 2 for y in ys)
    den = math.sqrt(den_x * den_y)
    if den == 0:
        return None
    return num / den


def _group_points(points: List[Dict[str, Any]]) -> Dict[Tuple[str, str, int], Dict[str, List[float]]]:
    groups: Dict[Tuple[str, str, int], Dict[str, List[float]]] = {}
    for pt in points:
        key = (pt['space'], pt['attribute'], int(pt['iteration']))
        g = groups.setdefault(key, {'r02': [], 's10': []})
        # 'x' and 'y' present in built pairs
        if 'x' in pt and 'y' in pt:
            # Decide which list to use based on scale by inspecting values
            # We'll let caller decide mapping; leave both empty here
            pass
    return groups


def _quantiles(values: List[float]) -> Tuple[Optional[float], Optional[float]]:
    try:
        import statistics as stats
        if len(values) < 2:
            return None, None
        qs = stats.quantiles(values, n=4, method="inclusive")
        return float(qs[0]), float(qs[2])
    except Exception:
        return None, None


def _aggregate_pairs_attr(
    points_raw: List[Dict[str, Any]],
    expert: Dict[str, Dict[str, Dict[str, float]]],
    iter_filter: str,
    iter_list: Optional[List[int]],
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    import statistics as stats
    # Build per-run pairs first
    p02, p10 = _build_pairs(points_raw, expert)
    # Group by (space, attribute, iteration)
    from collections import defaultdict
    g02: Dict[Tuple[str, str, int], List[float]] = defaultdict(list)
    g10: Dict[Tuple[str, str, int], List[float]] = defaultdict(list)
    for p in p02:
        g02[(p['space'], p['attribute'], int(p['iteration']))].append(float(p['y']))
    for p in p10:
        g10[(p['space'], p['attribute'], int(p['iteration']))].append(float(p['y']))

    # Determine iterations to include
    all_iters = sorted({it for (_, _, it) in set(list(g02.keys()) + list(g10.keys()))})
    if iter_list:
        include_iters = set(int(i) for i in iter_list)
    elif iter_filter == 'last' and all_iters:
        include_iters = {max(all_iters)}
    else:
        include_iters = set(all_iters)

    out02: List[Dict[str, Any]] = []
    out10: List[Dict[str, Any]] = []
    # 0-2 aggregates: median with Q1-Q3 asymmetric bars
    for (space, attr, it), vals in g02.items():
        if it not in include_iters:
            continue
        ex = _match_expert(expert, space, attr)
        if not ex or '0-2' not in ex:
            continue
        try:
            med = stats.median(vals)
        except Exception:
            med = None
        if med is None:
            continue
        q1, q3 = _quantiles(vals)
        low = max(0.0, med - q1) if q1 is not None else 0.0
        high = max(0.0, q3 - med) if q3 is not None else 0.0
        out02.append({'attribute': attr, 'space': space, 'iteration': it, 'x': float(ex['0-2']), 'y': float(med), 'yerr_low': float(low), 'yerr_high': float(high)})

    # 1-10 aggregates: mean with std (symmetric)
    for (space, attr, it), vals in g10.items():
        if it not in include_iters:
            continue
        ex = _match_expert(expert, space, attr)
        if not ex or '1-10' not in ex:
            continue
        try:
            mean = float(stats.fmean(vals))
        except Exception:
            mean = None
        if mean is None:
            continue
        std = 0.0
        if len(vals) >= 2:
            try:
                std = float(stats.stdev(vals))
            except Exception:
                std = 0.0
        out10.append({'attribute': attr, 'space': space, 'iteration': it, 'x': float(ex['1-10']), 'y': float(mean), 'yerr': float(std)})

    return out02, out10


def _facet_plot(pairs: List[Dict[str, Any]], out: Path, scale: str, color_by: str, jitter_x: float, jitter_y: float, alpha: float, size: float, seed: Optional[int], dpi: int, ncols: int = 3, aggregate: bool = False) -> None:
    if not pairs:
        raise SystemExit('No data to plot for this scale.')
    try:
        import matplotlib.pyplot as plt  # type: ignore
        import numpy as np  # type: ignore
    except Exception as e:
        raise SystemExit("Missing dependency: matplotlib (and numpy). Install with `pip install matplotlib numpy`.") from e

    if seed is not None:
        np.random.seed(int(seed))

    # facet by attribute
    attrs = sorted({p['attribute'] for p in pairs})
    # colors by iteration or space
    if color_by == 'space':
        color_keys = sorted({p['space'] for p in pairs})
    else:
        color_keys = sorted({str(p['iteration']) for p in pairs})
    c_map = _colors_for(color_keys)

    n = len(attrs)
    ncols = max(1, ncols)
    nrows = (n + ncols - 1) // ncols
    fig, axes = plt.subplots(nrows, ncols, figsize=(4.2 * ncols, 3.6 * nrows), sharex=False, sharey=False)
    if nrows == 1 and ncols == 1:
        axes = [[axes]]
    elif nrows == 1:
        axes = [list(axes)]
    elif ncols == 1:
        axes = [[ax] for ax in axes]

    # axis limits
    if scale == '0-2':
        xlim = (-0.5, 2.5); ylim = (-0.5, 2.5); xticks = [0, 1, 2]; yticks = [0, 1, 2]
        xlab = 'Expert Ratings (0,1,2)'; ylab = 'ChatGPT Ratings (0,1,2)'
    else:
        xlim = (0.5, 10.5); ylim = (0.5, 10.5); xticks = list(range(1, 11)); yticks = list(range(1, 11))
        xlab = 'Expert Ratings (1â€“10)'; ylab = 'ChatGPT Ratings (1â€“10)'

    # plot each attribute panel
    for i, a in enumerate(attrs):
        r = i // ncols; c = i % ncols
        ax = axes[r][c]
        data = [p for p in pairs if p['attribute'] == a]
        if not aggregate:
            xs = []; ys = []; cols = []
            for p in data:
                xs.append(float(p['x'])); ys.append(float(p['y']))
                key = p['space'] if color_by == 'space' else str(p['iteration'])
                cols.append(c_map.get(key))
            xs = np.array(xs, dtype=float); ys = np.array(ys, dtype=float)
            if jitter_x > 0: xs = xs + np.random.normal(0.0, jitter_x, size=xs.shape)
            if jitter_y > 0: ys = ys + np.random.normal(0.0, jitter_y, size=ys.shape)
            ax.scatter(xs, ys, s=size, alpha=alpha, c=cols, edgecolors='none')
        else:
            # aggregated: points with error bars per point
            xs = []; ys = []; cols = []
            yerrs: List[Tuple[float, float]] = []
            for p in data:
                xs.append(float(p['x'])); ys.append(float(p['y']))
                key = p['space'] if color_by == 'space' else str(p['iteration'])
                cols.append(c_map.get(key))
                if 'yerr_low' in p and 'yerr_high' in p:
                    yerrs.append((float(p['yerr_low']), float(p['yerr_high'])))
                else:
                    sym = float(p.get('yerr', 0.0)); yerrs.append((sym, sym))
            xs = np.array(xs, dtype=float); ys = np.array(ys, dtype=float)
            # In aggregated mode, do not jitter model central values; allow optional expert jitter only
            if jitter_x > 0: xs = xs + np.random.normal(0.0, jitter_x, size=xs.shape)
            # Allow y-jitter of central values only for 0â€“2 panels to reduce overplotting
            if scale == '0-2' and jitter_y > 0:
                ys = ys + np.random.normal(0.0, jitter_y, size=ys.shape)
            alpha_pts = max(alpha, 0.8)
            ax.scatter(xs, ys, s=max(12.0, size), alpha=alpha_pts, c=cols, edgecolors='none')
            for i_pt in range(len(xs)):
                low, high = yerrs[i_pt]
                ax.errorbar([xs[i_pt]], [ys[i_pt]], yerr=[[low], [high]], fmt='none', ecolor=cols[i_pt], elinewidth=1.0, alpha=min(1.0, alpha + 0.2))
        # identity line
        lo = max(xlim[0], ylim[0]); hi = min(xlim[1], ylim[1])
        ax.plot([lo, hi], [lo, hi], color='black', lw=1.2, alpha=0.85)
        ax.set_xlim(*xlim); ax.set_ylim(*ylim)
        ax.set_xticks(xticks); ax.set_yticks(yticks)
        ax.set_title(a)
        ax.grid(True, alpha=0.2)
        r = _pearson_r([float(p['x']) for p in data], [float(p['y']) for p in data])
        r_txt = "Correlation: NA" if r is None else f"Correlation: {r:.2f}"
        label = f"({chr(ord('a') + i)}) {r_txt}"
        ax.text(0.02, 0.96, label, transform=ax.transAxes, ha='left', va='top', fontsize=10)

    # hide unused axes
    for j in range(n, nrows * ncols):
        r = j // ncols; c = j % ncols
        axes[r][c].set_visible(False)

    # shared labels
    fig.text(0.5, 0.04, xlab, ha='center')
    fig.text(0.04, 0.5, ylab, va='center', rotation='vertical')

    # legend inside bottom-right of last visible panel
    import matplotlib.pyplot as _plt
    handles = [
        _plt.Line2D([0], [0], marker='o', color='none', markerfacecolor=c_map[k], markersize=max(3, math.sqrt(size)), label=str(k))
        for k in color_keys
    ]
    # Find last visible axes
    last_ax = None
    for rr in range(nrows-1, -1, -1):
        for cc in range(ncols-1, -1, -1):
            if hasattr(axes[rr][cc], 'lines'):
                last_ax = axes[rr][cc]
                break
        if last_ax:
            break
    if last_ax:
        last_ax.legend(handles=handles, loc='lower right', fontsize=8, frameon=True, title=('Iteration' if color_by!='space' else 'Space'))

    fig.tight_layout(rect=(0.06, 0.06, 0.98, 0.98))
    fig.savefig(out, dpi=dpi)
    _plt.close(fig)


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description='Per-attribute Expert vs Model scatter panels for 0â€“2 and 1â€“10 scales')
    parser.add_argument('--glob', default='results_*.xlsx', help='Glob for per-space results files')
    parser.add_argument('--xlsx', default='BD GPT Prompts.xlsx', help='Workbook containing Expert sheet')
    parser.add_argument('--out02', default='expert_vs_model_by_attr_0to2.png', help='Output image for 0â€“2 panels')
    parser.add_argument('--out10', default='expert_vs_model_by_attr_1to10.png', help='Output image for 1â€“10 panels')
    parser.add_argument('--color-by', choices=['iteration', 'space'], default='iteration', help='Color encoding in panels')
    parser.add_argument('--jitter02', type=float, default=0.06, help='Std dev of normal jitter for 0â€“2 x')
    parser.add_argument('--y-jitter02', type=float, default=0.10, help='Std dev of normal jitter for 0â€“2 y')
    parser.add_argument('--jitter10', type=float, default=0.08, help='Std dev of normal jitter for 1â€“10 x')
    parser.add_argument('--y-jitter10', type=float, default=0.15, help='Std dev of normal jitter for 1â€“10 y')
    parser.add_argument('--alpha', type=float, default=0.5, help='Marker alpha')
    parser.add_argument('--size', type=float, default=16.0, help='Marker size (points^2)')
    parser.add_argument('--seed', type=int, help='Random seed for jitter')
    parser.add_argument('--dpi', type=int, default=180, help='Figure DPI')
    parser.add_argument('--ncols', type=int, default=3, help='Number of columns (e.g., 3 => 2 rows for 6 attributes)')
    parser.add_argument('--aggregate', action='store_true', help='Plot central value with error bars instead of per-run scatter')
    parser.add_argument('--iter-filter', choices=['all', 'last'], default='all', help='Iterations to include when aggregating')
    parser.add_argument('--iterations', nargs='*', type=int, help='Explicit iteration list when aggregating')
    args = parser.parse_args()

    import glob as _glob
    paths = [Path(p) for p in sorted(_glob.glob(args.glob)) if p.lower().endswith('.xlsx')]
    if not paths:
        raise SystemExit('No result files matched.')

    expert = _read_expert_truths(Path(args.xlsx))
    points = _collect_points(paths)
    if args.aggregate:
        agg02, agg10 = _aggregate_pairs_attr(points, expert, iter_filter=str(args.iter_filter), iter_list=args.iterations)
        _facet_plot(agg02, Path(args.out02), scale='0-2', color_by=str(args.color_by), jitter_x=float(args.jitter02), jitter_y=float(args.y_jitter02), alpha=float(args.alpha), size=float(args.size), seed=args.seed, dpi=int(args.dpi), ncols=int(args.ncols), aggregate=True)
        _facet_plot(agg10, Path(args.out10), scale='1-10', color_by=str(args.color_by), jitter_x=float(args.jitter10), jitter_y=float(args.y_jitter10), alpha=float(args.alpha), size=float(args.size), seed=args.seed, dpi=int(args.dpi), ncols=int(args.ncols), aggregate=True)
    else:
        pairs02, pairs10 = _build_pairs(points, expert)
        _facet_plot(pairs02, Path(args.out02), scale='0-2', color_by=str(args.color_by), jitter_x=float(args.jitter02), jitter_y=float(args.y_jitter02), alpha=float(args.alpha), size=float(args.size), seed=args.seed, dpi=int(args.dpi), ncols=int(args.ncols))
        _facet_plot(pairs10, Path(args.out10), scale='1-10', color_by=str(args.color_by), jitter_x=float(args.jitter10), jitter_y=float(args.y_jitter10), alpha=float(args.alpha), size=float(args.size), seed=args.seed, dpi=int(args.dpi), ncols=int(args.ncols))
    print(f'Saved {args.out02} and {args.out10}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())


def _pearson_r(xs: List[float], ys: List[float]) -> Optional[float]:
    if len(xs) < 2:
        return None
    mean_x = sum(xs) / len(xs)
    mean_y = sum(ys) / len(ys)
    num = sum((x - mean_x) * (y - mean_y) for x, y in zip(xs, ys))
    den_x = sum((x - mean_x) ** 2 for x in xs)
    den_y = sum((y - mean_y) ** 2 for y in ys)
    den = math.sqrt(den_x * den_y)
    if den == 0:
        return None
    return num / den
