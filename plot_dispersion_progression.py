"""
Author: Hoda Alem
Date: November 2025
Context: Part of PhD research for a PhD in Environmental Design and Planning
         at Myers-Lawson School of Construction, Virginia Tech, Blacksburg, VA.
Purpose: Visualize aggregate dispersion over prompt iterations using raw results:
         pooled MAD for 0-2 and pooled standard deviation for 1-10, by attribute.
Features:
- Reads per-run rows from all results_*.xlsx files.
- Computes per-iteration, per-attribute dispersion from pooled deviations.
- Produces a two-panel figure with grouped bars per attribute and an aggregate line.
"""

import argparse
import statistics as stats
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple


def _normalize_tokens(s: str) -> List[str]:
    import re
    s = (s or "").lower()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return [t for t in s.split() if t]


def _norm_key(s: str) -> str:
    return " ".join(_normalize_tokens(s))


def _load_wb(path: Path):
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:
        raise SystemExit(
            "Missing dependency: openpyxl. Install with `pip install openpyxl` or `pip install -r requirements.txt`."
        ) from e
    return load_workbook(filename=str(path), read_only=True, data_only=True)


def _headers_from_sheet(ws) -> Dict[str, int]:
    row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers: Dict[str, int] = {}
    for i, v in enumerate(row):
        if isinstance(v, str):
            headers[v.strip()] = i
    return headers


def _collect_data_rows(ws) -> List[Dict[str, Any]]:
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers: Dict[str, int] = {}
    for i, v in enumerate(header):
        if isinstance(v, str):
            headers[v.strip()] = i
    data: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        rec: Dict[str, Any] = {}
        for k, idx in headers.items():
            rec[k] = row[idx] if idx < len(row) else None
        if isinstance(rec.get("run"), (int, float)):
            data.append(rec)
    return data


def _to_number(val: Any) -> Optional[float]:
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return None
        try:
            return float(s)
        except Exception:
            return None
    return None


def _attributes_from_expert(path: Path, sheet_name: str = "Expert") -> List[str]:
    wb = _load_wb(path)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    attrs: List[str] = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        v = row[1] if len(row) > 1 else None
        if v:
            attrs.append(str(v).strip())
    return [a for a in attrs if a]


def compute_dispersion_by_attribute(paths: List[Path]) -> Tuple[Dict[int, Dict[str, float]], Dict[int, Dict[str, float]], List[int], Dict[int, float], Dict[int, float], Dict[int, Dict[str, int]], Dict[int, int]]:
    """Return (std_by_it_attr, mad_by_it_attr, iterations, agg_mad_by_it, agg_std_by_it, mad_counts_by_it_attr, mad_counts_by_it).

    For each space-attribute-iteration group:
    - MAD: compute per-space median, then absolute deviations for that space's runs.
      Pool deviations across spaces and take the median (pooled MAD).
    - Std: compute per-space mean, then deviations for that space's runs.
      Pool deviations across spaces and take the standard deviation (pooled std).
    """
    pooled_mad_devs: Dict[int, Dict[str, List[float]]] = {}
    pooled_std_devs: Dict[int, Dict[str, List[float]]] = {}
    pooled_mad_all: Dict[int, List[float]] = {}
    pooled_std_all: Dict[int, List[float]] = {}
    iterations_set: set[int] = set()

    for p in paths:
        wb = _load_wb(p)
        for sheet in wb.sheetnames:
            if sheet in {"meta", "prompts"}:
                continue
            ws = wb[sheet]
            rows = _collect_data_rows(ws)
            # Prefer attribute name from the data, fallback to sheet name
            attr = sheet
            for rec in rows:
                a = rec.get("attribute")
                if isinstance(a, str) and a.strip():
                    attr = a.strip()
                    break
            by_it: Dict[int, List[Dict[str, Any]]] = {}
            for rec in rows:
                try:
                    it = int(rec.get("iteration"))
                except Exception:
                    continue
                by_it.setdefault(it, []).append(rec)
            for it, group in by_it.items():
                iterations_set.add(it)
                ratings: List[float] = []
                for r in group:
                    v = _to_number(r.get("rating_0_2"))
                    if v is not None:
                        ratings.append(v)
                if ratings:
                    med = stats.median(ratings)
                    devs = [abs(x - med) for x in ratings]
                    pooled_mad_devs.setdefault(it, {}).setdefault(attr, []).extend(devs)
                    pooled_mad_all.setdefault(it, []).extend(devs)
                scores: List[float] = []
                for r in group:
                    v = _to_number(r.get("score_1_10"))
                    if v is not None:
                        scores.append(v)
                if scores:
                    mean = stats.mean(scores)
                    devs = [x - mean for x in scores]
                    pooled_std_devs.setdefault(it, {}).setdefault(attr, []).extend(devs)
                    pooled_std_all.setdefault(it, []).extend(devs)

    iterations = sorted(iterations_set)
    mad_by_it_attr: Dict[int, Dict[str, float]] = {it: {} for it in iterations}
    std_by_it_attr: Dict[int, Dict[str, float]] = {it: {} for it in iterations}

    for it in iterations:
        for attr, devs in pooled_mad_devs.get(it, {}).items():
            if devs:
                mad_by_it_attr[it][attr] = float(stats.mean(devs))
        for attr, devs in pooled_std_devs.get(it, {}).items():
            if len(devs) >= 2:
                std_by_it_attr[it][attr] = float(stats.stdev(devs))
            elif devs:
                std_by_it_attr[it][attr] = 0.0

    agg_mad_by_it: Dict[int, float] = {}
    agg_std_by_it: Dict[int, float] = {}
    for it in iterations:
        devs = pooled_mad_all.get(it, [])
        if devs:
            agg_mad_by_it[it] = float(stats.mean(devs))
        devs_std = pooled_std_all.get(it, [])
        if len(devs_std) >= 2:
            agg_std_by_it[it] = float(stats.stdev(devs_std))
        elif devs_std:
            agg_std_by_it[it] = 0.0

    mad_counts_by_it_attr: Dict[int, Dict[str, int]] = {}
    mad_counts_by_it: Dict[int, int] = {}
    mad_stats_by_it_attr: Dict[int, Dict[str, Tuple[float, float, float]]] = {}
    for it, attrs in pooled_mad_devs.items():
        mad_counts_by_it_attr[it] = {a: len(v) for a, v in attrs.items()}
        mad_counts_by_it[it] = sum(mad_counts_by_it_attr[it].values())
        mad_stats_by_it_attr[it] = {}
        for a, devs in attrs.items():
            if devs:
                med = float(stats.median(devs))
                nonzero = sum(1 for d in devs if d != 0)
                pct_nonzero = (nonzero / len(devs)) * 100.0
                mad_stats_by_it_attr[it][a] = (med, pct_nonzero, float(len(devs)))

    return std_by_it_attr, mad_by_it_attr, iterations, agg_mad_by_it, agg_std_by_it, mad_counts_by_it_attr, mad_counts_by_it, mad_stats_by_it_attr


def plot_dispersion_bars(
    mad_by_it_attr: Dict[int, Dict[str, float]],
    std_by_it_attr: Dict[int, Dict[str, float]],
    attributes: List[str],
    iterations: List[int],
    out: Path,
    agg_mad_by_it: Dict[int, float],
    agg_std_by_it: Dict[int, float],
) -> None:
    """Draw two stacked bar charts: top=pooled MAD (0-2), bottom=pooled std (1-10)."""
    try:
        import matplotlib.pyplot as plt  # type: ignore
        import numpy as np  # type: ignore
    except Exception as e:
        raise SystemExit(
            "Missing dependency: matplotlib (and numpy). Install with `pip install matplotlib numpy` or `pip install -r requirements.txt`."
        ) from e

    x = np.arange(len(iterations))
    n_attr = len(attributes)
    width = min(0.16, 0.8 / max(1, n_attr))
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(11, 10), sharex=True)

    # Normalize attribute keys so Expert names map to sheet/data names
    available_keys: set[str] = set()
    for it_map in mad_by_it_attr.values():
        available_keys.update(it_map.keys())
    for it_map in std_by_it_attr.values():
        available_keys.update(it_map.keys())
    norm_map: Dict[str, str] = {_norm_key(k): k for k in available_keys}

    def _lookup_attr_key(attr: str) -> str:
        return norm_map.get(_norm_key(attr), attr)

    # Top: pooled MAD by attribute
    for i, attr in enumerate(attributes):
        key = _lookup_attr_key(attr)
        vals = [mad_by_it_attr.get(it, {}).get(key, 0.0) for it in iterations]
        ax1.bar(x + (i - (n_attr - 1) / 2) * width, vals, width=width, label=attr)
    agg_mad = [float(agg_mad_by_it.get(it, 0.0)) for it in iterations]
    ax1.plot(
        x,
        agg_mad,
        "-o",
        color="tab:cyan",
        linewidth=3.0,
        markersize=8,
        markerfacecolor="tab:cyan",
        markeredgecolor="tab:cyan",
        label="Aggregate across attributes",
    )
    ax1.set_ylabel("Pooled Mean Absolute Deviation (0,1,2)")
    ax1.set_ylim(bottom=0)
    ax1.grid(axis="y", alpha=0.3)
    ax1.text(0.01, 0.02, "(a)", transform=ax1.transAxes, ha="left", va="bottom", fontsize=11)

    # Bottom: pooled std by attribute
    for i, attr in enumerate(attributes):
        key = _lookup_attr_key(attr)
        vals = [std_by_it_attr.get(it, {}).get(key, 0.0) for it in iterations]
        ax2.bar(x + (i - (n_attr - 1) / 2) * width, vals, width=width, label=attr)
    agg_std = [float(agg_std_by_it.get(it, 0.0)) for it in iterations]
    ax2.plot(
        x,
        agg_std,
        "-o",
        color="tab:cyan",
        linewidth=3.0,
        markersize=8,
        markerfacecolor="tab:cyan",
        markeredgecolor="tab:cyan",
        label="Aggregate across attributes",
    )
    ax2.set_ylabel("Pooled Standard Deviation (1-10)")
    ax2.set_ylim(bottom=0)
    ax2.grid(axis="y", alpha=0.3)
    ax2.set_xticks(x)
    ax2.set_xticklabels([str(it) for it in iterations])
    ax2.set_xlabel("Prompt Iteration")
    ax2.legend(ncol=min(3, n_attr), fontsize=10)
    ax2.text(0.01, 0.02, "(b)", transform=ax2.transAxes, ha="left", va="bottom", fontsize=11)

    fig.tight_layout()
    fig.savefig(out, dpi=200)
    plt.close(fig)


def _write_aggregate_lines_xlsx(path: Path, iterations: List[int], agg_mad_by_it: Dict[int, float], agg_std_by_it: Dict[int, float]) -> None:
    try:
        from openpyxl import Workbook  # type: ignore
    except Exception as e:
        raise SystemExit(
            "Missing dependency: openpyxl. Install with `pip install openpyxl` or `pip install -r requirements.txt`."
        ) from e
    wb = Workbook()
    ws = wb.active
    ws.title = "aggregate_lines"
    ws.append(["iteration", "agg_mean_abs_dev_0_2", "agg_pooled_std_1_10"])
    for it in iterations:
        ws.append([it, float(agg_mad_by_it.get(it, 0.0)), float(agg_std_by_it.get(it, 0.0))])
    wb.save(str(path))


def main(argv: Optional[List[str]] = None) -> int:
    import sys
    ap = argparse.ArgumentParser(description="Aggregate dispersion over iterations: pooled std (1-10) and pooled MAD (0-2)")
    ap.add_argument("--glob", default="results_*.xlsx", help="Glob of per-space result workbooks")
    ap.add_argument("--xlsx", default="BD GPT Prompts.xlsx", help="Workbook containing Expert sheet (for attribute order)")
    ap.add_argument("--out", default="dispersion_progression.png", help="Output image path")
    ap.add_argument("--agg-out", default="dispersion_aggregate_lines.xlsx", help="Output xlsx for aggregate lines")
    args = ap.parse_args(argv or sys.argv[1:])

    import glob as _glob
    result_paths = [Path(p) for p in sorted(_glob.glob(args.glob)) if p.lower().endswith(".xlsx")]
    if not result_paths:
        raise SystemExit("No result files matched.")

    std_by_it_attr, mad_by_it_attr, iterations, agg_mad_by_it, agg_std_by_it, mad_counts_by_it_attr, mad_counts_by_it, mad_stats_by_it_attr = compute_dispersion_by_attribute(result_paths)
    attributes = _attributes_from_expert(Path(args.xlsx))
    if not attributes:
        attrs: set[str] = set()
        for p in result_paths:
            wb = _load_wb(p)
            attrs.update(s for s in wb.sheetnames if s.lower() not in {"meta", "prompts"})
        attributes = sorted(attrs)

    # Debug: MAD stats per iteration/attribute (median, % nonzero deviations, count)
    print("MAD pooled deviation stats by iteration:")
    for it in iterations:
        per_attr = mad_stats_by_it_attr.get(it, {})
        parts = ", ".join(
            f"{k}: med={v[0]:.3f}, nonzero={v[1]:.1f}%, n={int(v[2])}"
            for k, v in sorted(per_attr.items())
        )
        print(f"  it{it}: {parts}")

    print("MAD aggregate (median of pooled deviations) by iteration:")
    for it in iterations:
        print(f"  it{it}: {agg_mad_by_it.get(it, 0.0):.4f}")
    print("Std aggregate (pooled deviations) by iteration:")
    for it in iterations:
        print(f"  it{it}: {agg_std_by_it.get(it, 0.0):.4f}")

    plot_dispersion_bars(mad_by_it_attr, std_by_it_attr, attributes, iterations, Path(args.out), agg_mad_by_it, agg_std_by_it)
    _write_aggregate_lines_xlsx(Path(args.agg_out), iterations, agg_mad_by_it, agg_std_by_it)
    print(f"Saved {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
