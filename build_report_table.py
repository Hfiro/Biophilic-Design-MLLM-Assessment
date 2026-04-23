"""
Author: Hoda Alem
Date: November 2025
Context: Part of PhD research for a PhD in Environmental Design and Planning
         at Myers-Lawson School of Construction, Virginia Tech, Blacksburg, VA.
Purpose: Generate a compact report table for the last prompt iteration that
         mirrors the Expert layout: two columns per space (0â€“2 and 1â€“10).
Features:
- Reads order and Expert values from the Expert sheet.
- Pulls median/Q1/Q3 (0â€“2) and mean/std (1â€“10) from combined_summary.xlsx.
- Emits Excel with annotated mismatches (Expert values in parentheses).
- Optional figure (SVG/PNG/PDF) and LaTeX table export for publications.
"""

import argparse
import math
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple


def _norm_key(value: Any) -> str:
    """Normalize labels so matching ignores case and extra whitespace."""
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _match_space_dict(mapping: Dict[str, Dict[str, Any]], space: str) -> Dict[str, Any]:
    if space in mapping:
        return mapping[space]
    norm = _norm_key(space)
    for name, data in mapping.items():
        if _norm_key(name) == norm:
            return data
    return {}


def _match_attr_value(mapping: Dict[str, Any], attribute: str) -> Any:
    if attribute in mapping:
        return mapping[attribute]
    norm = _norm_key(attribute)
    for name, value in mapping.items():
        if _norm_key(name) == norm:
            return value
    return None


def _normalize_tokens(s: str) -> List[str]:
    import re
    s = (s or "").lower()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return [t for t in s.split() if t]


def _load_wb(path: Path):
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:
        raise SystemExit(
            "Missing dependency: openpyxl. Install with `pip install openpyxl` or `pip install -r requirements.txt`."
        ) from e
    return load_workbook(filename=str(path), read_only=True, data_only=True)


def _read_expert_orders_and_values(path: Path, sheet_name: str = "Expert") -> Tuple[List[str], List[str], Dict[str, Dict[str, float]], Dict[str, Dict[str, float]]]:
    """Return (spaces_order, attributes_order, expert02_map, expert10_map)."""
    wb = _load_wb(path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {path}")
    ws = wb[sheet_name]

    header1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    spaces_order: List[str] = []
    start_col = 3
    c = start_col
    while c <= ws.max_column:
        name = str(header1[c - 1] or '').strip()
        if name and (not spaces_order or spaces_order[-1] != name):
            spaces_order.append(name)
        c += 2

    attributes_order: List[str] = []
    r = 3
    while r <= ws.max_row:
        val = ws.cell(row=r, column=2).value
        if val is None:
            r += 1
            continue
        attr = str(val).strip()
        if attr:
            attributes_order.append(attr)
        r += 1

    expert02: Dict[str, Dict[str, float]] = {}
    expert10: Dict[str, Dict[str, float]] = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        attr_raw = row[1] if len(row) > 1 else None
        if not attr_raw:
            continue
        attribute = str(attr_raw).strip()
        c = start_col
        s_idx = 0
        while c <= ws.max_column and s_idx < len(spaces_order):
            space = spaces_order[s_idx]
            v02 = row[c - 1] if c - 1 < len(row) else None
            c10 = c + 1
            v10 = row[c10 - 1] if c10 - 1 < len(row) else None
            def to_num(x):
                try:
                    if x is None or (isinstance(x, float) and math.isnan(x)):
                        return None
                    if isinstance(x, (int, float)):
                        return float(x)
                    s = str(x).strip()
                    return float(s) if s else None
                except Exception:
                    return None
            r02 = to_num(v02)
            r10 = to_num(v10)
            if r02 is not None:
                expert02.setdefault(space, {})[attribute] = r02
            if r10 is not None:
                expert10.setdefault(space, {})[attribute] = r10
            c += 2
            s_idx += 1

    return spaces_order, attributes_order, expert02, expert10


def _resolve_attr_sheet(wb, desired: str) -> Optional[str]:
    if desired in wb.sheetnames:
        return desired
    tokens = set(_normalize_tokens(desired))
    cands = [s for s in wb.sheetnames if s.lower() != 'index']
    best_subset = [s for s in cands if tokens and tokens.issubset(set(_normalize_tokens(s)))]
    if best_subset:
        return sorted(best_subset, key=lambda x: (len(_normalize_tokens(x)), x))[0]
    scored = sorted(((len(tokens & set(_normalize_tokens(s))), s) for s in cands), reverse=True)
    if scored and scored[0][0] > 0:
        return scored[0][1]
    return None


def _headers_from_sheet(ws) -> Dict[str, int]:
    row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers: Dict[str, int] = {}
    for i, v in enumerate(row):
        if isinstance(v, str):
            headers[v.strip()] = i
    return headers


def _find_last_iteration(headers: Dict[str, int]) -> int:
    import re
    iters: List[int] = []
    for h in headers.keys():
        m = re.match(r"it(\d+)_", h)
        if m:
            iters.append(int(m.group(1)))
    if not iters:
        raise ValueError("No iteration columns found (expected headers like it1_*)")
    return max(iters)


def _format_int_or_1dp(x: Optional[float]) -> str:
    if x is None:
        return ""
    if abs(x - round(x)) < 1e-9:
        return str(int(round(x)))
    return f"{x:.1f}"


def _format_two_dec(x: Optional[float]) -> str:
    if x is None:
        return ""
    return f"{float(x):.2f}"


def build_table(summary_xlsx: Path, expert_xlsx: Path, out_xlsx: Path,
                return_structured: bool = False):
    """Create an Excel report table for the last iteration.

    - Reads order and expert values from the Expert sheet.
    - Pulls median/Q1/Q3 (0â€“2) and mean/std (1â€“10) from the consolidated workbook.
    - Annotates mismatches by appending the expert value in parentheses.
    If return_structured is True, also return a dict with body rows for figure/LaTeX export.
    """
    try:
        from openpyxl import Workbook
    except Exception as e:
        raise SystemExit(
            "Missing dependency: openpyxl. Install with `pip install openpyxl` or `pip install -r requirements.txt`."
        ) from e

    spaces, attributes, expert02, expert10 = _read_expert_orders_and_values(expert_xlsx)
    wb_sum = _load_wb(summary_xlsx)

    # Determine last iteration from any attribute sheet
    sample_sheet_name = _resolve_attr_sheet(wb_sum, attributes[0])
    if not sample_sheet_name:
        raise ValueError(f"Attribute sheet for '{attributes[0]}' not found in {summary_xlsx}")
    ws_sample = wb_sum[sample_sheet_name]
    headers = _headers_from_sheet(ws_sample)
    last_it = _find_last_iteration(headers)

    # Precompute column keys for last iteration
    key_median = f"it{last_it}_median_rating_0_2"
    key_q1 = f"it{last_it}_q1_rating_0_2"
    key_q3 = f"it{last_it}_q3_rating_0_2"
    key_mean = f"it{last_it}_mean_score_1_10"
    key_std = f"it{last_it}_std_score_1_10"

    out_wb = Workbook()
    ws_out = out_wb.active
    ws_out.title = f"Iteration {last_it}"

    # Two-row header mimicking Expert layout
    header1: List[Any] = [""]
    header2: List[Any] = ["Attribute"]
    for s in spaces:
        header1.extend([s, s])
        header2.extend(["0-2", "1-10"])
    ws_out.append(header1)
    ws_out.append(header2)

    # Fill rows per attribute
    for attr in attributes:
        sheet_name = _resolve_attr_sheet(wb_sum, attr)
        if not sheet_name:
            continue
        ws_attr = wb_sum[sheet_name]
        hmap = _headers_from_sheet(ws_attr)
        rows = list(ws_attr.iter_rows(min_row=2, values_only=True))
        # Build index of space -> row values (skip expert rows)
        per_space: Dict[str, Dict[str, Any]] = {}
        for r in rows:
            if not r or r[0] is None:
                continue
            name = str(r[0]).strip()
            if name.endswith("(Expert)"):
                continue
            # Map header values
            rec: Dict[str, Any] = {}
            for k, i in hmap.items():
                rec[k] = r[i] if i < len(r) else None
            per_space[name] = rec

        out_row: List[Any] = [attr]
        for s in spaces:
            rec = _match_space_dict(per_space, s)
            med = rec.get(key_median)
            q1 = rec.get(key_q1)
            q3 = rec.get(key_q3)
            mean = rec.get(key_mean)
            std = rec.get(key_std)

            # 0â€“2 string
            med_str = str(int(round(med))) if isinstance(med, (int, float)) and not math.isnan(float(med)) else ""
            q1_str = _format_int_or_1dp(float(q1)) if isinstance(q1, (int, float)) and not math.isnan(float(q1)) else ""
            q3_str = _format_int_or_1dp(float(q3)) if isinstance(q3, (int, float)) and not math.isnan(float(q3)) else ""
            cell_02 = med_str
            if q1_str or q3_str:
                cell_02 = f"{med_str} [{q1_str}, {q3_str}]"
            # Append expert if mismatch
            exp02 = _match_attr_value(_match_space_dict(expert02, s), attr)
            if exp02 is not None and med_str != "":
                try:
                    if float(med) != float(exp02):
                        cell_02 += f" ({int(round(exp02))})"
                except Exception:
                    pass

            # 1â€“10 string
            mean_str = _format_two_dec(float(mean)) if isinstance(mean, (int, float)) and not math.isnan(float(mean)) else ""
            std_str = _format_two_dec(float(std)) if isinstance(std, (int, float)) and not math.isnan(float(std)) else ""
            cell_10 = mean_str
            if mean_str:
                cell_10 = f"{mean_str} Â± {std_str or '0.00'}"
            exp10 = _match_attr_value(_match_space_dict(expert10, s), attr)
            if exp10 is not None and mean_str:
                try:
                    if float(mean) != float(exp10):
                        # Expert displayed with up to 2 decimals, trim trailing zeros
                        e = f"{float(exp10):.2f}".rstrip('0').rstrip('.')
                        cell_10 += f" ({e})"
                except Exception:
                    pass

            out_row.extend([cell_02, cell_10])

        ws_out.append(out_row)

    # Autosize columns
    try:
        for col in ws_out.columns:
            max_len = 0
            col_letter = getattr(col[0], "column_letter", None)
            for cell in col:
                val = cell.value
                if val is None:
                    continue
                max_len = max(max_len, len(str(val)))
            if col_letter:
                ws_out.column_dimensions[col_letter].width = min(max_len + 2, 60)
    except Exception:
        pass

    out_wb.save(out_xlsx)

    if return_structured:
        # Build structured arrays for optional figure export
        col_pairs = [(s, '0-2') for s in spaces] + [(s, '1-10') for s in spaces]
        # But we prefer a flat pair order [ (s, '0-2'), (s, '1-10') ] already encoded in rows
        # Construct cell matrix (no header rows) to be used by figure exporter
        body_rows: List[List[str]] = []
        # Recompute body rows similarly to above to avoid reading from worksheet again
        # Re-open summary workbook structures used above
        body_rows = []
        # Build dictionary: attr -> {space: (cell_02, cell_10)}
        # Reuse previously computed per_space records is non-trivial here; we will re-read minimally
        wb_sum2 = _load_wb(summary_xlsx)
        for attr in attributes:
            sheet_name = _resolve_attr_sheet(wb_sum2, attr)
            if not sheet_name:
                continue
            ws_attr = wb_sum2[sheet_name]
            hmap = _headers_from_sheet(ws_attr)
            rows2 = list(ws_attr.iter_rows(min_row=2, values_only=True))
            per_space2: Dict[str, Dict[str, Any]] = {}
            for r in rows2:
                if not r or r[0] is None:
                    continue
                name = str(r[0]).strip()
                if name.endswith("(Expert)"):
                    continue
                rec: Dict[str, Any] = {k: (r[i] if i < len(r) else None) for k, i in hmap.items()}
                per_space2[name] = rec

            row_cells: List[str] = [attr]
            for s in spaces:
                rec = _match_space_dict(per_space2, s)
                med = rec.get(key_median)
                q1 = rec.get(key_q1)
                q3 = rec.get(key_q3)
                mean = rec.get(key_mean)
                std = rec.get(key_std)

                med_str = str(int(round(med))) if isinstance(med, (int, float)) and not math.isnan(float(med)) else ""
                q1_str = _format_int_or_1dp(float(q1)) if isinstance(q1, (int, float)) and not math.isnan(float(q1)) else ""
                q3_str = _format_int_or_1dp(float(q3)) if isinstance(q3, (int, float)) and not math.isnan(float(q3)) else ""
                cell_02 = med_str
                if q1_str or q3_str:
                    cell_02 = f"{med_str} [{q1_str}, {q3_str}]"
                exp02 = _match_attr_value(_match_space_dict(expert02, s), attr)
                if exp02 is not None and med_str:
                    try:
                        if float(med) != float(exp02):
                            cell_02 += f" ({int(round(exp02))})"
                    except Exception:
                        pass

                mean_str = _format_two_dec(float(mean)) if isinstance(mean, (int, float)) and not math.isnan(float(mean)) else ""
                std_str = _format_two_dec(float(std)) if isinstance(std, (int, float)) and not math.isnan(float(std)) else ""
                cell_10 = mean_str
                if mean_str:
                    cell_10 = f"{mean_str} Â± {std_str or '0.00'}"
                exp10 = _match_attr_value(_match_space_dict(expert10, s), attr)
                if exp10 is not None and mean_str:
                    try:
                        if float(mean) != float(exp10):
                            e = f"{float(exp10):.2f}".rstrip('0').rstrip('.')
                            cell_10 += f" ({e})"
                    except Exception:
                        pass

                row_cells.extend([cell_02, cell_10])
            body_rows.append(row_cells)

        return {
            'iteration': last_it,
            'spaces': spaces,
            'attributes': attributes,
            'body_rows': body_rows,
        }


def _two_lines(s: str, max_first: int = 16) -> str:
    s = s.strip()
    if len(s) <= max_first:
        return s
    parts = s.split()
    if len(parts) < 2:
        return s
    mid = len(parts) // 2
    return " ".join(parts[:mid]) + "\n" + " ".join(parts[mid:])


def export_table_figure(structured: Dict[str, Any], out_path: Path, figsize: Tuple[float, float] = (12.0, 6.5)) -> None:
    """Export a compact figure of the table with grouped headers (space over 0â€“2 and 1â€“10)."""
    try:
        import matplotlib.pyplot as plt  # type: ignore
    except Exception as e:
        raise SystemExit(
            "Missing dependency: matplotlib. Install with `pip install matplotlib` or `pip install -r requirements.txt`."
        ) from e

    spaces: List[str] = structured['spaces']
    body_rows: List[List[str]] = structured['body_rows']
    iteration: int = structured['iteration']

    # Build headers
    col_labels = ["Attribute"] + sum((["0-2", "1-10"] for _ in spaces), [])
    n_cols = len(col_labels)

    fig, ax = plt.subplots(figsize=figsize)
    ax.axis('off')

    # Column widths: allocate 22% to Attribute, rest evenly across data columns
    attr_w = 0.22
    rem = 1.0 - attr_w
    per = rem / (2 * len(spaces) if spaces else 1)
    col_widths = [attr_w] + [per] * (n_cols - 1)

    # Build table
    the_table = ax.table(
        cellText=body_rows,
        colLabels=col_labels,
        cellLoc='center',
        colLoc='center',
        loc='upper center',
        colWidths=col_widths,
    )
    the_table.auto_set_font_size(False)
    the_table.set_fontsize(10)
    the_table.scale(1.0, 1.2)

    # Style header row
    for j in range(n_cols):
        cell = the_table[0, j]
        cell.set_text_props(weight='bold')
        cell.set_linewidth(1.2)

    # Emphasize top and bottom borders
    n_rows = len(body_rows) + 1  # includes header
    for j in range(n_cols):
        the_table[0, j].set_linewidth(1.5)  # top border via header top lines
        the_table[n_rows - 1, j].set_linewidth(1.5)  # bottom line for last row

    # Add space group labels above the two columns for each space (two-line if long)
    # Compute x centers in axes coords from col_widths
    x_positions: List[float] = []
    acc = 0.0
    for w in col_widths:
        x_positions.append(acc + w / 2.0)
        acc += w
    # y position slightly above header
    y = the_table[0, 0].get_window_extent(fig.canvas.get_renderer())
    # Use a fixed relative y offset in axes coordinates
    y_text = 1.02
    for i, s in enumerate(spaces):
        j0 = 1 + 2 * i
        j1 = j0 + 1
        x_center = (x_positions[j0] + x_positions[j1]) / 2.0
        ax.text(x_center, y_text, _two_lines(s), ha='center', va='bottom', transform=ax.transAxes, fontsize=12, fontweight='bold')

    ax.set_title(f"Iteration {iteration}", fontsize=12, pad=16)
    fig.tight_layout()
    fig.savefig(out_path, dpi=300)
    plt.close(fig)


def _latex_escape(s: str) -> str:
    repl = {
        "\\": r"\textbackslash{}",
        "&": r"\&",
        "%": r"\%",
        "$": r"\$",
        "#": r"\#",
        "_": r"\_",
        "{": r"\{",
        "}": r"\}",
        "~": r"\textasciitilde{}",
        "^": r"\textasciicircum{}",
    }
    out = []
    for ch in s:
        out.append(repl.get(ch, ch))
    return "".join(out)


def export_table_latex(structured: Dict[str, Any], out_path: Path, standalone: bool = True) -> None:
    """Export a LaTeX table using booktabs + tabularx.

    If standalone is False, emit only the table snippet (no preamble/document).
    """
    spaces: List[str] = structured['spaces']
    body_rows: List[List[str]] = structured['body_rows']
    iteration: int = structured['iteration']

    # Build LaTeX code with booktabs + tabularx; auto-wrap cells to avoid overflow
    n_pairs = len(spaces)
    # Column spec: l then 2*n_pairs Y columns (centered X columns)
    preamble = r"""
% Auto-generated table
\usepackage{booktabs}
\usepackage{tabularx}
\usepackage{array}
\newcolumntype{Y}{>{\centering\arraybackslash}X}
""".strip()

    header_top = ["", *(f"\\multicolumn{{2}}{{c}}{{{_latex_escape(s)}}}" for s in spaces)]
    header_sub = ["Attribute", *(["0--2", "1--10"] * n_pairs)]

    lines: List[str] = []
    if standalone:
        lines.append(r"\documentclass{article}")
        lines.append(r"\usepackage[margin=1in]{geometry}")
        lines.append(preamble)
        lines.append(r"\begin{document}")
        lines.append(rf"\section*{{Iteration {iteration}}}")
    else:
        lines.append(preamble)

    lines.append(r"\begingroup")
    lines.append(r"\small")
    lines.append(r"\setlength{\tabcolsep}{4pt}")
    colspec = "l " + " ".join(["Y Y"] * n_pairs)
    lines.append(rf"\begin{{tabularx}}{{\textwidth}}{{{colspec}}}")
    lines.append(r"\toprule")
    lines.append(" ".join(header_top) + " \\")
    # cmidrules per pair
    cmr = []
    # compute column index ranges (1-based) for \\cmidrule
    for i in range(n_pairs):
        c0 = 2 * i + 2  # first data col index (Attribute is col 1)
        c1 = c0 + 1
        cmr.append(rf"\cmidrule(lr){{{c0}-{c1}}}")
    lines.append("".join(cmr))
    lines.append(" ".join(header_sub) + " \\")
    lines.append(r"\midrule")

    for row in body_rows:
        # row[0] is attribute; remaining are cells already formatted
        esc = [_latex_escape(str(x)) for x in row]
        # Replace Â± with math \pm
        esc = [e.replace("Â±", r"$\pm$") for e in esc]
        lines.append(" & ".join(esc) + " \\")

    lines.append(r"\bottomrule")
    lines.append(r"\end{tabularx}")
    lines.append(r"\endgroup")
    if standalone:
        lines.append(r"\end{document}")

    out_path.write_text("\n".join(lines), encoding='utf-8')


def main(argv: Optional[List[str]] = None) -> int:
    import sys
    ap = argparse.ArgumentParser(description="Build a report table with 0â€“2 (median [Q1, Q3]) and 1â€“10 (mean Â± std) for last iteration; optional figure export")
    ap.add_argument('--summary', default='combined_summary.xlsx', help='Consolidated workbook path')
    ap.add_argument('--expert', default='BD GPT Prompts.xlsx', help='Workbook containing Expert sheet')
    ap.add_argument('--out', default='final_report_table.xlsx', help='Output Excel workbook')
    ap.add_argument('--fig-out', help='Optional path to export a nicely formatted table figure (SVG/PNG/PDF)')
    ap.add_argument('--latex-out', help='Optional path to export LaTeX table (standalone .tex or snippet)')
    ap.add_argument('--latex-snippet', action='store_true', help='Export LaTeX as a snippet (no preamble/document)')
    ap.add_argument('--fig-width', type=float, default=12.0, help='Figure width in inches')
    ap.add_argument('--fig-height', type=float, default=6.5, help='Figure height in inches')
    args = ap.parse_args(argv or sys.argv[1:])

    structured = build_table(Path(args.summary), Path(args.expert), Path(args.out), return_structured=bool(args.fig_out))
    print(f"Wrote {args.out}")
    if args.fig_out:
        export_table_figure(structured, Path(args.fig_out), figsize=(args.fig_width, args.fig_height))
        print(f"Wrote {args.fig_out}")
    if args.latex_out:
        export_table_latex(structured, Path(args.latex_out), standalone=not args.latex_snippet)
        print(f"Wrote {args.latex_out}")
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
