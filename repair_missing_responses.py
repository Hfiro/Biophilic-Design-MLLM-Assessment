# Author: Hoda Alem
# Date: November 2025
# Context: Part of PhD research for a PhD in Environmental Design and Planning
#          at Myers-Lawson School of Construction, Virginia Tech, Blacksburg, VA.

import argparse
import sys
import time
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple


def _headers(ws) -> Dict[str, int]:
    row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    out: Dict[str, int] = {}
    for i, v in enumerate(row, start=1):
        if isinstance(v, str):
            out[v.strip()] = i
    return out


def _cell_val(row, idx: Dict[str, int], key: str) -> Any:
    col = idx.get(key)
    if not col:
        return None
    return row[col - 1] if col - 1 < len(row) else None


def _is_number(x: Any) -> bool:
    return isinstance(x, (int, float))


def _issue_flags(row, idx: Dict[str, int]) -> Dict[str, bool]:
    j = _cell_val(row, idx, "justification")
    raw = _cell_val(row, idx, "raw")
    model = _cell_val(row, idx, "model")
    r02 = _cell_val(row, idx, "rating_0_2")
    s10 = _cell_val(row, idx, "score_1_10")
    return {
        "justification_error": (isinstance(j, str) and j.startswith("ERROR:")),
        "raw_missing": (not raw or (isinstance(raw, str) and not raw.strip())),
        "rating_score_missing": ((r02 is None or r02 == "") and (s10 is None or s10 == "")),
        "model_missing": (not model or (isinstance(model, str) and not str(model).strip())),
    }


def _needs_repair(row, idx: Dict[str, int], errors_only: bool = False) -> Tuple[bool, str]:
    flags = _issue_flags(row, idx)
    reasons: List[str] = []
    if errors_only:
        if flags["justification_error"]:
            return True, "justification ERROR"
        return False, ""
    # broader criteria
    if flags["justification_error"]:
        reasons.append("justification ERROR")
    if flags["raw_missing"]:
        reasons.append("raw missing")
    if flags["rating_score_missing"]:
        reasons.append("rating/score missing")
    if flags["model_missing"]:
        reasons.append("model missing")
    return (len(reasons) > 0), ", ".join(reasons)


def _load_prompts_map(xlsx_path: Path) -> Dict[Tuple[str, int], str]:
    from biophilic_eval import _read_attribute_prompts_from_sheet
    prompts = _read_attribute_prompts_from_sheet(xlsx_path)
    out: Dict[Tuple[str, int], str] = {}
    for rec in prompts:
        try:
            attr = str(rec.get("attribute", "")).strip()
            it = int(rec.get("iteration")) if rec.get("iteration") is not None else None
            prompt = str(rec.get("prompt", "")).strip()
        except Exception:
            continue
        if attr and it and prompt:
            out[(attr.lower(), it)] = prompt
    return out


def _collect_images_for_space(space: str) -> List[Path]:
    from pathlib import Path as _Path
    from biophilic_eval import _collect_five_jpgs
    p = _Path(space)
    if not p.is_absolute():
        p = _Path.cwd() / p
    return _collect_five_jpgs(p)


def _encode_images(paths: List[Path]) -> List[str]:
    from biophilic_eval import _encode_image_as_data_url
    return [_encode_image_as_data_url(str(p)) for p in paths]


def _call_model(image_data_urls: List[str], prompt_text: str, model: str, detail: str, timeout: float, allow_fallback: bool) -> Tuple[str, str]:
    from biophilic_eval import analyze_color
    return analyze_color(
        image_data_urls=image_data_urls,
        prompt_text=prompt_text,
        model=model,
        detail=detail,
        allow_fallback=allow_fallback,
        timeout=timeout,
    )


def _parse_response(text: str) -> Tuple[Optional[int], Optional[float], str]:
    from biophilic_eval import _extract_rating_and_score, _first_sentences
    rating, score = _extract_rating_and_score(text)
    just = _first_sentences(text, max_sentences=3)
    return rating, score, just


def _load_env():
    try:
        from biophilic_eval import _load_env_from_dotenv  # type: ignore
        _load_env_from_dotenv()
    except Exception:
        # Best-effort; fall through if helper is unavailable
        pass


def repair_file(
    path: Path,
    prompts_map: Dict[Tuple[str, int], str],
    model: str,
    detail: str,
    timeout: float,
    allow_fallback: bool,
    dry_run: bool,
    copy_mode: bool,
    limit: Optional[int] = None,
    spaces_filter: Optional[List[str]] = None,
    errors_only: bool = False,
) -> Tuple[int, int]:
    from openpyxl import load_workbook

    target_path = path
    if copy_mode:
        target_path = path.with_name(f"{path.stem}_repaired{path.suffix}")

    wb = load_workbook(filename=str(path), read_only=False, data_only=True)
    total_issues = 0
    fixed = 0

    for sheet in wb.sheetnames:
        if sheet in {"meta", "prompts"}:
            continue
        ws = wb[sheet]
        idx = _headers(ws)
        header = list(idx.keys())
        if not header:
            continue
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row is None:
                continue
            try:
                space = _cell_val(row, idx, "space") or None
                if isinstance(space, str):
                    space = space.strip()
                attr = _cell_val(row, idx, "attribute") or sheet
                if isinstance(attr, str):
                    attr = attr.strip()
                it_raw = _cell_val(row, idx, "iteration")
                it = int(it_raw) if it_raw is not None else None
                run_val = _cell_val(row, idx, "run")
            except Exception:
                continue

            # Only consider true per-run rows; skip summary/blank rows
            if not _is_number(run_val):
                continue

            if spaces_filter and space and str(space) not in spaces_filter:
                continue

            is_issue, reason = _needs_repair(row, idx, errors_only=errors_only)
            if not is_issue:
                continue

            total_issues += 1
            if limit is not None and fixed >= limit:
                continue

            key = (str(attr).lower(), int(it) if it is not None else None)
            prompt_text = None
            if key[1] is not None:
                prompt_text = prompts_map.get((key[0], key[1]))
            if not prompt_text:
                # No prompt available; skip safely but log in dry-run mode already; in repair mode, also log
                if not dry_run:
                    print(f"{path.name} | sheet={sheet} | space={space} | attribute={attr} | iteration={it} | run={run_val} | SKIP: prompt not found")
                continue

            # Collect images
            try:
                img_paths = _collect_images_for_space(str(space))
                data_urls = _encode_images(img_paths)
            except Exception as e:
                # Can't find images; skip
                if not dry_run:
                    print(f"{path.name} | sheet={sheet} | space={space} | attribute={attr} | iteration={it} | run={run_val} | SKIP: images missing ({e})")
                continue

            if dry_run:
                run_val_out = _cell_val(row, idx, "run")
                task_val = _cell_val(row, idx, "task")
                print(f"{path.name} | sheet={sheet} | space={space} | attribute={attr} | iteration={it} | run={run_val_out} | task={task_val} | reason={reason}")
                continue

            # Call model with retries
            attempts = 0
            max_attempts = 3
            backoff = [2, 5, 10]
            got_text: Optional[str] = None
            used_model = model
            last_err: Optional[Exception] = None
            while attempts < max_attempts:
                try:
                    txt, used_model = _call_model(data_urls, prompt_text, model, detail, timeout, allow_fallback)
                    got_text = txt
                    break
                except Exception as e:  # noqa: BLE001
                    last_err = e
                    time.sleep(backoff[min(attempts, len(backoff) - 1)])
                    attempts += 1
            if got_text is None:
                # Leave row unchanged on repeated failure
                print(f"{path.name} | sheet={sheet} | space={space} | attribute={attr} | iteration={it} | run={run_val} | FAIL: {last_err}")
                continue

            rating, score, just = _parse_response(got_text)

            # Write back into row cells
            def _set(name: str, value: Any):
                col = idx.get(name)
                if col:
                    ws.cell(row=row_idx, column=col, value=value)

            _set("model", used_model)
            _set("rating_0_2", rating if rating is not None else "")
            _set("score_1_10", score if score is not None else "")
            _set("justification", just)
            _set("raw", got_text)

            fixed += 1

    if not dry_run and fixed > 0:
        wb.save(str(target_path))
    wb.close()
    return total_issues, fixed


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Repair missing/errored responses in results_*.xlsx by re-calling the API for those rows only.")
    parser.add_argument("--glob", default="results_*.xlsx", help="Glob for per-space result workbooks")
    parser.add_argument("--xlsx", default="BD GPT Prompts.xlsx", help="Workbook containing Prompts sheet")
    parser.add_argument("--model", default="gpt-4o", help="OpenAI model")
    parser.add_argument("--detail", default="high", choices=["low", "high"], help="Image detail level")
    parser.add_argument("--timeout", type=float, default=90.0, help="Per-request timeout in seconds")
    parser.add_argument("--no-fallback", action="store_true", help="Disable fallback to gpt-4o-mini on quota errors")
    parser.add_argument("--dry-run", action="store_true", help="List issues but do not repair")
    parser.add_argument("--copy", action="store_true", help="Write changes to <file>_repaired.xlsx instead of modifying in place")
    parser.add_argument("--limit", type=int, help="Maximum number of rows to repair across all files")
    parser.add_argument("--spaces", nargs="*", help="Optional list of spaces to restrict repairs")
    parser.add_argument("--errors-only", action="store_true", help="Only flag rows with explicit connection/API errors (justification starts with 'ERROR:')")
    args = parser.parse_args(argv or sys.argv[1:])

    # Load .env if present
    _load_env()

    # Build prompts lookup
    prompts_map = _load_prompts_map(Path(args.xlsx))
    if not prompts_map:
        print(f"No prompts found in {args.xlsx}", file=sys.stderr)

    import glob as _glob
    files = [Path(p) for p in sorted(_glob.glob(args.glob)) if p.lower().endswith('.xlsx')]
    if not files:
        print("No result files matched.")
        return 0

    total_issues = 0
    total_fixed = 0
    remaining = args.limit if args.limit is not None else None

    for p in files:
        lim = remaining if remaining is not None else None
        issues, fixed = repair_file(
            p,
            prompts_map,
            model=args.model,
            detail=args.detail,
            timeout=float(args.timeout),
            allow_fallback=not args.no_fallback,
            dry_run=bool(args.dry_run),
            copy_mode=bool(args.copy),
            limit=lim,
            spaces_filter=args.spaces,
            errors_only=bool(args.errors_only),
        )
        total_issues += issues
        total_fixed += fixed
        if remaining is not None:
            remaining = max(0, remaining - fixed)
            if remaining == 0:
                break
        print(f"{p.name}: issues={issues}, fixed={fixed}")

    print(f"Total issues found: {total_issues}")
    if args.dry_run:
        print("No changes written (dry run)")
    else:
        print(f"Total rows repaired: {total_fixed}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
