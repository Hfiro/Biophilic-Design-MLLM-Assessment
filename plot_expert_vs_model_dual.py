"""
Author: Hoda Alem
Date: November 2025
Context: Part of PhD research for a PhD in Environmental Design and Planning
         at Myers-Lawson School of Construction, Virginia Tech, Blacksburg, VA.

Dual scatter: Expert vs Model for both scales (0â€“2 and 1â€“10).

Top panel: 0â€“2 rating (x=Expert, y=Model), with normal jitter.
Bottom panel: 1â€“10 score (x=Expert, y=Model), with normal jitter.

Color can encode iteration or attribute (or space). Identity line y=x is drawn on both.
"""

import argparse
import math
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple


def _normalize_tokens(s: str) -> List[str]:
    import re
    s = (s or "").lower()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return [t for t in s.split() if t]


def _norm_key(s: str) -> str:
    return " ".join(_normalize_tokens(s or ""))


def _load_wb(path: Path):
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:
        raise SystemExit(
            "Missing dependency: openpyxl. Install with `pip install openpyxl`."
        ) from e
    return load_workbook(filename=str(path), read_only=True, data_only=True)


def _read_expert_truths(path: Path, sheet_name: str = "Expert") -> Dict[str, Dict[str, Dict[str, float]]]:
    """Return mapping: space -> attribute -> {'0-2': v, '1-10': v}."""
    wb = _load_wb(path)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"Sheet '{sheet_name}' not found in {path}")
    ws = wb[sheet_name]

    header1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    start_col = 3

    space_cols: List[Tuple[str, int]] = []
    c = start_col
    while c <= ws.max_column:
        space = str(header1[c - 1] or '').strip()
        if space:
            space_cols.append((space, c))
        c += 2

    out: Dict[str, Dict[str, Dict[str, float]]] = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        attr_raw = row[1] if len(row) > 1 else None
        if not attr_raw:
            continue
        attribute = str(attr_raw).strip()
        for space, c02 in space_cols:
            v02 = row[c02 - 1] if c02 - 1 < len(row) else None
            c10 = c02 + 1
            v10 = row[c10 - 1] if c10 - 1 < len(row) else None
            try:
                val02 = float(v02) if isinstance(v02, (int, float)) else float(str(v02).strip()) if v02 else None
            except Exception:
                val02 = None
            try:
                val10 = float(v10) if isinstance(v10, (int, float)) else float(str(v10).strip()) if v10 else None
            except Exception:
                val10 = None
            if val02 is None and val10 is None:
                continue
            out.setdefault(space, {}).setdefault(attribute, {})
            if val02 is not None:
                out[space][attribute]['0-2'] = val02
            if val10 is not None:
                out[space][attribute]['1-10'] = val10
    return out


def _headers(ws) -> Dict[str, int]:
    row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers: Dict[str, int] = {}
    for i, v in enumerate(row):
        if isinstance(v, str):
            headers[v.strip()] = i
    return headers


def _is_number(x: Any) -> bool:
    return isinstance(x, (int, float)) and not (isinstance(x, float) and math.isnan(x))


def _collect_points(paths: List[Path]) -> List[Dict[str, Any]]:
    """Collect per-run rows across all result files."""
    pts: List[Dict[str, Any]] = []
    for p in paths:
        wb = _load_wb(p)
        # derive space name
        space_name: Optional[str] = None
        if 'meta' in wb.sheetnames:
            try:
                ws_meta = wb['meta']
                for r in ws_meta.iter_rows(min_row=1, values_only=True):
                    if r and str(r[0]).strip().lower() == 'space':
                        space_name = str(r[1]).strip() if len(r) > 1 and r[1] is not None else None
                        break
            except Exception:
                pass
        for sheet in wb.sheetnames:
            if sheet in {"meta", "prompts"}:
                continue
            ws = wb[sheet]
            idx = _headers(ws)
            if not idx:
                continue
            def val(row, key: str):
                i = idx.get(key)
                return row[i] if i is not None and i < len(row) else None
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row is None:
                    continue
                rv = val(row, 'run')
                if not _is_number(rv):
                    continue
                r02 = val(row, 'rating_0_2')
                s10 = val(row, 'score_1_10')
                it = val(row, 'iteration')
                if not _is_number(r02) or not _is_number(s10) or not _is_number(it):
                    continue
                attr = val(row, 'attribute') or sheet
                sp = val(row, 'space') or space_name or p.stem
                pts.append({
                    'file': p.name,
                    'sheet': sheet,
                    'space': str(sp),
                    'attribute': str(attr),
                    'iteration': int(it),
                    'run': int(rv),
                    'rating_0_2': float(r02),
                    'score_1_10': float(s10),
                })
    return pts


def _match_expert(expert: Dict[str, Dict[str, Dict[str, float]]], space: str, attribute: str) -> Optional[Dict[str, float]]:
    # exact
    if space in expert and attribute in expert[space]:
        return expert[space][attribute]
    # normalized
    s_norm = _norm_key(space)
    a_norm = _norm_key(attribute)
    for s, amap in expert.items():
        if _norm_key(s) != s_norm:
            continue
        for a, vals in amap.items():
            if _norm_key(a) == a_norm:
                return vals
    return None


def _build_pairs(points: List[Dict[str, Any]], expert: Dict[str, Dict[str, Dict[str, float]]]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """Return (pairs02, pairs10) with expert/model pairs and metadata.
    pairs entries: {x, y, space, attribute, iteration}
    """
    out02: List[Dict[str, Any]] = []
    out10: List[Dict[str, Any]] = []
    for pt in points:
        vals = _match_expert(expert, pt['space'], pt['attribute'])
        if not vals:
            continue
        if '0-2' in vals:
            out02.append({
                'x': float(vals['0-2']),
                'y': float(pt['rating_0_2']),
                'space': pt['space'],
                'attribute': pt['attribute'],
                'iteration': pt['iteration'],
            })
        if '1-10' in vals:
            out10.append({
                'x': float(vals['1-10']),
                'y': float(pt['score_1_10']),
                'space': pt['space'],
                'attribute': pt['attribute'],
                'iteration': pt['iteration'],
            })
    return out02, out10


def _group_points(points: List[Dict[str, Any]]) -> Dict[Tuple[str, str, int], Dict[str, List[float]]]:
    groups: Dict[Tuple[str, str, int], Dict[str, List[float]]] = {}
    for pt in points:
        key = (pt['space'], pt['attribute'], int(pt['iteration']))
        g = groups.setdefault(key, {'r02': [], 's10': []})
        g['r02'].append(float(pt['rating_0_2']))
        g['s10'].append(float(pt['score_1_10']))
    return groups


def _quantiles(values: List[float]) -> Tuple[Optional[float], Optional[float]]:
    try:
        import statistics as stats
        if len(values) < 2:
            return None, None
        qs = stats.quantiles(values, n=4, method="inclusive")
        return float(qs[0]), float(qs[2])
    except Exception:
        return None, None


def _aggregate_pairs(
    points: List[Dict[str, Any]],
    expert: Dict[str, Dict[str, Dict[str, float]]],
    iter_filter: str,
    iter_list: Optional[List[int]],
    rating_err_mode: str,  # 'q1q3' or 'mad'
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    import statistics as stats
    groups = _group_points(points)
    all_iters = sorted({it for (_, _, it) in groups.keys()})
    if iter_list:
        include_iters = set(int(i) for i in iter_list)
    elif iter_filter == 'last' and all_iters:
        include_iters = {max(all_iters)}
    else:
        include_iters = set(all_iters)

    out02: List[Dict[str, Any]] = []
    out10: List[Dict[str, Any]] = []
    for (space, attribute, it), vals in groups.items():
        if it not in include_iters:
            continue
        ex = _match_expert(expert, space, attribute)
        if not ex:
            continue
        r_list = vals['r02']
        s_list = vals['s10']
        # 0-2
        if '0-2' in ex and r_list:
            try:
                med = stats.median(r_list)
            except Exception:
                med = None
            low_err = high_err = None
            if rating_err_mode == 'q1q3':
                q1, q3 = _quantiles(r_list)
                if q1 is not None and med is not None:
                    low_err = max(0.0, med - q1)
                if q3 is not None and med is not None:
                    high_err = max(0.0, q3 - med)
            else:
                try:
                    if med is not None:
                        mad = stats.median([abs(x - med) for x in r_list])
                        low_err = mad
                        high_err = mad
                except Exception:
                    pass
            if med is not None:
                out02.append({
                    'x': float(ex['0-2']),
                    'y': float(med),
                    'yerr_low': float(low_err) if low_err is not None else 0.0,
                    'yerr_high': float(high_err) if high_err is not None else 0.0,
                    'space': space,
                    'attribute': attribute,
                    'iteration': it,
                })
        # 1-10
        if '1-10' in ex and s_list:
            try:
                mean = float(stats.fmean(s_list))
            except Exception:
                mean = None
            std = None
            if len(s_list) >= 2:
                try:
                    std = float(stats.stdev(s_list))
                except Exception:
                    std = None
            if mean is not None:
                out10.append({
                    'x': float(ex['1-10']),
                    'y': float(mean),
                    'yerr': float(std) if std is not None else 0.0,
                    'space': space,
                    'attribute': attribute,
                    'iteration': it,
                })
    return out02, out10


def _colors_for(keys: List[str]):
    import matplotlib.pyplot as plt
    cmap = plt.get_cmap('tab20')
    return {k: cmap((i % 20) / 20.0) for i, k in enumerate(sorted(set(keys)))}


def _pearson_r(xs: List[float], ys: List[float]) -> Optional[float]:
    if len(xs) < 2:
        return None
    mean_x = sum(xs) / len(xs)
    mean_y = sum(ys) / len(ys)
    num = sum((x - mean_x) * (y - mean_y) for x, y in zip(xs, ys))
    den_x = sum((x - mean_x) ** 2 for x in xs)
    den_y = sum((y - mean_y) ** 2 for y in ys)
    den = math.sqrt(den_x * den_y)
    if den == 0:
        return None
    return num / den


def plot_dual(
    pairs02: List[Dict[str, Any]],
    pairs10: List[Dict[str, Any]],
    out: Path,
    color_by: str,
    jitter02: float,
    yjitter02: float,
    jitter10: float,
    yjitter10: float,
    alpha: float,
    size: float,
    seed: Optional[int],
    dpi: int,
    aggregate: bool = False,
) -> None:
    if not pairs02 and not pairs10:
        raise SystemExit('No data to plot.')
    try:
        import matplotlib.pyplot as plt  # type: ignore
        import numpy as np  # type: ignore
    except Exception as e:
        raise SystemExit("Missing dependency: matplotlib (and numpy). Install with `pip install matplotlib numpy`.") from e

    if seed is not None:
        np.random.seed(int(seed))

    # Prepare colors
    if color_by == 'attribute':
        keys02 = [p['attribute'] for p in pairs02]
        keys10 = [p['attribute'] for p in pairs10]
    elif color_by == 'space':
        keys02 = [p['space'] for p in pairs02]
        keys10 = [p['space'] for p in pairs10]
    else:  # iteration
        keys02 = [str(p['iteration']) for p in pairs02]
        keys10 = [str(p['iteration']) for p in pairs10]
    all_keys = sorted(set(keys02 + keys10))
    color_map = _colors_for(all_keys)

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 6), sharex=False, sharey=False)

    def _plot(ax, pairs, jitter, yjitter, xlim, ylim, xticks, yticks, xlab, ylab, allow_y_jitter: bool, panel_label: str):
        if not pairs:
            ax.set_visible(False)
            return
        if not aggregate:
            xs = []
            ys = []
            cols = []
            for p in pairs:
                xs.append(p['x'])
                ys.append(p['y'])
                k = (
                    p['attribute'] if color_by == 'attribute' else p['space'] if color_by == 'space' else str(p['iteration'])
                )
                cols.append(color_map.get(k))
            raw_xs = [float(x) for x in xs]
            raw_ys = [float(y) for y in ys]
            xs = np.array(raw_xs, dtype=float)
            ys = np.array(raw_ys, dtype=float)
            if jitter > 0:
                xs = xs + np.random.normal(0.0, jitter, size=xs.shape)
            if yjitter > 0:
                ys = ys + np.random.normal(0.0, yjitter, size=ys.shape)
            ax.scatter(xs, ys, s=size, alpha=alpha, c=cols, edgecolors='none')
        else:
            xs = []
            ys = []
            cols = []
            yerrs: List[Tuple[float, float]] = []  # (low, high) per point
            for p in pairs:
                xs.append(p['x'])
                ys.append(p['y'])
                if 'yerr_low' in p and 'yerr_high' in p:
                    low = float(p.get('yerr_low', 0.0))
                    high = float(p.get('yerr_high', 0.0))
                else:
                    sym = float(p.get('yerr', 0.0))
                    low = sym
                    high = sym
                yerrs.append((low, high))
                k = (
                    p['attribute'] if color_by == 'attribute' else p['space'] if color_by == 'space' else str(p['iteration'])
                )
                cols.append(color_map.get(k))
            raw_xs = [float(x) for x in xs]
            raw_ys = [float(y) for y in ys]
            xs = np.array(raw_xs, dtype=float)
            ys = np.array(raw_ys, dtype=float)
            # In aggregated mode, keep expert jitter if desired but do NOT jitter model central values
            if jitter > 0:
                xs = xs + np.random.normal(0.0, jitter, size=xs.shape)
            if allow_y_jitter and yjitter > 0:
                ys = ys + np.random.normal(0.0, yjitter, size=ys.shape)
            # Draw points
            alpha_pts = max(alpha, 0.8)
            ax.scatter(xs, ys, s=max(12.0, size), alpha=alpha_pts, c=cols, edgecolors='none')
            # Draw per-point error bars with matching colors
            for i in range(len(xs)):
                low, high = yerrs[i]
                # Build asym yerr for a single point
                yerr = [[low], [high]]
                ax.errorbar([xs[i]], [ys[i]], yerr=yerr, fmt='none', ecolor=cols[i], elinewidth=1.0, alpha=min(1.0, alpha + 0.2))
        # identity line within overlapping domain
        lo = max(xlim[0], ylim[0])
        hi = min(xlim[1], ylim[1])
        ax.plot([lo, hi], [lo, hi], color='black', lw=1.5, alpha=0.8)
        ax.set_xlim(*xlim)
        ax.set_ylim(*ylim)
        if xticks is not None:
            ax.set_xticks(xticks)
        if yticks is not None:
            ax.set_yticks(yticks)
        ax.set_xlabel(xlab)
        ax.set_ylabel(ylab)
        ax.grid(True, alpha=0.25)
        r = _pearson_r(raw_xs, raw_ys)
        r_txt = "Correlation: NA" if r is None else f"Correlation: {r:.2f}"
        ax.text(0.02, 0.96, f"{panel_label} {r_txt}", transform=ax.transAxes, ha='left', va='top', fontsize=10)

    # Panel configs
    xlim02 = (-0.5, 2.5)
    ylim02 = (-0.5, 2.5)
    xticks02 = [0, 1, 2]
    yticks02 = [0, 1, 2]
    xlab02 = 'Expert Ratings (0,1,2)'
    ylab02 = 'ChatGPT Ratings (0,1,2)'

    xlim10 = (0.5, 10.5)
    ylim10 = (0.5, 10.5)
    xticks10 = list(range(1, 11))
    yticks10 = list(range(1, 11))
    xlab10 = 'Expert Ratings (1â€“10)'
    ylab10 = 'ChatGPT Ratings (1â€“10)'

    # For aggregated mode: allow y-jitter only on the 0â€“2 panel
    _plot(ax1, pairs02, jitter02, yjitter02, xlim02, ylim02, xticks02, yticks02, xlab02, ylab02, allow_y_jitter=True, panel_label='(a)')
    _plot(ax2, pairs10, jitter10, yjitter10, xlim10, ylim10, xticks10, yticks10, xlab10, ylab10, allow_y_jitter=False, panel_label='(b)')

    # Build legend using color_map keys, place inside the right panel (bottom-right)
    import matplotlib.pyplot as _plt
    handles = [
        _plt.Line2D([0], [0], marker='o', color='none', markerfacecolor=color_map[k], markersize=max(3, math.sqrt(size)), label=str(k))
        for k in all_keys
    ]
    ax2.legend(handles=handles, loc='lower right', fontsize=8, title=color_by.capitalize(), frameon=True)

    fig.tight_layout(rect=(0.03, 0.03, 0.95, 0.97))
    fig.savefig(out, dpi=dpi)
    _plt.close(fig)


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description='Dual Expert vs Model scatter for 0â€“2 and 1â€“10 scales')
    parser.add_argument('--glob', default='results_*.xlsx', help='Glob for per-space results files')
    parser.add_argument('--xlsx', default='BD GPT Prompts.xlsx', help='Workbook containing Expert sheet')
    parser.add_argument('--out', default='expert_vs_model.png', help='Output image path')
    parser.add_argument('--color-by', choices=['iteration', 'attribute', 'space'], default='iteration', help='Color encoding')
    parser.add_argument('--jitter02', type=float, default=0.08, help='Std dev of normal jitter for 0â€“2 x-axis (Expert)')
    parser.add_argument('--y-jitter02', type=float, default=0.08, help='Std dev of normal jitter for 0â€“2 y-axis (Model)')
    parser.add_argument('--jitter10', type=float, default=0.12, help='Std dev of normal jitter for 1â€“10 x-axis (Expert)')
    parser.add_argument('--y-jitter10', type=float, default=0.12, help='Std dev of normal jitter for 1â€“10 y-axis (Model)')
    parser.add_argument('--alpha', type=float, default=0.5, help='Marker alpha')
    parser.add_argument('--size', type=float, default=18.0, help='Marker size (points^2)')
    parser.add_argument('--seed', type=int, help='Random seed for jitter')
    parser.add_argument('--dpi', type=int, default=180, help='Figure DPI')
    parser.add_argument('--aggregate', action='store_true', help='Plot aggregated central tendency with error bars instead of per-run scatter')
    parser.add_argument('--rating-error', choices=['q1q3', 'mad'], default='q1q3', help='Error bar mode for 0â€“2 when aggregating')
    parser.add_argument('--iter-filter', choices=['all', 'last'], default='all', help='Which iterations to include when aggregating')
    parser.add_argument('--iterations', nargs='*', type=int, help='Explicit list of iterations to include when aggregating')
    args = parser.parse_args()

    import glob as _glob
    paths = [Path(p) for p in sorted(_glob.glob(args.glob)) if p.lower().endswith('.xlsx')]
    if not paths:
        raise SystemExit('No result files matched.')

    expert = _read_expert_truths(Path(args.xlsx))
    points = _collect_points(paths)
    if args.aggregate:
        pairs02, pairs10 = _aggregate_pairs(points, expert, iter_filter=str(args.iter_filter), iter_list=args.iterations, rating_err_mode=str(args.rating_error))
    else:
        pairs02, pairs10 = _build_pairs(points, expert)
    plot_dual(
        pairs02,
        pairs10,
        Path(args.out),
        color_by=str(args.color_by),
        jitter02=float(args.jitter02),
        yjitter02=float(args.y_jitter02),
        jitter10=float(args.jitter10),
        yjitter10=float(args.y_jitter10),
        alpha=float(args.alpha),
        size=float(args.size),
        seed=args.seed,
        dpi=int(args.dpi),
        aggregate=bool(args.aggregate),
    )
    print(f'Saved {args.out}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())


def _pearson_r(xs: List[float], ys: List[float]) -> Optional[float]:
    if len(xs) < 2:
        return None
    mean_x = sum(xs) / len(xs)
    mean_y = sum(ys) / len(ys)
    num = sum((x - mean_x) * (y - mean_y) for x, y in zip(xs, ys))
    den_x = sum((x - mean_x) ** 2 for x in xs)
    den_y = sum((y - mean_y) ** 2 for y in ys)
    den = math.sqrt(den_x * den_y)
    if den == 0:
        return None
    return num / den
