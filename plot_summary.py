"""
Author: Hoda Alem
Date: November 2025
Context: Part of PhD research for a PhD in Environmental Design and Planning
         at Myers-Lawson School of Construction, Virginia Tech, Blacksburg, VA.
Purpose: Visualize per-space, per-attribute progression across prompt iterations
         using consolidated summaries (combined_summary.xlsx)).
Features:
- Resolves attribute sheets robustly and loads model/expert rows.
- Plots 1â€“10 mean Â± standard deviation and 0â€“1â€“2 median with asymmetric
  Q1â€“Q3 whiskers (fallback to Â±IQR/2 when quartiles are unavailable)).
- Supports single-plot export and multi-panel grid (attributes Ã— spaces).
- Minimal, publication-friendly axes, legends, and layout.
"""

import argparse
from pathlib import Path
from typing import Dict, List, Any, Optional


def _normalize_name(s: str) -> List[str]:
    import re
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return [tok for tok in s.split() if tok]


def _labels_match(a: str, b: str) -> bool:
    if a == b:
        return True
    toks_a = _normalize_name(a)
    toks_b = _normalize_name(b)
    return bool(toks_a and toks_b) and set(toks_a) == set(toks_b)


def _strip_expert_suffix(label: str) -> tuple[str, bool]:
    clean = label.strip()
    lower = clean.lower()
    if lower.endswith("(expert)"):
        return clean[: lower.rfind("(expert)")].strip(), True
    return clean, False


def _two_lines(text: str) -> str:
    # Split into roughly two balanced lines by words
    words = text.split()
    if len(words) <= 2:
        return text
    mid = len(words) // 2
    # Ensure at least one word per line
    line1 = " ".join(words[:mid])
    line2 = " ".join(words[mid:])
    return f"{line1}\n{line2}"


def load_attribute_sheet(xlsx: Path, attribute: str) -> Dict[str, Any]:
    """Load one attribute sheet from the consolidated workbook.

    Resolves the sheet name flexibly (token matching) and returns a dict with
    keys: headers (list[str]), idx (header->index), rows (data rows), title (sheet title).
    """
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:
        raise SystemExit(
            "Missing dependency: openpyxl. Install with `pip install openpyxl` or `pip install -r requirements.txt`."
        ) from e

    if not xlsx.exists():
        raise FileNotFoundError(f"Workbook not found: {xlsx}")

    wb = load_workbook(filename=str(xlsx), read_only=True, data_only=True)
    # Prefer exact match
    if attribute in wb.sheetnames:
        sheet_name = attribute
    else:
        # Flexible token-based resolution (ignoring punctuation/spacing)
        attr_tokens = set(_normalize_name(attribute))
        candidates = [s for s in wb.sheetnames if s.lower() != 'index']
        scored: List[tuple[int, str]] = []
        best_subset: List[str] = []
        for s in candidates:
            toks = set(_normalize_name(s))
            if attr_tokens and attr_tokens.issubset(toks):
                best_subset.append(s)
            overlap = len(attr_tokens & toks)
            scored.append((overlap, s))
        if best_subset:
            # If multiple, choose the one with smallest token count (most specific)
            sheet_name = sorted(best_subset, key=lambda x: (len(_normalize_name(x)), x))[0]
        else:
            # Fallback to highest overlap
            scored.sort(reverse=True)
            if scored and scored[0][0] > 0:
                sheet_name = scored[0][1]
            else:
                raise ValueError(f"Sheet '{attribute}' not found. Available: {wb.sheetnames}")

    ws = wb[sheet_name]
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h) if h is not None else "" for h in header]
    idx = {h: i for i, h in enumerate(headers)}

    rows: List[List[Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        rows.append(list(row))
    return {"headers": headers, "idx": idx, "rows": rows, "title": sheet_name}


def _read_space_order_and_map(names_xlsx: Optional[Path]) -> tuple[list[str], Dict[str, str]]:
    order: list[str] = []
    mapping: Dict[str, str] = {}
    if not names_xlsx:
        return order, mapping
    try:
        from openpyxl import load_workbook  # type: ignore
        if not names_xlsx.exists():
            return order, mapping
        wb = load_workbook(filename=str(names_xlsx), read_only=True, data_only=True)
        if 'Spaces' not in wb.sheetnames:
            return order, mapping
        ws = wb['Spaces']
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r:
                continue
            original = str(r[0]).strip() if r[0] is not None else ''
            if not original:
                continue
            order.append(original)
            if len(r) >= 2 and r[1] is not None:
                display = str(r[1]).strip()
                if display:
                    mapping[original] = display
    except Exception:
        return order, mapping
    return order, mapping


def _clean_label(s: str) -> str:
    return s.replace('–', '-').replace('—', '-')


def plot_space_attribute(xlsx: Path, attribute: str, space: str, out: Path, display_map: Optional[Dict[str, str]] = None) -> None:
    """Render a single plot for one attribute and one space.

    Plots 1â€“10 meanÂ±std and 0â€“2 median with Q1â€“Q3 whiskers (fallback: Â±IQR/2).
    """
    data = load_attribute_sheet(xlsx, attribute)
    headers = data["headers"]
    idx = data["idx"]
    rows = data["rows"]

    # Locate the two rows: model row = space; expert row = f"{space} (Expert)"
    model_row: Optional[List[Any]] = None
    expert_row: Optional[List[Any]] = None
    for r in rows:
        raw = str(r[0]) if len(r) > 0 and r[0] is not None else ""
        label, is_expert = _strip_expert_suffix(raw)
        if not raw:
            continue
        if not is_expert and model_row is None and _labels_match(label, space):
            model_row = r
        elif is_expert and expert_row is None and _labels_match(label, space):
            expert_row = r
    if model_row is None:
        raise ValueError(f"Space '{space}' not found in attribute sheet '{attribute}'.")
    if expert_row is None:
        # allow plotting without expert line
        pass

    # Determine iterations dynamically from headers
    import re
    iters = sorted({int(m.group(1)) for h in headers for m in [re.match(r"it(\d+)_mean_score_1_10$", str(h))] if m})
    xs = iters or [1]
    means: List[float] = []
    errs: List[float] = []
    meds02: List[float] = []
    q1_02: List[float] = []
    q3_02: List[float] = []
    iqr02: List[float] = []
    for it in xs:
        mean_col = f"it{it}_mean_score_1_10"
        std_col = f"it{it}_std_score_1_10"
        med02_col = f"it{it}_median_rating_0_2"
        q1_col = f"it{it}_q1_rating_0_2"
        q3_col = f"it{it}_q3_rating_0_2"
        iqr02_col = f"it{it}_iqr_rating_0_2"
        m = model_row[idx.get(mean_col, -1)] if idx.get(mean_col) is not None else None
        v = model_row[idx.get(std_col, -1)] if idx.get(std_col) is not None else None
        m02 = model_row[idx.get(med02_col, -1)] if idx.get(med02_col, -1) is not None else None
        q1v = model_row[idx.get(q1_col, -1)] if idx.get(q1_col, -1) is not None else None
        q3v = model_row[idx.get(q3_col, -1)] if idx.get(q3_col, -1) is not None else None
        q02 = model_row[idx.get(iqr02_col, -1)] if idx.get(iqr02_col, -1) is not None else None
        means.append(float(m) if m is not None else float('nan'))
        errs.append(float(v) if v is not None else 0.0)
        meds02.append(float(m02) if m02 is not None else float('nan'))
        q1_02.append(float(q1v) if q1v is not None else float('nan'))
        q3_02.append(float(q3v) if q3v is not None else float('nan'))
        iqr02.append(float(q02) if q02 is not None else 0.0)

    expert_val = None
    expert02 = None
    if expert_row is not None:
        # Expert rows repeat the 1-10 score under each iteration's mean column; take it1 as canonical
        mcol = f"it1_mean_score_1_10"
        mv = expert_row[idx.get(mcol, -1)] if idx.get(mcol) is not None else None
        if mv is not None:
            expert_val = float(mv)
        rcol = f"it1_median_rating_0_2"
        rv = expert_row[idx.get(rcol, -1)] if idx.get(rcol) is not None else None
        if rv is not None:
            expert02 = float(rv)

    # Plot
    try:
        import matplotlib.pyplot as plt  # type: ignore
    except Exception as e:
        raise SystemExit(
            "Missing dependency: matplotlib. Install with `pip install matplotlib` or `pip install -r requirements.txt`."
        ) from e

    fig, ax = plt.subplots(figsize=(8, 5))
    # 1-10 series (mean +/- std)
    ax.errorbar(
        xs,
        means,
        yerr=errs,
        fmt='o-',
        capsize=5,
        color='tab:blue',
        label=_clean_label('ChatGPT 1-10 (+/- std)'),
    )
    # 0,1,2 series (median with Q1-Q3 whiskers when available; fallback: IQR/2)
    import math
    lower_err: List[float] = []
    upper_err: List[float] = []
    for m, q1, q3, iq in zip(meds02, q1_02, q3_02, iqr02):
        if not math.isnan(q1) and not math.isnan(q3):
            lower = max(0.0, m - q1)
            upper = max(0.0, q3 - m)
        else:
            half = (iq / 2.0) if iq is not None else 0.0
            lower = upper = half
        lower_err.append(lower)
        upper_err.append(upper)
    ax.errorbar(
        xs,
        meds02,
        yerr=[lower_err, upper_err],
        fmt='s-',
        capsize=5,
        color='tab:green',
        label=_clean_label('ChatGPT 0,1,2 median (Q1-Q3)'),
    )
    if expert_val is not None:
        ax.hlines(
            expert_val,
            xmin=min(xs),
            xmax=max(xs),
            colors='tab:red',
            linestyles='dashed',
            label=_clean_label('Expert 1-10'),
        )
    if expert02 is not None:
        ax.hlines(
            expert02,
            xmin=min(xs),
            xmax=max(xs),
            colors='tab:orange',
            linestyles='dashed',
            label=_clean_label('Expert 0,1,2'),
        )
    ax.set_xticks(xs)
    ax.set_xlim(min(xs) - 0.2, max(xs) + 0.2)
    ax.set_xlabel('Iteration')
    ax.set_ylabel(f'{attribute}')
    ax.set_ylim(0, 10)
    ax.set_yticks(list(range(0, 11)))
    label_space = display_map.get(space, space) if display_map else space
    ax.set_title(f"{label_space}")
    ax.grid(True, alpha=0.3)
    ax.legend(loc='best')
    fig.tight_layout()
    fig.savefig(out, dpi=200)
    plt.close(fig)


def plot_space_attribute_on_ax(
    ax,
    data: Dict[str, Any],
    attribute_label: str,
    space: str,
    show_ylabel: bool,
    show_xticks: bool,
    show_title: bool,
    add_legend: bool,
    show_xlabel_text: bool = False,
    xlabel_text: str = 'Iteration',
    display_map: Optional[Dict[str, str]] = None,
) -> None:
    """Low-level drawer used by the grid plot.

    Controls tick visibility and legend placement for compact layouts.
    """
    headers = data["headers"]
    idx = data["idx"]
    rows = data["rows"]

    # Find rows
    model_row: Optional[List[Any]] = None
    expert_row: Optional[List[Any]] = None
    for r in rows:
        raw = str(r[0]) if len(r) > 0 and r[0] is not None else ""
        label, is_expert = _strip_expert_suffix(raw)
        if not raw:
            continue
        if not is_expert and model_row is None and _labels_match(label, space):
            model_row = r
        elif is_expert and expert_row is None and _labels_match(label, space):
            expert_row = r
    if model_row is None:
        # Nothing to draw
        ax.set_visible(False)
        return

    import re
    iters = sorted({int(m.group(1)) for h in headers for m in [re.match(r"it(\d+)_mean_score_1_10$", str(h))] if m})
    xs = iters or [1]
    means: List[float] = []
    stds: List[float] = []
    meds02: List[float] = []
    q1_02: List[float] = []
    q3_02: List[float] = []
    iqr02: List[float] = []
    for it in xs:
        mean_col = f"it{it}_mean_score_1_10"
        std_col = f"it{it}_std_score_1_10"
        med02_col = f"it{it}_median_rating_0_2"
        q1_col = f"it{it}_q1_rating_0_2"
        q3_col = f"it{it}_q3_rating_0_2"
        iqr02_col = f"it{it}_iqr_rating_0_2"
        m = model_row[idx.get(mean_col, -1)] if idx.get(mean_col) is not None else None
        s = model_row[idx.get(std_col, -1)] if idx.get(std_col) is not None else None
        m02 = model_row[idx.get(med02_col, -1)] if idx.get(med02_col, -1) is not None else None
        q1v = model_row[idx.get(q1_col, -1)] if idx.get(q1_col, -1) is not None else None
        q3v = model_row[idx.get(q3_col, -1)] if idx.get(q3_col, -1) is not None else None
        q02 = model_row[idx.get(iqr02_col, -1)] if idx.get(iqr02_col, -1) is not None else None
        means.append(float(m) if m is not None else float('nan'))
        stds.append(float(s) if s is not None else 0.0)
        meds02.append(float(m02) if m02 is not None else float('nan'))
        q1_02.append(float(q1v) if q1v is not None else float('nan'))
        q3_02.append(float(q3v) if q3v is not None else float('nan'))
        iqr02.append(float(q02) if q02 is not None else 0.0)

    expert10 = None
    expert02 = None
    if expert_row is not None:
        mv = expert_row[idx.get("it1_mean_score_1_10", -1)] if idx.get("it1_mean_score_1_10") is not None else None
        rv = expert_row[idx.get("it1_median_rating_0_2", -1)] if idx.get("it1_median_rating_0_2") is not None else None
        if mv is not None:
            expert10 = float(mv)
        if rv is not None:
            expert02 = float(rv)

    # Draw
    ax.errorbar(
        xs, means, yerr=stds, fmt='o-', capsize=4, lw=1.6, color='tab:blue',
        label=_clean_label('ChatGPT 1-10 (+/- std)'),
    )
    # Asymmetric Q1-Q3 whiskers when available; fallback to IQR/2
    import math
    lower_err2: List[float] = []
    upper_err2: List[float] = []
    for m, q1, q3, iq in zip(meds02, q1_02, q3_02, iqr02):
        if not math.isnan(q1) and not math.isnan(q3):
            lower = max(0.0, m - q1)
            upper = max(0.0, q3 - m)
        else:
            half = (iq / 2.0) if iq is not None else 0.0
            lower = upper = half
        lower_err2.append(lower)
        upper_err2.append(upper)
    ax.errorbar(
        xs, meds02, yerr=[lower_err2, upper_err2], fmt='s-', capsize=4, lw=1.6, color='tab:green',
        label=_clean_label('ChatGPT 0,1,2 (Q1-Q3)'),
    )
    if expert10 is not None:
        ax.hlines(expert10, xmin=min(xs), xmax=max(xs), colors='tab:red', linestyles='dashed', lw=1.2, label=_clean_label('Expert 1-10'))
    if expert02 is not None:
        ax.hlines(expert02, xmin=min(xs), xmax=max(xs), colors='tab:orange', linestyles='dashed', lw=1.2, label=_clean_label('Expert 0,1,2'))
    # Axes cosmetics
    ax.set_xlim(min(xs) - 0.2, max(xs) + 0.2)
    ax.set_ylim(-0.5, 10)
    ax.set_xticks(xs)
    ax.set_yticks(list(range(0, 11)))
    # Control tick labels and axis label independently
    ax.tick_params(labelbottom=show_xticks)
    ax.set_xlabel(xlabel_text if show_xlabel_text else '')
    if show_ylabel:
        ax.set_ylabel(f'{_two_lines(attribute_label)}')
        ax.tick_params(labelleft=True)
    else:
        ax.set_ylabel('')
        ax.tick_params(labelleft=False)
    if show_title:
        label_space = display_map.get(space, space) if display_map else space
        ax.set_title(_two_lines(_clean_label(label_space)))
    ax.grid(True, alpha=0.25)
    if add_legend:
        ax.legend(loc='best', fontsize=8)


def plot_grid(xlsx: Path, out: Path, attributes: List[str], spaces: List[str], display_map: Optional[Dict[str, str]] = None) -> None:
    """Render an attributes Ã— spaces grid of small multiples from the consolidated workbook."""
    try:
        import matplotlib.pyplot as plt  # type: ignore
    except Exception as e:
        raise SystemExit(
            "Missing dependency: matplotlib. Install with `pip install matplotlib` or `pip install -r requirements.txt`."
        ) from e

    # Preload sheets per attribute
    sheets: Dict[str, Dict[str, Any]] = {}
    for attr in attributes:
        sheets[attr] = load_attribute_sheet(xlsx, attr)

    n_rows = len(attributes)
    n_cols = len(spaces)
    fig, axes = plt.subplots(n_rows, n_cols, figsize=(2.8*n_cols, 2.6*n_rows), sharex=True, sharey=True)

    for r, attr in enumerate(attributes):
        for c, space in enumerate(spaces):
            ax = axes[r][c] if n_rows > 1 else axes[c]
            show_ylabel = (c == 0)
            show_xticks = (r == n_rows - 1)
            show_title = (r == 0)
            add_legend = (r == 0 and c == 0)
            # Only center of bottom row shows the x-axis label text
            mid_col = len(spaces) // 2
            show_xlabel_text = (r == n_rows - 1 and c == mid_col)
            plot_space_attribute_on_ax(
                ax,
                sheets[attr],
                sheets[attr]["title"],
                space,
                show_ylabel,
                show_xticks,
                show_title,
                add_legend,
                show_xlabel_text=show_xlabel_text,
                xlabel_text='Prompt Iteration',
                display_map=display_map,
            )

    # Tighter outer margins and gutters to reduce unused whitespace
    plt.subplots_adjust(left=0.07, right=0.995, top=0.94, bottom=0.05, wspace=0.06, hspace=0.08)
    fig.savefig(out, dpi=200)
    plt.close(fig)


def main(argv: Optional[List[str]] = None) -> int:  # type: ignore[name-defined]
    import sys
    parser = argparse.ArgumentParser(description='Plot ratings from consolidated workbook')
    parser.add_argument('--xlsx', required=True, help='Path to consolidated workbook (combined_summary.xlsx)')
    parser.add_argument('--names-xlsx', help='Optional workbook with Spaces sheet (col1=original, col2=display)')
    # Single plot options
    parser.add_argument('--attribute', help='Attribute/sheet name (e.g., Color)')
    parser.add_argument('--space', help='Built space name (row label)')
    parser.add_argument('--out', default='plot.png', help='Output image path')
    # Grid options
    parser.add_argument('--grid', action='store_true', help='Render grid of plots (attributes Ã— spaces)')
    parser.add_argument('--attr-list', nargs='*', help='Attributes in order (default: all sheets)')
    parser.add_argument('--space-list', nargs='*', help='Spaces in order (default: all spaces)')
    # Error bars use standard deviation from the consolidated workbook
    args = parser.parse_args(argv or sys.argv[1:])

    xlsx = Path(args.xlsx)
    display_map: Dict[str, str] = {}
    spaces_order: List[str] = []
    if args.names_xlsx:
        spaces_order, display_map = _read_space_order_and_map(Path(args.names_xlsx))
    if args.grid:
        attrs = args.attr_list if args.attr_list else None
        spaces = args.space_list if args.space_list else None
        if attrs is None:
            # Use all attribute sheets (excluding index)
            try:
                from openpyxl import load_workbook  # type: ignore
                wb = load_workbook(filename=str(xlsx), read_only=True, data_only=True)
                attrs = [s for s in wb.sheetnames if s.lower() != 'index']
            except Exception:
                attrs = []
        if spaces is None:
            if spaces_order:
                spaces = spaces_order
            else:
                # Derive spaces from the first attribute sheet, ignoring Expert rows
                try:
                    from openpyxl import load_workbook  # type: ignore
                    wb = load_workbook(filename=str(xlsx), read_only=True, data_only=True)
                    first = [s for s in wb.sheetnames if s.lower() != 'index'][0]
                    ws = wb[first]
                    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                    # spaces are in column 1 of data rows
                    spaces = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row and row[0]:
                            label, is_expert = _strip_expert_suffix(str(row[0]))
                            if is_expert:
                                continue
                            spaces.append(label)
                except Exception:
                    spaces = []
        plot_grid(xlsx, Path(args.out), attrs, spaces, display_map=display_map)
        print(f"Saved {args.out}")
        return 0

    if not args.attribute or not args.space:
        raise SystemExit("For a single plot, provide both --attribute and --space. Or use --grid.")
    plot_space_attribute(xlsx, args.attribute, args.space, Path(args.out), display_map=display_map)
    print(f"Saved {args.out}")
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
