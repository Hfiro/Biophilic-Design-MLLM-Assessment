"""
Author: Hoda Alem
Date: November 2025
Context: Part of PhD research for a PhD in Environmental Design and Planning
         at Myers-Lawson School of Construction, Virginia Tech, Blacksburg, VA.
Purpose: Compute 1â€“10 accuracy matrices per iteration with a tolerance window
         (e.g., Â±2 from Expert) and export as an Excel workbook.
Features:
- Preserves attribute/space order from the Expert sheet.
- Parameterized correctness tolerance (default Â±2).
- For each iteration, produces a 5Ã—5 matrix of percentages plus row/column
  aggregates and a grand total cell.
- Simple CLI for globbing result workbooks and choosing output path.
"""

import argparse
import glob
import math
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple


def _normalize_tokens(s: str) -> List[str]:
    import re
    s = (s or "").lower()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return [t for t in s.split() if t]


def _norm_key(s: str) -> str:
    return " ".join(_normalize_tokens(s))


def _resolve_sheet_name(wb, desired: str) -> Optional[str]:
    if desired in wb.sheetnames:
        return desired
    tokens = set(_normalize_tokens(desired))
    cands = [s for s in wb.sheetnames if s.lower() != 'index']
    best_subset = [s for s in cands if tokens and tokens.issubset(set(_normalize_tokens(s)))]
    if best_subset:
        return sorted(best_subset, key=lambda x: (len(_normalize_tokens(x)), x))[0]
    scored = sorted(((len(tokens & set(_normalize_tokens(s))), s) for s in cands), reverse=True)
    if scored and scored[0][0] > 0:
        return scored[0][1]
    return None


def _load_wb(path: Path):
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:
        raise SystemExit(
            "Missing dependency: openpyxl. Install with `pip install openpyxl` or `pip install -r requirements.txt`."
        ) from e
    return load_workbook(filename=str(path), read_only=True, data_only=True)


def _collect_data_rows(ws) -> List[Dict[str, Any]]:
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers: Dict[str, int] = {}
    for i, v in enumerate(header):
        if isinstance(v, str):
            headers[v.strip()] = i
    data: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        rec: Dict[str, Any] = {}
        for k, idx in headers.items():
            rec[k] = row[idx] if idx < len(row) else None
        if isinstance(rec.get("run"), (int, float)):
            data.append(rec)
    return data


def _read_expert_order_and_values_10(path: Path, sheet_name: str = "Expert") -> Tuple[List[str], List[str], Dict[str, Dict[str, float]]]:
    """Return (spaces_order, attributes_order, expert10_map) for 1â€“10 scores.

    Expert 1â€“10 values are in the second column of each space pair (c+1).
    """
    wb = _load_wb(path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {path}")
    ws = wb[sheet_name]

    header1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))

    # Ordered spaces from row 1 starting at column 3 (every 2 columns)
    spaces_order: List[str] = []
    start_col = 3
    c = start_col
    while c <= ws.max_column:
        name = str(header1[c - 1] or '').strip()
        if name and (not spaces_order or spaces_order[-1] != name):
            spaces_order.append(name)
        c += 2

    # Attributes order from column 2 (row >= 3)
    attributes_order: List[str] = []
    r = 3
    while r <= ws.max_row:
        val = ws.cell(row=r, column=2).value
        if val is None:
            r += 1
            continue
        attr = str(val).strip()
        if attr:
            attributes_order.append(attr)
        r += 1

    # Expert 1â€“10 map (use the c+1 column)
    expert10: Dict[str, Dict[str, float]] = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        attr_raw = row[1] if len(row) > 1 else None
        if not attr_raw:
            continue
        attribute = str(attr_raw).strip()
        c = start_col
        s_idx = 0
        while c <= ws.max_column and s_idx < len(spaces_order):
            space = spaces_order[s_idx]
            c10 = c + 1
            v10 = row[c10 - 1] if c10 - 1 < len(row) else None
            try:
                if v10 is None or (isinstance(v10, float) and math.isnan(v10)):
                    val10 = None
                elif isinstance(v10, (int, float)):
                    val10 = float(v10)
                else:
                    s10 = str(v10).strip()
                    val10 = float(s10) if s10 else None
            except Exception:
                val10 = None
            if val10 is not None:
                expert10.setdefault(space, {})[attribute] = val10
            c += 2
            s_idx += 1

    return spaces_order, attributes_order, expert10


def _space_from_results(path: Path) -> Optional[str]:
    wb = _load_wb(path)
    if 'meta' not in wb.sheetnames:
        return None
    ws = wb['meta']
    for r in ws.iter_rows(min_row=1, values_only=True):
        if r and str(r[0]).strip().lower() == 'space':
            return str(r[1]).strip() if r[1] else None
    return None


def _gather_result_paths_by_space(paths: List[Path]) -> Dict[str, Path]:
    out: Dict[str, Path] = {}
    for p in paths:
        s = _space_from_results(p)
        if s:
            out[s] = p
    return out


def _count_matches_for_score(ws, iteration: int, expert_val: float, tolerance: float = 2.0) -> Tuple[int, int]:
    rows = _collect_data_rows(ws)
    matched = 0
    total = 0
    for rec in rows:
        try:
            it = int(rec.get('iteration'))
        except Exception:
            continue
        if it != iteration:
            continue
        sv = rec.get('score_1_10')
        if isinstance(sv, (int, float)):
            total += 1
            if abs(float(sv) - float(expert_val)) <= tolerance:
                matched += 1
    return matched, total


def compute_accuracy_matrix_10(
    result_paths: List[Path],
    expert_xlsx: Path,
    iterations: List[int],
    tolerance: float = 2.0,
) -> Tuple[List[str], List[str], Dict[int, Dict[str, Dict[str, Tuple[int, int]]]]]:
    """Return (spaces, attributes, per-iteration matrices of (matched, total)).

    A per-cell count is considered correct if abs(model_score - expert_score)
    is within the given tolerance.
    """
    spaces, attributes, expert10 = _read_expert_order_and_values_10(expert_xlsx)
    by_space = _gather_result_paths_by_space(result_paths)
    norm_space_paths = {_norm_key(name): path for name, path in by_space.items()}

    out: Dict[int, Dict[str, Dict[str, Tuple[int, int]]]] = {it: {a: {} for a in attributes} for it in iterations}
    for space in spaces:
        path = by_space.get(space)
        if not path:
            path = norm_space_paths.get(_norm_key(space))
        if not path:
            continue
        wb = _load_wb(path)
        for attr in attributes:
            exp_map = expert10.get(space)
            if exp_map is None:
                norm_space = _norm_key(space)
                for candidate_space, candidate_map in expert10.items():
                    if _norm_key(candidate_space) == norm_space:
                        exp_map = candidate_map
                        break
            if not exp_map:
                continue
            expert_val = exp_map.get(attr)
            if expert_val is None:
                norm_attr = _norm_key(attr)
                for attr_name, val in exp_map.items():
                    if _norm_key(attr_name) == norm_attr:
                        expert_val = val
                        break
            if expert_val is None:
                continue
            sheet_name = _resolve_sheet_name(wb, attr)
            if not sheet_name:
                continue
            ws = wb[sheet_name]
            for it in iterations:
                m, t = _count_matches_for_score(ws, it, expert_val, tolerance=tolerance)
                out[it][attr][space] = (m, t)
    return spaces, attributes, out

def _available_iterations(paths: List[Path]) -> List[int]:
    """Scan all result workbooks to collect the set of iteration indices present."""
    iters: set[int] = set()
    for p in paths:
        wb = _load_wb(p)
        for sheet in wb.sheetnames:
            if sheet in {"meta", "prompts"}:
                continue
            ws = wb[sheet]
            rows = _collect_data_rows(ws)
            for rec in rows:
                try:
                    it = int(rec.get('iteration'))
                    iters.add(it)
                except Exception:
                    continue
    return sorted(iters)


def _fmt_pct(matched: int, total: int) -> float:
    if total <= 0:
        return 0.0
    return round(100.0 * matched / total, 1)


def export_accuracy_xlsx(
    spaces: List[str],
    attributes: List[str],
    matrices: Dict[int, Dict[str, Dict[str, Tuple[int, int]]]],
    out_xlsx: Path,
) -> None:
    try:
        from openpyxl import Workbook
    except Exception as e:
        raise SystemExit(
            "Missing dependency: openpyxl. Install with `pip install openpyxl` or `pip install -r requirements.txt`."
        ) from e

    wb = Workbook()
    ws0 = wb.active
    wb.remove(ws0)

    for it, attr_map in matrices.items():
        ws = wb.create_sheet(f"Iteration {it}")
        header = ["Attribute"] + spaces + ["All Spaces"]
        ws.append(header)

        col_totals: List[Tuple[int, int]] = [(0, 0) for _ in spaces]
        grand_m = 0
        grand_t = 0

        for attr in attributes:
            row_vals: List[Any] = [attr]
            rm = 0
            rt = 0
            for si, space in enumerate(spaces):
                m, t = attr_map.get(attr, {}).get(space, (0, 0))
                rm += m
                rt += t
                cm, ct = col_totals[si]
                col_totals[si] = (cm + m, ct + t)
                row_vals.append(_fmt_pct(m, t))
            row_vals.append(_fmt_pct(rm, rt))
            grand_m += rm
            grand_t += rt
            ws.append(row_vals)

        footer: List[Any] = ["All Attributes"]
        for (cm, ct) in col_totals:
            footer.append(_fmt_pct(cm, ct))
        footer.append(_fmt_pct(grand_m, grand_t))
        ws.append(footer)

        try:
            for col in ws.columns:
                max_len = 0
                col_letter = getattr(col[0], "column_letter", None)
                for cell in col:
                    val = cell.value
                    if val is None:
                        continue
                    max_len = max(max_len, len(str(val)))
                if col_letter:
                    ws.column_dimensions[col_letter].width = min(max_len + 2, 40)
        except Exception:
            pass

    wb.save(out_xlsx)


def main(argv: Optional[List[str]] = None) -> int:
    import sys
    ap = argparse.ArgumentParser(description="Aggregate 1â€“10 accuracy (within Â±2) per iteration into a 5Ã—5 matrix with aggregates")
    ap.add_argument('--glob', default='results_*.xlsx', help='Glob for per-space result workbooks')
    ap.add_argument('--xlsx', default='BD GPT Prompts.xlsx', help='Workbook containing Expert sheet')
    ap.add_argument('--out', default='accuracy_by_iteration_1to10.xlsx', help='Output Excel workbook')
    ap.add_argument('--tolerance', type=float, default=2.0, help='Tolerance for correctness (absolute difference)')
    args = ap.parse_args(argv or sys.argv[1:])

    result_paths = [Path(p) for p in sorted(glob.glob(args.glob)) if p.lower().endswith('.xlsx')]
    if not result_paths:
        raise SystemExit('No result files matched.')

    iterations = _available_iterations(result_paths)
    spaces, attributes, matrices = compute_accuracy_matrix_10(result_paths, Path(args.xlsx), iterations, tolerance=args.tolerance)
    export_accuracy_xlsx(spaces, attributes, matrices, Path(args.out))
    print(f"Wrote {args.out}")
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
