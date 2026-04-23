"""
Scatter all per-run results across all result workbooks.

Features:
- x-axis: 0/1/2 rating (with jitter)
- y-axis: 1..10 score
- color by built space
- optional marker shape by attribute
- optional in-point iteration labels

Usage examples:
- python plot_scatter_all.py --glob "results_*.xlsx" --out scatter_all.png
- python plot_scatter_all.py --glob "results_*.xlsx" --out scatter_all.png --label-iteration --markers-by-attribute --jitter 0.15 --alpha 0.6
"""
# Author: Hoda Alem
# Date: November 2025
# Context: Part of PhD research for a PhD in Environmental Design and Planning
#          at Myers-Lawson School of Construction, Virginia Tech, Blacksburg, VA.

import argparse
import math
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple


def _load_wb(path: Path):
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:
        raise SystemExit(
            "Missing dependency: openpyxl. Install with `pip install openpyxl` or use requirements.txt."
        ) from e
    return load_workbook(filename=str(path), read_only=True, data_only=True)


def _headers_from_sheet(ws) -> Dict[str, int]:
    row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers: Dict[str, int] = {}
    for i, v in enumerate(row):
        if isinstance(v, str):
            headers[v.strip()] = i
    return headers


def _is_number(x: Any) -> bool:
    return isinstance(x, (int, float)) and not (isinstance(x, float) and math.isnan(x))


def _pearson_r(xs: List[float], ys: List[float]) -> Optional[float]:
    if len(xs) < 2:
        return None
    mean_x = sum(xs) / len(xs)
    mean_y = sum(ys) / len(ys)
    num = sum((x - mean_x) * (y - mean_y) for x, y in zip(xs, ys))
    den_x = sum((x - mean_x) ** 2 for x in xs)
    den_y = sum((y - mean_y) ** 2 for y in ys)
    den = math.sqrt(den_x * den_y)
    if den == 0:
        return None
    return num / den

def _collect_points(paths: List[Path]) -> List[Dict[str, Any]]:
    points: List[Dict[str, Any]] = []
    for p in paths:
        wb = _load_wb(p)
        # space name from meta sheet if present
        space_name: Optional[str] = None
        try:
            if 'meta' in wb.sheetnames:
                ws_meta = wb['meta']
                for r in ws_meta.iter_rows(min_row=1, values_only=True):
                    if r and str(r[0]).strip().lower() == 'space':
                        space_name = str(r[1]).strip() if len(r) > 1 and r[1] is not None else None
                        break
        except Exception:
            pass
        for sheet in wb.sheetnames:
            if sheet in {"meta", "prompts"}:
                continue
            ws = wb[sheet]
            idx = _headers_from_sheet(ws)
            if not idx:
                continue
            def val(row, key: str):
                i = idx.get(key)
                return row[i] if i is not None and i < len(row) else None
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row is None:
                    continue
                rv = val(row, 'run')
                if not _is_number(rv):
                    continue
                r02 = val(row, 'rating_0_2')
                s10 = val(row, 'score_1_10')
                if not _is_number(r02) or not _is_number(s10):
                    continue
                it = val(row, 'iteration')
                attr = val(row, 'attribute') or sheet
                sp = val(row, 'space') or space_name or p.stem
                points.append({
                    'file': p.name,
                    'sheet': sheet,
                    'space': str(sp),
                    'attribute': str(attr),
                    'iteration': int(it) if isinstance(it, (int, float)) else it,
                    'run': int(rv),
                    'rating_0_2': float(r02),
                    'score_1_10': float(s10),
                })
    return points


def _distinct_markers(n: int) -> List[str]:
    # A small cycle of distinct markers for attributes
    base = ['o', 's', '^', 'D', 'P', 'X', 'v', '<', '>', 'h']
    if n <= len(base):
        return base[:n]
    # repeat if more attributes than markers
    out: List[str] = []
    while len(out) < n:
        out.extend(base)
    return out[:n]


def _read_spaces_order_and_map(names_xlsx: Optional[Path]) -> Tuple[List[str], Dict[str, str]]:
    order: List[str] = []
    mapping: Dict[str, str] = {}
    if not names_xlsx:
        return order, mapping
    try:
        from openpyxl import load_workbook  # type: ignore
        if not names_xlsx.exists():
            return order, mapping
        wb = load_workbook(filename=str(names_xlsx), read_only=True, data_only=True)
        if 'Spaces' not in wb.sheetnames:
            return order, mapping
        ws = wb['Spaces']
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r:
                continue
            orig = str(r[0]).strip() if r[0] is not None else ''
            if not orig:
                continue
            order.append(orig)
            if len(r) >= 2 and r[1] is not None:
                disp = str(r[1]).strip()
                if disp:
                    mapping[orig] = disp
    except Exception:
        return [], {}
    return order, mapping


def plot_scatter(points: List[Dict[str, Any]], out: Path, jitter: float, y_jitter: float, alpha: float, size: float, label_iteration: bool, markers_by_attribute: bool, seed: Optional[int], dpi: int, highlight_suspicious: bool, suspect_tol: float, suspicious_csv: Optional[Path], fit_line: bool, spaces_order: Optional[List[str]] = None, display_map: Optional[Dict[str, str]] = None) -> None:
    if not points:
        raise SystemExit('No points to plot (no per-run rows with both rating_0_2 and score_1_10).')
    try:
        import matplotlib.pyplot as plt  # type: ignore
        import numpy as np  # type: ignore
    except Exception as e:
        raise SystemExit("Missing dependency: matplotlib (and numpy). Install with `pip install matplotlib numpy`.") from e

    if seed is not None:
        np.random.seed(int(seed))

    # Encode color by space (respect Spaces sheet order if provided)
    spaces_set = {pt['space'] for pt in points}
    if spaces_order:
        spaces = [s for s in spaces_order if s in spaces_set]
        extras = sorted(list(spaces_set - set(spaces)))
        spaces.extend(extras)
    else:
        spaces = sorted(spaces_set)
    cmap = plt.get_cmap('tab20')
    color_map: Dict[str, Tuple[float, float, float, float]] = {s: cmap((i % 20) / 20.0) for i, s in enumerate(spaces)}

    # Optional markers by attribute
    attr_mark_map: Dict[str, str] = {}
    if markers_by_attribute:
        attrs = sorted({pt['attribute'] for pt in points})
        markers = _distinct_markers(len(attrs))
        attr_mark_map = {a: m for a, m in zip(attrs, markers)}

    fig, ax = plt.subplots(figsize=(10, 6))

    # Group by space (for legend coloring) and attribute (for marker)
    for s in spaces:
        pts_s = [pt for pt in points if pt['space'] == s]
        # Further split by attribute if markers vary
        if markers_by_attribute:
            attrs_s = sorted({pt['attribute'] for pt in pts_s})
            for a in attrs_s:
                pts_sa = [pt for pt in pts_s if pt['attribute'] == a]
                x = np.array([pt['rating_0_2'] for pt in pts_sa], dtype=float)
                y = np.array([pt['score_1_10'] for pt in pts_sa], dtype=float)
                if jitter > 0:
                    x = x + np.random.normal(loc=0.0, scale=jitter, size=x.shape)
                if y_jitter > 0:
                    y = y + np.random.normal(loc=0.0, scale=y_jitter, size=y.shape)
                label_s = display_map.get(s, s) if display_map else s
                ax.scatter(x, y, s=size, alpha=alpha, color=color_map[s], edgecolors='none', marker=attr_mark_map.get(a, 'o'), label=f"{label_s} / {a}")
                if label_iteration:
                    for xi, yi, it in zip(x, y, [pt['iteration'] for pt in pts_sa]):
                        try:
                            ax.text(xi, yi, str(it), fontsize=7, ha='center', va='center', color='black')
                        except Exception:
                            pass
        else:
            x = np.array([pt['rating_0_2'] for pt in pts_s], dtype=float)
            y = np.array([pt['score_1_10'] for pt in pts_s], dtype=float)
            if jitter > 0:
                x = x + np.random.normal(loc=0.0, scale=jitter, size=x.shape)
            if y_jitter > 0:
                y = y + np.random.normal(loc=0.0, scale=y_jitter, size=y.shape)
            label_s = display_map.get(s, s) if display_map else s
            ax.scatter(x, y, s=size, alpha=alpha, color=color_map[s], edgecolors='none', marker='o', label=label_s)
            if label_iteration:
                for xi, yi, it in zip(x, y, [pt['iteration'] for pt in pts_s]):
                    try:
                        ax.text(xi, yi, str(it), fontsize=7, ha='center', va='center', color='black')
                    except Exception:
                        pass

    # Find suspicious points
    # Criteria: rating in {0,1,2} and score is near any target in this map (within suspect_tol)
    # Expanded targets include: (0,3), (0,4), (0,5), (1,8), (2,5), (2,6) in addition to (1,1), (2,2)
    target_map = {
        0.0: [3.0, 4.0, 5.0],
        1.0: [1.0, 8.0],
        2.0: [2.0, 5.0, 6.0],
    }
    susp: List[Dict[str, Any]] = []
    for pt in points:
        r = float(pt['rating_0_2'])
        s = float(pt['score_1_10'])
        targets = target_map.get(r)
        if not targets:
            continue
        if any(abs(s - t) <= suspect_tol for t in targets):
            susp.append(pt)
    if highlight_suspicious and susp:
        xs = np.array([pt['rating_0_2'] for pt in susp], dtype=float)
        ys = np.array([pt['score_1_10'] for pt in susp], dtype=float)
        if jitter > 0:
            xs = xs + np.random.normal(loc=0.0, scale=jitter, size=xs.shape)
        if y_jitter > 0:
            ys = ys + np.random.normal(loc=0.0, scale=y_jitter, size=ys.shape)
        ax.scatter(xs, ys, s=max(size, 36), facecolors='none', edgecolors='red', linewidths=1.2, label='Suspicious (scoreâ‰ˆrating)')

    # Fit a global line (y vs x) across all points (without jitter)
    if fit_line:
        import numpy as _np
        x_all = _np.array([pt['rating_0_2'] for pt in points], dtype=float)
        y_all = _np.array([pt['score_1_10'] for pt in points], dtype=float)
        try:
            m, b = _np.polyfit(x_all, y_all, deg=1)
            x_line = _np.array([0.0, 2.0])
            y_line = m * x_line + b
            ax.plot(x_line, y_line, color='black', linewidth=2.0, alpha=0.9, label='Fit')
        except Exception:
            pass

    if suspicious_csv is not None:
        import csv
        with suspicious_csv.open('w', newline='', encoding='utf-8') as f:
            w = csv.writer(f)
            w.writerow(['file', 'sheet', 'space', 'attribute', 'iteration', 'run', 'rating_0_2', 'score_1_10'])
            for pt in susp:
                w.writerow([pt['file'], pt['sheet'], pt['space'], pt['attribute'], pt['iteration'], pt['run'], pt['rating_0_2'], pt['score_1_10']])

    # Axes cosmetics
    ax.set_xlabel("ChatGPT's Ratings (0,1,2)")
    ax.set_ylabel("ChatGPT's Ratings (1-10)")
    ax.set_xlim(-0.5, 2.5)
    ax.set_xticks([0, 1, 2])
    ax.set_ylim(0, 10.5)
    ax.set_yticks(list(range(0, 11)))
    ax.grid(True, alpha=0.25)
    # Correlation label (using raw, non-jittered values)
    raw_xs = [pt['rating_0_2'] for pt in points]
    raw_ys = [pt['score_1_10'] for pt in points]
    r = _pearson_r(raw_xs, raw_ys)
    r_txt = "Correlation: NA" if r is None else f"Correlation: {r:.2f}"
    ax.text(0.98, 0.02, r_txt, transform=ax.transAxes, ha='right', va='bottom', fontsize=10)
    # Manage legend size
    if markers_by_attribute:
        ax.legend(loc='best', fontsize=8)
    else:
        ax.legend(loc='best', fontsize=9)
    fig.tight_layout()
    fig.savefig(out, dpi=dpi)
    plt.close(fig)


def plot_scatter_facet(
    points: List[Dict[str, Any]],
    out: Path,
    facet: str,  # 'space' or 'attribute'
    jitter: float,
    y_jitter: float,
    alpha: float,
    size: float,
    label_iteration: bool,
    seed: Optional[int],
    dpi: int,
    ncols: int,
    panel_w: float,
    panel_h: float,
    fit_line: bool,
    spaces_order: Optional[List[str]] = None,
    display_map: Optional[Dict[str, str]] = None,
    legend_loc: str = 'bottom',
    no_legend: bool = False,
) -> None:
    if not points:
        raise SystemExit('No points to plot.')
    try:
        import matplotlib.pyplot as plt  # type: ignore
        import numpy as np  # type: ignore
    except Exception as e:
        raise SystemExit("Missing dependency: matplotlib (and numpy). Install with `pip install matplotlib numpy`.") from e

    if seed is not None:
        np.random.seed(int(seed))

    if facet == 'space':
        spaces_set = {pt['space'] for pt in points}
        if spaces_order:
            panels = [s for s in spaces_order if s in spaces_set]
            extras = sorted(list(spaces_set - set(panels)))
            panels.extend(extras)
        else:
            panels = sorted(spaces_set)
        color_keys = sorted({pt['attribute'] for pt in points})
        color_label = 'Attribute'
        def panel_filter(p):
            return [pt for pt in points if pt['space'] == p]
        def color_of(pt):
            return pt['attribute']
    else:  # facet by attribute
        panels = sorted({pt['attribute'] for pt in points})
        spaces_set = {pt['space'] for pt in points}
        if spaces_order:
            ordered = [s for s in spaces_order if s in spaces_set]
            extras = sorted(list(spaces_set - set(ordered)))
            color_keys = ordered + extras
        else:
            color_keys = sorted(spaces_set)
        color_label = 'Space'
        def panel_filter(p):
            return [pt for pt in points if pt['attribute'] == p]
        def color_of(pt):
            return pt['space']

    cmap = plt.get_cmap('tab20')
    color_map: Dict[str, Tuple[float, float, float, float]] = {k: cmap((i % 20) / 20.0) for i, k in enumerate(color_keys)}

    n = len(panels)
    ncols = max(1, int(ncols))
    nrows = (n + ncols - 1) // ncols
    fig, axes = plt.subplots(nrows, ncols, figsize=(panel_w * ncols, panel_h * nrows), sharex=True, sharey=True)
    # Normalize axes indexing
    if nrows == 1 and ncols == 1:
        axes = [[axes]]
    elif nrows == 1:
        axes = [list(axes)]
    elif ncols == 1:
        axes = [[ax] for ax in axes]

    # plot each panel
    for i, pkey in enumerate(panels):
        r = i // ncols
        c = i % ncols
        ax = axes[r][c]
        pts = panel_filter(pkey)
        # group by color key
        for ck in color_keys:
            pts_c = [pt for pt in pts if color_of(pt) == ck]
            if not pts_c:
                continue
            x = np.array([pt['rating_0_2'] for pt in pts_c], dtype=float)
            y = np.array([pt['score_1_10'] for pt in pts_c], dtype=float)
            if jitter > 0:
                x = x + np.random.normal(loc=0.0, scale=jitter, size=x.shape)
            if y_jitter > 0:
                y = y + np.random.normal(loc=0.0, scale=y_jitter, size=y.shape)
            ax.scatter(x, y, s=size, alpha=alpha, color=color_map[ck], edgecolors='none', marker='o', label=str(ck))
            if label_iteration:
                for xi, yi, it in zip(x, y, [pt['iteration'] for pt in pts_c]):
                    try:
                        ax.text(xi, yi, str(it), fontsize=6, ha='center', va='center', color='black')
                    except Exception:
                        pass

        # optional local fit line
        if fit_line and pts:
            try:
                x_all = np.array([pt['rating_0_2'] for pt in pts], dtype=float)
                y_all = np.array([pt['score_1_10'] for pt in pts], dtype=float)
                m, b = np.polyfit(x_all, y_all, deg=1)
                ax.plot([0.0, 2.0], [m*0.0 + b, m*2.0 + b], color='black', lw=1.5, alpha=0.9)
            except Exception:
                pass

        title = display_map.get(pkey, pkey) if (display_map and facet == 'space') else str(pkey)
        ax.set_title(title)
        ax.set_xlim(-0.5, 2.5)
        ax.set_xticks([0, 1, 2])
        ax.set_ylim(0, 10.5)
        ax.grid(True, alpha=0.2)
        # Panel correlation and label
        raw_xs = [pt['rating_0_2'] for pt in pts]
        raw_ys = [pt['score_1_10'] for pt in pts]
        r = _pearson_r(raw_xs, raw_ys)
        r_txt = "Correlation: NA" if r is None else f"Correlation: {r:.2f}"
        label_txt = f"({chr(ord('a') + i)}) {r_txt}"
        ax.text(0.02, 0.96, label_txt, transform=ax.transAxes, ha='left', va='top', fontsize=10)

    # remove unused axes
    for j in range(n, nrows * ncols):
        r = j // ncols
        c = j % ncols
        axes[r][c].set_visible(False)

    # legend and shared labels
    if not no_legend:
        handles = []
        labels = []
        for k in color_keys:
            lab = display_map.get(k, k) if (display_map and color_label.lower() == 'space') else str(k)
            # Smaller for space facet (more entries), larger for attribute facet
            legend_ms = 2.5 if facet == 'space' else 4.0
            handles.append(plt.Line2D([0], [0], marker='o', color='none', markerfacecolor=color_map[k], markersize=legend_ms, label=lab))
            labels.append(lab)
        legend_cols = min(len(color_keys), 10)
        if legend_loc == 'right':
            # place legend to the right, outside subplots
            fig.legend(
                handles,
                labels,
                loc='center left',
                ncol=1,
                title=color_label,
                fontsize=(5 if facet == 'space' else 7),
                frameon=True,
                bbox_to_anchor=(0.98, 0.5),
                borderaxespad=0.4,
                handlelength=1.0,
                columnspacing=0.8,
                handletextpad=0.4,
            )
            # reserve right margin for legend; place figure-level labels
            fig.subplots_adjust(left=0.07, right=0.86, top=0.98, bottom=0.11)
            try:
                fig.supxlabel("ChatGPT's Ratings (0,1,2)", y=0.055, fontsize=9)
                fig.supylabel("ChatGPT's Ratings (1-10)", x=0.055, fontsize=9)
            except Exception:
                fig.text(0.5, 0.055, "ChatGPT's Ratings (0,1,2)", ha='center', fontsize=9)
                fig.text(0.055, 0.5, "ChatGPT's Ratings (1-10)", va='center', rotation='vertical', fontsize=9)
        else:
            # bottom-center legend outside subplots; fit width by limiting columns
            legend_cols = min(len(color_keys), 8 if facet == 'attribute' else 6)
            fig.legend(
                handles,
                labels,
                loc='lower center',
                ncol=legend_cols,
                title=color_label,
                fontsize=(5 if facet == 'space' else 7),
                frameon=True,
                bbox_to_anchor=(0.5, 0.02),
                borderaxespad=0.18,
                handlelength=(0.8 if facet == 'space' else 1.0),
                columnspacing=(0.5 if facet == 'space' else 0.7),
                handletextpad=(0.18 if facet == 'space' else 0.25),
                labelspacing=(0.12 if facet == 'space' else 0.20),
            )
            # reserve space: shrink plot area a bit more for balance
            if facet == 'attribute':
                fig.subplots_adjust(left=0.07, right=0.98, top=0.93, bottom=0.16)
            else:
                fig.subplots_adjust(left=0.06, right=0.98, top=0.95, bottom=0.13)
            try:
                # Push x-axis title up and move y-label slightly closer to plot
                fig.supxlabel("ChatGPT's Ratings (0,1,2)", y=(0.09 if facet == 'attribute' else 0.085), fontsize=9)
                ylabel_x = (0.026 if facet == 'attribute' else 0.03)
                fig.supylabel("ChatGPT's Ratings (1-10)", x=ylabel_x, fontsize=9)
            except Exception:
                fig.text(0.5, (0.09 if facet == 'attribute' else 0.085), "ChatGPT's Ratings (0,1,2)", ha='center', fontsize=9)
                ylabel_x = (0.026 if facet == 'attribute' else 0.03)
                fig.text(ylabel_x, 0.5, "ChatGPT's Ratings (1-10)", va='center', rotation='vertical', fontsize=9)
    else:
        # no legend: use full width/height and place labels close to panels
        fig.subplots_adjust(left=0.07, right=0.98, top=0.98, bottom=0.11)
        try:
            fig.supxlabel("ChatGPT's Ratings (0,1,2)", y=0.055)
            fig.supylabel("ChatGPT's Ratings (1-10)", x=0.055)
        except Exception:
            fig.text(0.5, 0.055, "ChatGPT's Ratings (0,1,2)", ha='center')
            fig.text(0.055, 0.5, "ChatGPT's Ratings (1-10)", va='center', rotation='vertical')
    fig.savefig(out, dpi=dpi)
    plt.close(fig)


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description='Scatter all per-run points: x=0/1/2 (jitter), y=1â€“10, color by space, optional marker by attribute, optional iteration labels')
    parser.add_argument('--glob', default='results_*.xlsx', help='Glob for per-space result workbooks')
    parser.add_argument('--out', default='scatter_all.png', help='Output image path')
    parser.add_argument('--names-xlsx', help='Optional workbook with Spaces sheet for enforced order and display labels')
    parser.add_argument('--jitter', type=float, default=0.25, help='Std dev of normal jitter for x (0â€“2)')
    parser.add_argument('--y-jitter', type=float, default=0.35, help='Std dev of normal jitter for y (1â€“10)')
    parser.add_argument('--alpha', type=float, default=0.5, help='Marker alpha (lower = more transparent)')
    parser.add_argument('--size', type=float, default=16.0, help='Marker size (points^2)')
    parser.add_argument('--label-iteration', action='store_true', help='Draw iteration number on each point')
    parser.add_argument('--markers-by-attribute', action='store_true', help='Use different marker shapes per attribute')
    parser.add_argument('--seed', type=int, help='Random seed for jitter reproducibility')
    parser.add_argument('--dpi', type=int, default=180, help='Figure DPI')
    parser.add_argument('--no-fit-line', action='store_true', help='Disable global best-fit line')
    parser.add_argument('--facet', choices=['none', 'space', 'attribute'], default='none', help='Create a panel grid by space or by attribute')
    parser.add_argument('--ncols', type=int, default=3, help='Number of columns in facet grid')
    parser.add_argument('--panel-w', type=float, default=4.0, help='Facet panel width in inches')
    parser.add_argument('--panel-h', type=float, default=3.0, help='Facet panel height in inches')
    parser.add_argument('--legend-loc', choices=['right', 'bottom'], default='bottom', help='Legend placement for facet plots')
    parser.add_argument('--no-legend', action='store_true', help='Hide legend on facet plots')
    parser.add_argument('--highlight-suspicious', action='store_true', help='Highlight points where scoreâ‰ˆrating (1 or 2)')
    parser.add_argument('--suspect-tol', type=float, default=0.05, help='Tolerance for suspicious detection (|score-rating| â‰¤ tol)')
    parser.add_argument('--suspicious-csv', help='Optional CSV path to export suspicious points')
    args = parser.parse_args()

    import glob as _glob
    paths = [Path(p) for p in sorted(_glob.glob(args.glob)) if p.lower().endswith('.xlsx')]
    if not paths:
        raise SystemExit('No result files matched.')

    points = _collect_points(paths)
    spaces_order: List[str] = []
    display_map: Dict[str, str] = {}
    if args.names_xlsx:
        spaces_order, display_map = _read_spaces_order_and_map(Path(args.names_xlsx))
    suspicious_path = Path(args.suspicious_csv) if args.suspicious_csv else None
    if args.facet == 'none':
        plot_scatter(
            points,
            Path(args.out),
            jitter=float(args.jitter),
            y_jitter=float(args.y_jitter),
            alpha=float(args.alpha),
            size=float(args.size),
            label_iteration=bool(args.label_iteration),
            markers_by_attribute=bool(args.markers_by_attribute),
            seed=args.seed,
            dpi=int(args.dpi),
            highlight_suspicious=bool(args.highlight_suspicious),
            suspect_tol=float(args.suspect_tol),
            suspicious_csv=suspicious_path,
        fit_line=(not bool(args.no_fit_line)),
        spaces_order=spaces_order if spaces_order else None,
        display_map=display_map if display_map else None,
    )
    else:
        # When faceting, the color denotes the opposite category
        plot_scatter_facet(
            points,
            Path(args.out),
            facet=args.facet,
            jitter=float(args.jitter),
            y_jitter=float(args.y_jitter),
            alpha=float(args.alpha),
            size=float(args.size),
            label_iteration=bool(args.label_iteration),
            seed=args.seed,
            dpi=int(args.dpi),
            ncols=int(args.ncols),
            panel_w=float(args.panel_w),
            panel_h=float(args.panel_h),
            fit_line=(not bool(args.no_fit_line)),
            spaces_order=spaces_order if spaces_order else None,
            display_map=display_map if display_map else None,
            legend_loc=str(args.legend_loc),
            no_legend=bool(args.no_legend),
        )
    print(f'Saved {args.out}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())



