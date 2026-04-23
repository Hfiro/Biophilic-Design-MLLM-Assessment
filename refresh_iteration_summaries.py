"""
Author: Hoda Alem
Date: November 2025
Context: Part of PhD research for a PhD in Environmental Design and Planning
         at Myers-Lawson School of Construction, Virginia Tech, Blacksburg, VA.

Recompute per-iteration summaries in results_*.xlsx files, in-place.

For each attribute sheet in each workbook, the tool:
- Groups true per-run rows by 'iteration'
- Recomputes:
  - rating_0_2: median and MAD (median absolute deviation)
  - score_1_10: mean and standard deviation (sample)
- Writes the two summary rows that follow each iteration block:
  - Row 1 (mean/median): rating_0_2 <- median, score_1_10 <- mean
  - Row 2 (dispersion):  rating_0_2 <- MAD,    score_1_10 <- std

This matches the structure produced by biophilic_eval.py, but replaces
the previous variance with standard deviation for 1-10.
"""

import argparse
import math
import statistics as stats
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple


def _is_number(x: Any) -> bool:
    return isinstance(x, (int, float)) and not (isinstance(x, float) and math.isnan(x))


def _headers(ws) -> Dict[str, int]:
    row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    idx: Dict[str, int] = {}
    for i, v in enumerate(row, start=1):
        if isinstance(v, str):
            idx[v.strip()] = i
    return idx


def _collect_block(ws, start_row: int, idx: Dict[str, int]) -> Tuple[int, Optional[int], List[int], List[float]]:
    """Collect a contiguous block of per-run rows starting at start_row.

    Returns (end_row, iteration, ratings, scores).
    If no run row at start_row, returns (start_row, None, [], []).
    """
    col_run = idx.get("run")
    col_it = idx.get("iteration")
    col_r02 = idx.get("rating_0_2")
    col_s10 = idx.get("score_1_10")
    max_row = ws.max_row

    if not col_run or not col_it:
        return start_row, None, [], []

    def cell(row: int, col: int) -> Any:
        return ws.cell(row=row, column=col).value

    # Must start at a run row
    first_run = cell(start_row, col_run)
    if not _is_number(first_run):
        return start_row, None, [], []

    it_val = cell(start_row, col_it)
    try:
        it_num = int(it_val) if it_val is not None else None
    except Exception:
        it_num = None

    ratings: List[int] = []
    scores: List[float] = []

    r = start_row
    while r <= max_row:
        rv = cell(r, col_run)
        if not _is_number(rv):
            break
        # Stop block if iteration changes (safety)
        iv = cell(r, col_it)
        try:
            iv_num = int(iv) if iv is not None else None
        except Exception:
            iv_num = None
        if it_num is not None and iv_num is not None and iv_num != it_num:
            break
        # Accumulate values
        if col_r02:
            v = cell(r, col_r02)
            if _is_number(v):
                try:
                    ratings.append(int(v))
                except Exception:
                    pass
        if col_s10:
            v = cell(r, col_s10)
            if _is_number(v):
                try:
                    scores.append(float(v))
                except Exception:
                    pass
        r += 1

    end_row = r - 1
    return end_row, it_num, ratings, scores


def _median(values: List[int]) -> Optional[float]:
    if not values:
        return None
    try:
        return stats.median(values)
    except Exception:
        return None


def _mad(values: List[int]) -> Optional[float]:
    if not values:
        return None
    try:
        med = stats.median(values)
        return stats.median([abs(x - med) for x in values])
    except Exception:
        return None


def _mean(values: List[float]) -> Optional[float]:
    if not values:
        return None
    try:
        return float(stats.fmean(values))
    except Exception:
        return None


def _std(values: List[float]) -> Optional[float]:
    if len(values) < 2:
        return None
    try:
        return float(stats.stdev(values))
    except Exception:
        return None


def refresh_file(path: Path) -> Tuple[int, int]:
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(path), read_only=False, data_only=True)
    updated_blocks = 0
    total_blocks = 0

    for sheet in wb.sheetnames:
        if sheet in {"meta", "prompts"}:
            continue
        ws = wb[sheet]
        idx = _headers(ws)
        if not idx:
            continue

        r = 2
        while r <= ws.max_row:
            end_row, it_num, r_list, s_list = _collect_block(ws, r, idx)
            if it_num is None:
                r += 1
                continue
            total_blocks += 1

            # Compute stats
            med_rating = _median(r_list)
            mad_rating = _mad(r_list)
            mean_score = _mean(s_list)
            std_score = _std(s_list)

            # Layout per biophilic_eval.py:
            #   end_row + 1 => spacer (blank)
            #   end_row + 2 => mean/median row (rating<-median, score<-mean)
            #   end_row + 3 => dispersion row (rating<-MAD, score<-STD)
            spacer = end_row + 1
            row_mean_median = end_row + 2
            row_dispersion = end_row + 3
            trailing_spacer = end_row + 4
            # Write only to known columns
            col_r02 = idx.get("rating_0_2")
            col_s10 = idx.get("score_1_10")
            def setval(row: int, col: Optional[int], value: Optional[float]):
                if col:
                    ws.cell(row=row, column=col, value=value)

            # Clear spacer rows fully (all columns), both before and after summaries
            def clear_row(row_idx: int):
                try:
                    for c in range(1, ws.max_column + 1):
                        # Use empty string to visibly clear in Excel UI
                        ws.cell(row=row_idx, column=c, value="")
                except Exception:
                    # Best-effort
                    pass

            clear_row(spacer)

            # Write summaries to the intended rows
            setval(row_mean_median, col_r02, med_rating)
            setval(row_mean_median, col_s10, mean_score)
            setval(row_dispersion, col_r02, mad_rating)
            setval(row_dispersion, col_s10, std_score)

            # Ensure trailing spacer row exists and is blank
            clear_row(trailing_spacer)

            updated_blocks += 1
            # Advance past spacer + two summary rows + trailing spacer
            r = trailing_spacer + 1

    if updated_blocks > 0:
        wb.save(str(path))
    wb.close()
    return total_blocks, updated_blocks


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Recompute per-iteration summaries in-place for results_*.xlsx")
    parser.add_argument("--glob", default="results_*.xlsx", help="Glob for results workbooks")
    args = parser.parse_args()

    import glob as _glob
    files = [Path(p) for p in sorted(_glob.glob(args.glob)) if p.lower().endswith('.xlsx')]
    if not files:
        print("No files matched.")
        return 0

    grand_total = 0
    grand_updated = 0
    for p in files:
        total, updated = refresh_file(p)
        grand_total += total
        grand_updated += updated
        print(f"{p.name}: iteration_blocks={total}, updated={updated}")

    print(f"All files: blocks={grand_total}, updated={grand_updated}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
